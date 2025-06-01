import requests
import pandas as pd
import threading
import base64
import os
import win32gui
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser
from tkinter.simpledialog import _setup_dialog
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from autoicon import icon as autoicon
from json import dumps, dump, load
from re import findall
from pandas import (
    Series,
    DataFrame,
    ExcelWriter,
    concat,
    notna,
    pivot_table,
    read_json,
    read_excel,
)
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from concurrent.futures import ThreadPoolExecutor, wait
from shutil import rmtree
from typing import Callable, Iterable
from traceback import format_tb, format_exc
from operator import sub, add, lt, le, gt, ge, eq, ne
from excel2img import export_img
from PIL import Image
from globals import *

VERSION = "3.4.4"
TITLE = "兴效能"
IGNORE = "ignore"
CACHE = os.getenv("LOCALAPPDATA")  # C:\\Users\\admin\\AppData\\Local
CACHE = rf"{CACHE if CACHE else '.'}\pkxee"  # 'C:\\Users\\admin\\AppData\\Local\\pkxee\\pkxee'
HISTORY = (
    CACHE + r"\history_{}.{}"
)  # 'C:\\Users\\admin\\AppData\\Local\\pkxee\\history_{}.{}'
MONI = rf"{CACHE}\moni.xlsx"  # 'C:\\Users\\admin\\AppData\\Local\\pkxee\\moni.xlsx'
LOG = rf"{CACHE}\pkxee.log"  # 'C:\\Users\\admin\\AppData\\Local\\pkxee\\pkxee.log'
SETTINGS = rf"{CACHE}\settings.json"  # 'C:\\Users\\admin\\AppData\\Local\\pkxee\\settings.json'

inside = lambda min_, value, max_: max(min_, min(max_, value))
timestamp_to_delay = lambda x: (
    f"提前{int(-x / 60)}分钟"
    if x < 0
    else f"延误{int(x / 60)}分钟" if x > 0 else "准点"
)
timedelta_to_delay = lambda x: timestamp_to_delay(x.total_seconds())
isnumeric = lambda x: str(x).count(".") <= 1 and str(x).replace(".", "").isnumeric()
true = lambda *x: True  # 接收任意数量的位置参数，并返回True
state = lambda x: tk.NORMAL if x else tk.DISABLED
state_ = lambda x: "readonly" if x else tk.DISABLED
series_any = lambda *x: np.any(x, axis=0)
series_all = lambda *x: np.all(x, axis=0)

ONOFFS = {"onvalue": True, "offvalue": False}
ALIGN_LEFT = Alignment(horizontal="left", wrapText=True)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center")
ALIGN_CENTER_LEFT = Alignment(vertical="center", horizontal="left", wrapText=True)
ALIGN_CENTER_CENTER = Alignment(vertical="center", horizontal="center", wrapText=True)
SIDE = Side("thin", "000000")
BORDER = Border(SIDE, SIDE, SIDE, SIDE)
INIT = datetime(2019, 9, 25, 0, 0)
TYPECODE = ("W/Z", "C/B", "L/W", "J/B", "B/W", "Z/P")


COLORSCALE = (
    (18, "#00128F"),
    (16, "#082693"),
    (14, "#103A97"),
    (12, "#184E9C"),
    (10, "#2062A0"),
    (8, "#2775A4"),
    (6, "#2F89A8"),
    (4, "#379DAD"),
    (2, "#3FB1B1"),
    (0, "#47C5B5"),
)
RATES_COLOR = {
    "正常性": (
        (95, "#008409"),
        (94, "#168708"),
        (93, "#2C8907"),
        (92, "#438C06"),
        (91, "#598F05"),
        (90, "#6F9104"),
        (89, "#859403"),
        (88, "#9C9702"),
        (87, "#B29901"),
        (86, "#C89C00"),
        (85, "#CD7A00"),
        (84, "#C56C02"),
        (83, "#BD5F04"),
        (82, "#B55106"),
        (81, "#AD4408"),
        (80, "#A43609"),
        (79, "#9C290B"),
        (78, "#941B0D"),
        (77, "#8C0E0F"),
        (76, "#840011"),
        (75, "#71000F"),
    ),
    "CTOT推点": COLORSCALE,
    "延误未起飞": COLORSCALE,
    "维持": "dimgrey",
    "预警": "mediumblue",
    "黄色": "orange",
    "橙色": "orangered",
    "红色": "red",
}


class GuiError(Exception): ...


class Terminates(Exception): ...


class AutoGui(tk.Tk):
    # 初始化
    def __init__(self, settings: str = SETTINGS, **kwargs) -> None:
        """
        Override default `kwargs` with `settings` file "*.json" (allow inexistent).
        Settings file will be saved upon exit, unless
        an error occurs upon loading itself or settings == False.
        Creat a new settings file if settings == True.
        """
        super().__init__()
        self.exists = tk.BooleanVar(self, True)
        self.reset = False
        if not os.path.exists(CACHE):
            os.mkdir(CACHE)
        if os.path.isfile(settings) and os.path.exists(settings):
            try:
                with open(settings) as json:
                    settings_dict = load(json)
                assert isinstance(settings_dict, dict)
                if settings_dict.get("version", "") <= VERSION:
                    kwargs.update(settings_dict)
                else:
                    self.showwarning("设置文件来自更高版本，未载入")
            except Exception:
                self.showwarning("设置加载失败，应为字典构成的JSON文件")
                settings = ""
        elif settings:
            settings = settings if isinstance(settings, str) else SETTINGS
        self.title(TITLE)

        def set_var(**kw):
            if kw or messagebox.askyesno(TITLE, "确认还原所有参数？"):
                for k in global_vars:
                    self.__setattr__(k, kw.get(k, globals().get(k)))
                for k, v in self.TAGS.items():
                    if (
                        ("时" in v or "T" in v)
                        and "标签" not in v
                        and "CONVERT" not in kw
                    ):
                        self.CONVERT.append(k)
                    v = v.split("_")
                    for i in v[:-1]:
                        if i in self.TAG_MAP:
                            self.TAG_MAP[i][k] = v[-1]
                        else:
                            self.TAG_MAP[i] = {k: v[-1]}
                if not kw:
                    update_datetime_format(False)
                    self.update_log("参数已还原为初始值")

        self.TAG_MAP = {}
        self.CONVERT = kwargs.get("CONVERT", [])
        set_var(_=True, **kwargs)
        self.LOGIN = kwargs.get("LOGIN", {})
        self.MONI_INTERVAL = tk.IntVar(self, kwargs.get("MONI_INTERVAL", 2))
        self.UPDATE_INTERVAL = tk.IntVar(self, kwargs.get("UPDATE_INTERVAL", 0))
        self.UPDATE_LIMIT = tk.IntVar(self, kwargs.get("UPDATE_LIMIT", 2000))
        self.UPDATE_OFFSET = tk.IntVar(self, kwargs.get("UPDATE_OFFSET", 60))
        self.UPDATE_RANGE = tk.IntVar(self, kwargs.get("UPDATE_RANGE", 3))
        self.UPDATE_RANGE_ = tk.IntVar(self, kwargs.get("UPDATE_RANGE_", 0))
        self.DATA_LIMIT = tk.IntVar(self, kwargs.get("DATA_LIMIT", 2000))

        for i in "AUTO", "EXPORT", "INFO", "FUNC":
            self.__setattr__(
                f"{i}_PATH", tk.StringVar(self, kwargs.get(f"{i}_PATH", "."))
            )
            self.__setattr__(
                f"{i}_PATH_", tk.BooleanVar(self, kwargs.get(f"{i}_PATH_", False))
            )
        for i in "T", "D":
            self.__setattr__(
                f"AUTO_{i}", tk.BooleanVar(self, kwargs.get(f"AUTO_{i}", False))
            )
        self.PATH = tk.StringVar(self)

        self.CONVERT_ = tk.StringVar(
            self,
            list(self.DATETIME_FORMAT)[0] if self.DATETIME_FORMAT else "标准字符串",
        )
        self.USER = tk.StringVar(self, kwargs.get("USER", "选择用户..."))
        self.SCHEDULE = tk.StringVar(self, kwargs.get("SCHEDULE", "仅自动获取数据"))
        self.TIMESLOT = tk.BooleanVar(self, kwargs.get("TIMESLOT", True))
        self.DELAY = tk.StringVar(self, kwargs.get("DELAY", "outLastTot"))
        self.INITIAL = tk.BooleanVar(self, kwargs.get("INITIAL", False))
        self.EXCLUDE = tk.BooleanVar(self, kwargs.get("EXCLUDE", True))
        self.AUTODELAY = tk.BooleanVar(self, kwargs.get("AUTODELAY", False))
        self.AUTORETRY = tk.BooleanVar(self, kwargs.get("AUTORETRY", True))

        self.SIDE = tk.BooleanVar(self, kwargs.get("SIDE", True))
        self.WRAP = tk.BooleanVar(self, kwargs.get("WARP", True))
        self.LOGLIMIT = tk.IntVar(self, kwargs.get("LOGLIMIT", 100))
        self.TOPMOST = tk.BooleanVar(self, kwargs.get("TOPMOST", False))
        self.CACHE = tk.DoubleVar(self, kwargs.get("CACHE", 3.0))
        self.AUTOSEL = tk.BooleanVar(self, True)

        self.ADJ_TYPE = ["W/Z", "C/B", "J/B", "L/W"]
        self.PRE_ADJ = {}
        self.GET_DATA = {
            "执行": self.get_flight_info,
            "航班": self.get_flight_data,
            "旅客": self.get_passenger_info,
            "延误": self.get_delay_data,
            "流控": self.get_tmi_info,
            "综合": self.get_all,
        }
        self.DATA = {}
        self.PUSH = DataFrame(columns=["信息", "图片"])
        monicol = ["航班号", "日期", "类型", "目的地", "登机门", "机位", "地服", "描述"]
        self.MONIE = DataFrame(columns=monicol).set_index(monicol[:3], drop=True)
        self.MONI = self.MONIE.copy()
        if os.path.exists(MONI):
            try:
                self.MONI = (
                    read_excel(MONI)
                    .reindex(columns=monicol)
                    .set_index(monicol[:3], drop=True)
                )
            except Exception:
                pass

        self.NOW, self.NOW_ = tk.StringVar(self), tk.BooleanVar(self)
        # self.NOW, self.NOW_ = tk.StringVar(self, "2024-02-18 09:00"), tk.BooleanVar(self, True)

        self.RUNNING = {}
        self.msg_para = [tk.BooleanVar(self) for _ in range(3)]
        self.yesterday = tk.BooleanVar(self)
        self.all_schedule = (
            (self.auto_msg, self.PATH),
            (self.stock_msg, self.PATH),
            (self.get_ctot, 0, self.PATH),
            (self.long_delay, 1, self.PATH),
        )
        self.SCHEDULES = {
            "仅自动获取数据": False,
            "自动更新信息": {0: self.all_schedule},
            "自动同步信息": True,
            "每一小时生成信息": {60: self.all_schedule},
            "每半小时生成信息": {30: self.all_schedule},
            "每10分钟生成信息": {10: self.all_schedule},
        }
        if self.SCHEDULE.get() not in self.SCHEDULES:
            self.SCHEDULE.set(list(self.SCHEDULES.keys())[0])
        self.func_names = {
            "auto_msg": "效率席短信",
            "stock_msg": "当前运行概述",
            "get_ctot": "CTOT推点",
            "long_delay": "延误未起飞",
        }
        self.lock = threading.RLock()

        self.TIMEDELTA = {
            "outLastTotDelay",
            "outStotDelay",
            "attt",
            "sttt",
            "ttt",
            "mttt",
            "outMttt",
        }
        for k in TAGS.keys():
            self.__setattr__(k.replace(".", "_"), k)
        self.progress = tk.IntVar(self)
        self.login_timestamp = self.session = 0
        self.update_timestamp = tk.DoubleVar(self)

        # styles
        style = ttk.Style()
        style.configure("ChangeDt.TButton", font=("微软雅黑", 8), width=4)
        style.configure("ChangeDtm.TButton", font=("微软雅黑", 8), width=5)
        style.configure("Datetime.TButton", font=("Arial", 10, "bold"), width=20)
        style.configure("Editor.TButton", font=("微软雅黑", 8), width=8)
        style.configure("EditorIcon.TButton", font=("微软雅黑", 9), width=3)
        style.configure("Launch.TButton", font=("微软雅黑", 14, "bold"), width=5)
        style.configure("Main.TButton", font=("微软雅黑", 12), width=25)
        style.configure("Settings.TButton", font=("微软雅黑", 8), width=16)
        style.configure("White.TFrame", background="white")

        # functions
        func = {
            "更新所有信息": lambda: threading.Thread(
                target=self.update_info,
            ).start(),
            "综合效率席短信": lambda: threading.Thread(
                target=self.auto_msg,
            ).start(),
            "当前运行概述": lambda: threading.Thread(
                target=self.stock_msg,
            ).start(),
            "CTOT推点航班明细": lambda: threading.Thread(
                target=self.get_ctot,
            ).start(),
            "延误1h未起飞航班明细": lambda: threading.Thread(
                target=self.long_delay, args=(1,)
            ).start(),
            "延误未起飞航班明细": lambda: threading.Thread(
                target=self.long_delay,
            ).start(),
            "航司机场查询...": self.search,
            "企微信息推送...": self.push,
            "CTOT历史查询...": self.history,
            "自动化调时...": self.adj,
            "导出今日生效流控表格": lambda: self.submit_export(
                datetime.now(),
                self.today(1, minutes=-1),
                "导出今日生效流控表格",
                "流控",
            ),
            "导出今日计划离港表格": lambda: self.submit_export(
                self.today(0),
                self.today(1, minutes=-1),
                "导出今日计划离港表格",
                "航班",
                departureMode="outS",
            ),
            "导出今日计划进港表格": lambda: self.submit_export(
                self.today(0),
                self.today(1, minutes=-1),
                "导出今日计划进港表格",
                "航班",
                departureMode="inS",
            ),
            "导出昨日延误表格": lambda: self.submit_export(
                self.today(-1), self.today(0, minutes=-1), "导出昨日延误表格", "延误"
            ),
            "导出今日延误表格": lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日延误表格", "延误"
            ),
            "导出昨日执行概览表格": lambda: self.submit_export(
                self.today(-1),
                self.today(0, minutes=-1),
                "导出昨日执行概览表格",
                "执行",
            ),
            "导出今日执行概览表格": lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日执行概览表格", "执行"
            ),
            "导出明日执行概览表格": lambda: self.submit_export(
                self.today(1), self.today(2, minutes=-1), "导出明日执行概览表格", "执行"
            ),
        }

        # menus
        menu = tk.Menu(self, name="menu")
        self.config(menu=menu)

        # schedule
        def schedule():
            schedule = self.SCHEDULE.get()
            schedules = self.SCHEDULES.get(schedule)
            if schedules:
                if not self.AUTO_PATH_.get():
                    self.AUTO_PATH_.set(True)
                    update_path("AUTO")
                if not self.AUTO_PATH_.get():
                    raise Terminates("请选择导出路径以开始任务计划")
                menu_schedule[8].entryconfig(
                    0, offvalue=True, command=lambda: update_path("AUTO", True, False)
                )
                self.RUNNING["sync" if "同步" in schedule else "get_info"] = dict()

            self.RUNNING["get_data"] = {0: ""}
            self.get_data()
            menu_export[0].entryconfig(10, state=tk.NORMAL)

            if self.MONI_INTERVAL.get():
                self.RUNNING["get_monitor"] = {0: ""}
                threading.Thread(target=self.get_monitor).start()
            for interval, args in (
                schedules.items() if isinstance(schedules, dict) else []
            ):
                if not interval:
                    self.get_info(*args, interval=interval)
                else:
                    self.RUNNING["get_info"][interval] = self.after(
                        self.get_next_update(interval, self.UPDATE_OFFSET.get() - 10),
                        threading.Thread(
                            target=self.get_info,
                            args=args,
                            kwargs={"interval": interval},
                        ).start,
                    )

        def refresh(enabled: bool = True):
            enabled = state(enabled)
            launch_button.config(state=enabled)
            refresh_button.config(state=enabled)
            menu_schedule[0].entryconfig(0, state=enabled)
            menu_schedule[0].entryconfig(1, state=enabled)

        self.refresh = refresh
        self.retry = self.username = ""

        def launch():
            if s_ := ss_.get():
                ss_.set(False)
                if ss.get():
                    self.after_cancel(ss.get())
            if self.retry:
                self.after_cancel(self.retry)
            if self.SCHEDULE.get() not in self.SCHEDULES:
                raise GuiError("请选择任务计划")
            try:
                self.get_session()
                menu_schedule[0].entryconfig(3, state=tk.DISABLED)
                menu_schedule[0].entryconfig(0, label="停止", command=terminate)
                launch_button.config(text="停止", command=terminate)
                status = "自动化{}，用户：{}".format(
                    "计划自启" if s_ else "重启" if self.retry else "启动",
                    self.username if self.username else self.USER.get(),
                )
                self.retry = ""
                if self.NOW_.get():
                    status += "，更改当前时间功能已禁用，当前时间将与系统保持一致"
                    self.showinfo(status)
                    self.NOW_.set(False), set_datetime()
                menu_setting[0].entryconfig(0, state=tk.DISABLED)
                menu_setting[0].entryconfig(2, state=tk.DISABLED)
                menu_setting[3].entryconfig(0, state=tk.DISABLED)
                user_combobox.config(state=tk.DISABLED)
                user_button.config(state=tk.DISABLED)
                schedule_combobox.config(state=tk.DISABLED)
                default_rates()
                frames[1].pack(padx=0, pady=0, fill="x", before=frames[2])
                refresh(True)
                self.status_bar.config(background="palegreen")
                threading.Thread(target=schedule).start()
                self.update_log(status, "schedule", end="\n\n")
            except Exception as exception:
                terminate(exception)

        def terminate(*args):
            if se_.get():
                se_.set(False)
                if se.get():
                    self.after_cancel(se.get())
            menu_schedule[0].entryconfig(3, state=tk.NORMAL)
            menu_schedule[0].entryconfig(0, label="开始", command=launch)
            menu_schedule[8].entryconfig(
                0, offvalue=False, command=lambda: update_path("AUTO")
            )
            launch_button.config(text="开始", command=launch)
            menu_setting[0].entryconfig(0, state=tk.NORMAL)
            menu_setting[0].entryconfig(2, state=tk.NORMAL)
            menu_setting[3].entryconfig(0, state=tk.NORMAL)
            user_combobox.config(state="readonly")
            user_button.config(state=tk.NORMAL)
            schedule_combobox.config(state="readonly")
            frames[1].forget()

            for i in list(self.RUNNING.keys()):
                for i in self.RUNNING.pop(i).values():
                    if i:
                        self.after_cancel(i)
            if args:
                t = repr(args[0]).replace("<class ", "").replace("'", "")
                msg = f"自动化终止 ({t[: t.find('(')]})"
                if self.AUTORETRY.get():
                    t = 5 if datetime.now().hour <= 5 else 1
                    self.retry = self.after(t * 60 * 1000, launch)
                    msg += f"，{t}分钟后尝试重启自动化"
                self.update_log(msg, "warn", "schedule", end="\n\n")
            else:
                self.update_log("自动化停止", "schedule", end="\n\n")
                self.login_timestamp = 0
                self.username = ""
            self.update_status()
            refresh()
            self.title(f"{TITLE} - 自动化停止")
            threading.Thread(target=self.save_history).start()

        self.terminate = terminate

        menu_export = [tk.Menu(menu, name="menu_export", tearoff=False)]
        menu_others = tk.Menu(menu, name="menu_others", tearoff=False)
        menu_info = [tk.Menu(menu, name="menu_info", tearoff=False)]
        menu_schedule = [
            tk.Menu(menu, name="menu_schedule", tearoff=False),
        ]
        for _ in range(10):
            menu_schedule.append(tk.Menu(menu, tearoff=False))

        path_config = {
            "AUTO": (self.AUTO_PATH, self.AUTO_PATH_, menu_schedule[8], 8),
            "EXPORT": (self.EXPORT_PATH, self.EXPORT_PATH_, menu_export[0], 8),
            "INFO": (self.INFO_PATH, self.INFO_PATH_, menu_info[0], 8),
            "FUNC": (self.FUNC_PATH, self.FUNC_PATH_, menu_others, 12),
        }

        def split_path(__str: str, __max: int) -> str:
            s = ""
            for __str in reversed(__str.rsplit("/")):
                if len(s) + len(__str) <= __max:
                    s = rf"\{__str}" + s
                else:
                    break
            return s

        def update_path(
            config: str,
            ask: bool = True,
            cancelable: bool = True,
            index: int = 0,
        ):
            path, path_, menu, max_path = path_config[config]
            if path_.get():
                s = (
                    filedialog.askdirectory(
                        initialdir=path.get(),
                        mustexist=True,
                        parent=self,
                        title="选择导出路径...",
                    )
                    if ask
                    else path.get()
                )
                if s:
                    path.set(s)
                    s = "导出至" + (
                        "..." + split_path(s, max_path)
                        if "/" in s and len(s) > max_path
                        else s
                    )
                    menu.entryconfig(index, label=s)
                elif cancelable:
                    path_.set(False)
            if not path_.get():
                menu.entryconfig(index, label="设置导出路径...")

        self.update_path = update_path

        menu.add_cascade(label="自动化", menu=menu_schedule[0])
        menu_schedule[0].add_command(label="开始", command=launch)
        menu_schedule[0].add_command(
            label="立刻更新",
            command=lambda: threading.Thread(target=self.get_data).start(),
        )
        menu_schedule[0].add_separator()
        menu_schedule[0].add_cascade(label="选择任务计划", menu=menu_schedule[6])
        for i in self.SCHEDULES.keys():
            menu_schedule[6].add_radiobutton(variable=self.SCHEDULE, value=i, label=i)

        menu_schedule[0].add_cascade(label="自动生成文件", menu=menu_schedule[8])

        menu_schedule[8].add_checkbutton(
            label="设置根路径...",
            variable=self.AUTO_PATH_,
            command=lambda: update_path("AUTO"),
            **ONOFFS,
        )
        update_path("AUTO", False)
        menu_schedule[8].add_separator()
        menu_schedule[8].add_checkbutton(
            label="按信息类型分文件夹", variable=self.AUTO_T, **ONOFFS
        )
        menu_schedule[8].add_checkbutton(
            label="按日期分文件夹", variable=self.AUTO_D, **ONOFFS
        )

        menu_schedule[0].add_cascade(label="自动延误判定", menu=menu_schedule[9])
        menu_schedule[9].add_radiobutton(
            label="开启",
            value=True,
            variable=self.AUTODELAY,
            command=self.update_status,
        )
        menu_schedule[9].add_radiobutton(
            label="关闭",
            value=False,
            variable=self.AUTODELAY,
            command=self.update_status,
        )
        menu_schedule[9].add_separator()
        menu_schedule[9].add_command(label="判定策略...", command=self.auto_delay)
        # menu_schedule[9].add_command(label="预先判定...", command=self.set_delay)

        menu_schedule[0].add_checkbutton(
            label="中断自动重启", variable=self.AUTORETRY, **ONOFFS
        )

        menu_schedule[0].add_separator()
        menu_schedule[0].add_cascade(label="告警检测间隔", menu=menu_schedule[10])
        for i in 1, 2, 5:
            menu_schedule[10].add_radiobutton(
                variable=self.MONI_INTERVAL, value=i, label=f"{i}分钟"
            )

        menu_schedule[0].add_cascade(label="数据更新间隔", menu=menu_schedule[1])
        menu_schedule[1].add_radiobutton(
            variable=self.UPDATE_INTERVAL, value=0, label="自动"
        )
        menu_schedule[1].add_separator()
        for i in 2, 5, 10, 15:
            menu_schedule[1].add_radiobutton(
                variable=self.UPDATE_INTERVAL, value=i, label=f"{i}分钟"
            )

        menu_schedule[0].add_cascade(label="更新提前量", menu=menu_schedule[7])
        for i in 30, 60, 90, 120:
            menu_schedule[7].add_radiobutton(
                variable=self.UPDATE_OFFSET, value=i, label=self.min_sec_format(i)
            )

        def update_range():
            for i in range(2):
                menu_schedule[2].entryconfig(
                    i + 2, state=state(self.UPDATE_RANGE_.get())
                )

        menu_schedule[0].add_cascade(label="数据更新范围", menu=menu_schedule[2])
        menu_schedule[2].add_checkbutton(
            variable=self.UPDATE_RANGE_,
            onvalue=0,
            offvalue=3,
            label="自动",
            command=update_range,
        )
        menu_schedule[2].add_separator()
        menu_schedule[2].add_cascade(
            label="自今日0时",
            menu=menu_schedule[3],
            state=state(self.UPDATE_RANGE_.get()),
        )
        for i in 0, 3, 6, 12, 24:
            menu_schedule[3].add_radiobutton(
                variable=self.UPDATE_RANGE, value=i, label=f"{i}小时前"
            )

        menu_schedule[2].add_cascade(
            label="至当前", menu=menu_schedule[4], state=state(self.UPDATE_RANGE_.get())
        )
        for i in 6, 5, 4, 3, 2:
            menu_schedule[4].add_radiobutton(
                variable=self.UPDATE_RANGE_, value=i, label=f"{i}小时后"
            )

        menu_schedule[0].add_cascade(label="数据量限制", menu=menu_schedule[5])
        for i in 500, 1000, 2000, 5000:
            menu_schedule[5].add_radiobutton(
                variable=self.UPDATE_LIMIT, value=i, label=f"{i}条"
            )

        self.delay_type = [
            ["0112:其它天气原因", "本场"],
            ["0404:其它军事活动原因", "本场"],
            ["0112:其它天气原因", "外站"],
            ["0404:其它军事活动原因", "外站"],
            ["0201:公司计划", "	其他（与外部原因一致）"],
            # ['0202:运行保障', ''],
            # ['0203:空勤组', ''],
            # ['0204:工程机务', ''],
            # ['0205:公司销售', ''],
            # ['0206:地面服务', ''],
            # ['0207:食品供应', ''],
            # ['0208:货物运输', ''],
            # ['0209:后勤保障', ''],
            # ['0210:代理机构', ''],
            # ['0211:其它航空公司原因', ''],
            # ['1001:等待旅客', ''],
            # ['1002:登机手续不符合规定', ''],
            # ['1003:旅客突发疾病', ''],
            # ['1004:旅客丢失登机牌，重新办理乘机手续', ''],
            # ['1005:旅客登机后要求下机，重新进行客舱及行李舱安全检查', ''],
            # ['1006:旅客拒绝登机、霸占飞机', ''],
            # ['1007:其它旅客原因', ''],
            # ['1101:因举办大型活动或发生突发事件，造成保障能力下降或安检时间延长', ''],
            # ['1102:航班遭到劫持、爆炸威胁', ''],
            # ['1103:发生可能影响飞行安全的事件(如机场周边燃放烟花导致能见度下降，发现不明飞行物、气球、风筝，地震、海啸等自然灾害)', ''],
            # ['1104:公共卫生事件', ''],
            # ['1105:其它公共安全原因', ''],
        ]

        # export
        menu_index = 1
        for _ in range(8):
            menu_export.append(tk.Menu(menu, tearoff=False))
        menu.add_cascade(label="导出数据", menu=menu_export[0])
        menu_export[0].add_checkbutton(
            label="设置导出路径...",
            variable=self.EXPORT_PATH_,
            command=lambda: update_path("EXPORT"),
            **ONOFFS,
        )
        update_path("EXPORT", False)
        menu_export[0].add_cascade(label="导出数据限制", menu=menu_export[menu_index])
        for i in 500, 1000, 2000, 5000:
            menu_export[1].add_radiobutton(
                variable=self.DATA_LIMIT, value=i, label=f"{i}条"
            )

        menu_index += 1
        menu_export[0].add_cascade(label="日期时间格式", menu=menu_export[menu_index])
        menu_export[menu_index].add_radiobutton(
            variable=self.CONVERT_, value="标准字符串", label="标准字符串"
        )
        menu_export[menu_index].add_separator()
        for i in self.DATETIME_FORMAT.keys():
            menu_export[menu_index].add_radiobutton(
                variable=self.CONVERT_, value=i, label=i
            )

        menu_index += 1
        menu_export[0].add_separator()
        menu_export[0].add_cascade(label="航班详情表格", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="按计划离港选择...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "航班",
                "导出航班详情表格 - 计划离港",
                departureMode="outS",
            ),
        )
        menu_export[menu_index].add_command(
            label="按计划进港选择...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "航班",
                "导出航班详情表格 - 计划进港",
                departureMode="inS",
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="按实际离港选择...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "航班",
                "导出航班详情表格 - 实际离港",
                departureMode="outA",
            ),
        )
        menu_export[menu_index].add_command(
            label="按实际进港选择...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "航班",
                "导出航班详情表格 - 实际进港",
                departureMode="inA",
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="今日计划离港航班",
            command=lambda: self.submit_export(
                self.today(0),
                self.today(1, minutes=-1),
                "导出今日计划离港表格",
                "航班",
                departureMode="outS",
            ),
        )
        menu_export[menu_index].add_command(
            label="今日计划进港航班",
            command=lambda: self.submit_export(
                self.today(0),
                self.today(1, minutes=-1),
                "导出今日计划进港表格",
                "航班",
                departureMode="inS",
            ),
        )

        menu_index += 1
        menu_export[0].add_cascade(label="延误详情表格", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="自定义时段...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "延误",
                "导出延误航班表格 - 自定义时段",
                "导出时段内计划离港延误航班",
                (INIT, self.datetime_now),
                (INIT, self.today(1)),
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="昨日延误航班",
            command=lambda: self.submit_export(
                self.today(-1), self.today(0, minutes=-1), "导出昨日延误表格", "延误"
            ),
        )
        menu_export[menu_index].add_command(
            label="今日延误航班",
            command=lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日延误表格", "延误"
            ),
        )

        menu_index += 1
        menu_export[0].add_cascade(label="流控信息表格", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="自定义时段...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "流控",
                "导出流控信息表格 - 自定义时段",
                "导出时段内生效的流控信息",
                (INIT, self.today(7)),
                (INIT, self.today(7)),
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="今日生效流控信息",
            command=lambda: self.submit_export(
                self.datetime_now(),
                self.today(1, minutes=-1),
                "导出今日生效流控表格",
                "流控",
            ),
        )
        menu_export[menu_index].add_command(
            label="昨日流控信息",
            command=lambda: self.submit_export(
                self.today(-1), self.today(0, minutes=-1), "导出昨日流控表格", "流控"
            ),
        )
        menu_export[menu_index].add_command(
            label="今日流控信息",
            command=lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日流控表格", "流控"
            ),
        )
        menu_export[menu_index].add_command(
            label="明日流控信息",
            command=lambda: self.submit_export(
                self.today(1), self.today(2, minutes=-1), "导出明日流控表格", "流控"
            ),
        )

        menu_index += 1
        menu_export[0].add_cascade(label="执行概览表格", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="自定义时段...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "执行",
                "导出执行概览表格 - 自定义时段",
                "导出时段内进离港航班执行概览",
                (INIT, self.today(7)),
                (INIT, self.today(7)),
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="昨日执行概览",
            command=lambda: self.submit_export(
                self.today(-1),
                self.today(0, minutes=-1),
                "导出昨日执行概览表格",
                "执行",
            ),
        )
        menu_export[menu_index].add_command(
            label="今日执行概览",
            command=lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日执行概览表格", "执行"
            ),
        )
        menu_export[menu_index].add_command(
            label="明日执行概览",
            command=lambda: self.submit_export(
                self.today(1), self.today(2, minutes=-1), "导出明日执行概览表格", "执行"
            ),
        )

        menu_index += 1
        menu_export[0].add_separator()
        menu_export[0].add_cascade(label="上述所有表格", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="自定义时段...",
            command=lambda: self.ask_export(
                self.today(),
                self.today(1),
                "综合",
                "导出综合信息表格",
                "导出应用提供的所有表格",
            ),
        )
        menu_export[menu_index].add_separator()
        menu_export[menu_index].add_command(
            label="昨日所有表格",
            command=lambda: self.submit_export(
                self.today(-1),
                self.today(0, minutes=-1),
                "导出昨日综合信息表格",
                "综合",
            ),
        )
        menu_export[menu_index].add_command(
            label="今日所有表格",
            command=lambda: self.submit_export(
                self.today(0), self.today(1, minutes=-1), "导出今日综合信息表格", "综合"
            ),
        )
        menu_export[menu_index].add_command(
            label="明日所有表格",
            command=lambda: self.submit_export(
                self.today(1), self.today(2, minutes=-1), "导出明日综合信息表格", "综合"
            ),
        )

        def get_history_rates():
            if self.HISTORY["RATES"].empty:
                self.showinfo("暂无效率指标历史记录")
                return 1
            title, filename = "导出效率指标历史表格", self.FILENAME.get("效率指标")
            file = (
                f"{self.EXPORT_PATH.get()}/{filename}"
                if self.EXPORT_PATH_.get()
                else filedialog.asksaveasfilename(
                    filetypes=(("Xlsx表格文件", "*.xlsx"),),
                    confirmoverwrite=True,
                    parent=self,
                    title=title,
                    initialdir=self.EXPORT_PATH.get(),
                    initialfile=filename,
                )
            )
            threading.Thread(
                target=self.save_excel,
                args=(self, title, file, self.get_history_rates),
                kwargs={"data": self.HISTORY["RATES"].copy()},
            ).start()
            return 0

        menu_index += 1
        menu_export[0].add_cascade(label="历史记录", menu=menu_export[menu_index])
        menu_export[menu_index].add_command(
            label="CTOT和COBT...",
            command=lambda: self.ask_export(
                self.today(),
                self.datetime_now(),
                self.get_history_cal,
                "选择导出时段",
                "导出STD在时段内的离港航班CTOT和COBT历史",
                (INIT, lambda: datetime.now() + timedelta(hours=6)),
                (INIT, lambda: datetime.now() + timedelta(hours=6)),
            ),
        )
        menu_export[menu_index].add_command(label="效率指标", command=get_history_rates)
        menu_export[0].add_command(label="数据透视表格...", command=self.pivot_export)

        # info
        def ask_color(color: str, parent: tk.Misc):
            if color_ := colorchooser.askcolor(
                color=f"#{color}",
                initialcolor=f"#{color}",
                title="选择颜色",
                parent=parent,
            )[1]:
                color = color_[1:]
            return color

        def update_config(__item: str, title: str):
            item = self.__getattribute__(f"{__item}CONFIG")
            config = tk.Toplevel(self, name="config")
            config.attributes("-topmost", self.TOPMOST.get())
            config.title(title)
            config.resizable(False, False)
            config.bind("<Escape>", lambda *x: config.destroy())
            config.geometry(
                f"+{self.winfo_rootx() + self.winfo_width() // 10}+{self.winfo_rooty() + self.winfo_height() // 10}"
            )
            var = dict()
            if __item == "CTOT":
                rules = {
                    "B": "标注受流控影响航班号为粗体",
                    "D": "标注预计延误到达航班（未落/未起）",
                    "J": "标注推点时间色阶（0分钟/大于80%）",
                    "K": "标注过站时间不足航班",
                    "M": "标注兴快线和已登机航班",
                    "N": "标注CTOT推迟的航班",
                }
                set_color = {
                    ("B", 2): lambda: var["B"][2].set(
                        ask_color(var["B"][2].get(), config)
                    ),
                    ("D", 2): lambda: var["D"][2].set(
                        ask_color(var["D"][2].get(), config)
                    ),
                    ("D", 3): lambda: var["D"][3].set(
                        ask_color(var["D"][3].get(), config)
                    ),
                    ("J", 2): lambda: var["J"][2].set(
                        ask_color(var["J"][2].get(), config)
                    ),
                    ("J", 3): lambda: var["J"][3].set(
                        ask_color(var["J"][3].get(), config)
                    ),
                    ("K", 2): lambda: var["K"][2].set(
                        ask_color(var["K"][2].get(), config)
                    ),
                    ("M", 2): lambda: var["M"][2].set(
                        ask_color(var["M"][2].get(), config)
                    ),
                    ("N", 2): lambda: var["N"][2].set(
                        ask_color(var["N"][2].get(), config)
                    ),
                    ("延误", 1): lambda: var["延误"][1].set(
                        ask_color(var["延误"][1].get(), config)
                    ),
                }
            else:
                rules = {
                    "J": "标注延误时间色阶（最小/最大）",
                    "K": "标注距CTOT色阶（最小/最大）",
                    "L": "标注机上等待色阶（最小/最大）",
                    "M": "标注已登机航班状态为粗体",
                }
                set_color = {
                    ("J", 2): lambda: var["J"][2].set(
                        ask_color(var["J"][2].get(), config)
                    ),
                    ("J", 3): lambda: var["J"][3].set(
                        ask_color(var["J"][3].get(), config)
                    ),
                    ("K", 2): lambda: var["K"][2].set(
                        ask_color(var["K"][2].get(), config)
                    ),
                    ("K", 3): lambda: var["K"][3].set(
                        ask_color(var["K"][3].get(), config)
                    ),
                    ("L", 2): lambda: var["L"][2].set(
                        ask_color(var["L"][2].get(), config)
                    ),
                    ("L", 3): lambda: var["L"][3].set(
                        ask_color(var["L"][3].get(), config)
                    ),
                    ("M", 2): lambda: var["M"][2].set(
                        ask_color(var["M"][2].get(), config)
                    ),
                }

            for row, (k, v) in enumerate(item.items()):
                var[k] = [tk.BooleanVar(config, v[1])]
                if k.isupper():
                    ttk.Label(config, text=f"{v[0]}（{k}列）").grid(
                        row=row, column=0, padx=0, pady=1, sticky="e"
                    )
                    ttk.Checkbutton(
                        config, text="隐藏", variable=var[k][0], **ONOFFS
                    ).grid(row=row, column=1, padx=2, pady=1)
                    if rule := rules.get(k):
                        var[k].append(tk.BooleanVar(config, v[2]))
                        ttk.Checkbutton(
                            config, text=rule, variable=var[k][1], **ONOFFS
                        ).grid(row=row, column=2, padx=2, pady=1, sticky="w")
                        for i, v in enumerate(v[3:]):
                            var[k].append(tk.StringVar(config, v))
                            ttk.Button(
                                config,
                                text="调整颜色",
                                width=8,
                                command=set_color[(k, i + 2)],
                            ).grid(row=row, column=i + 3, padx=2, pady=1)

                else:
                    ttk.Checkbutton(
                        config, text=v[0], variable=var[k][0], **ONOFFS
                    ).grid(row=row, column=2, padx=2, pady=1, sticky="w")
                    for i, v in enumerate(v[2:]):
                        var[k].append(tk.StringVar(config, v))
                        ttk.Button(
                            config,
                            text="调整颜色",
                            width=8,
                            command=set_color[(k, i + 1)],
                        ).grid(row=row, column=3, padx=2, pady=1)

            def reset():
                if messagebox.askyesno(
                    TITLE, "确定恢复默认表格配置？此操作不能撤销", parent=config
                ):
                    item.update(globals().get(f"{__item}CONFIG"))
                    update_config(__item, title)

            def confirm():
                for k, v in var.items():
                    for i, v in enumerate(v):
                        item[k][i + 1] = v.get()
                config.destroy()

            row += 1
            ttk.Button(config, text="确定", command=confirm, width=10).grid(
                row=row, column=0, columnspan=5, sticky="ws", padx=2, pady=5
            )
            ttk.Button(config, text="恢复默认", command=reset, width=10).grid(
                row=row, column=0, columnspan=5, sticky="s", padx=2, pady=5
            )
            ttk.Button(config, text="取消", command=config.destroy, width=10).grid(
                row=row, column=0, columnspan=5, sticky="es", padx=2, pady=5
            )
            config.focus_set()
            config.wait_window()

        for _ in range(3):
            menu_info.append(tk.Menu(menu_info[0], tearoff=False))
        menu.add_cascade(label="导出信息", menu=menu_info[0])
        menu_info[0].add_checkbutton(
            label="设置导出路径...",
            variable=self.INFO_PATH_,
            command=lambda: update_path("INFO"),
            **ONOFFS,
        )
        update_path("INFO", False)

        menu_info[0].add_command(label="企微信息推送...", command=self.push)
        menu_info[0].add_cascade(label="信息设置", menu=menu_info[1])
        menu_info[1].add_checkbutton(label="吸附时间", variable=self.TIMESLOT, **ONOFFS)
        menu_info[1].add_checkbutton(
            label="昨日续报",
            variable=self.yesterday,
            command=lambda: (
                (self.update_timestamp.set(0), self.update_status())
                if self.yesterday.get()
                else self.update_status()
            ),
            **ONOFFS,
        )
        menu_info[1].add_separator()
        menu_info[1].add_checkbutton(
            label="自动选择信息与状态", variable=self.AUTOSEL, **ONOFFS
        )
        menu_info[1].add_checkbutton(
            label="大面积航延", variable=self.msg_para[2], **ONOFFS
        )
        menu_info[1].add_checkbutton(
            label="霜天气",
            variable=self.msg_para[0],
            command=lambda: (
                self.msg_para[1].set(False) if self.msg_para[1].get() else None
            ),
            **ONOFFS,
        )
        menu_info[1].add_checkbutton(
            label="冰雪天气",
            variable=self.msg_para[1],
            command=lambda: (
                self.msg_para[0].set(False) if self.msg_para[0].get() else None
            ),
            **ONOFFS,
        )

        menu_info[1].add_separator()
        menu_info[1].add_radiobutton(
            label="表格按放行延误",
            variable=self.DELAY,
            value="outLastTot",
            command=self.update_status,
        )
        menu_info[1].add_radiobutton(
            label="表格按起飞延误",
            variable=self.DELAY,
            value="outStot",
            command=self.update_status,
        )

        menu_info[1].add_separator()
        menu_info[1].add_radiobutton(
            label="按A-CDM始发标记",
            variable=self.INITIAL,
            value=True,
        )
        menu_info[1].add_radiobutton(
            label="按计算始发标记",
            variable=self.INITIAL,
            value=False,
        )

        def update_ctot_type():
            self.CTOTTYPE = self.ask_type("配置CTOT表格航班性质", self, self.CTOTTYPE)
            self.update_status()

        menu_info[1].add_separator()
        menu_info[1].add_command(label="CTOT航班性质...", command=update_ctot_type)
        menu_info[1].add_command(
            label="CTOT表格定制...",
            command=lambda: update_config("CTOT", "CTOT表格定制"),
        )
        menu_info[1].add_separator()
        menu_info[1].add_checkbutton(
            label="延误未起排除未落地", variable=self.EXCLUDE, **ONOFFS
        )
        menu_info[1].add_command(
            label="延误未起表格定制...",
            command=lambda: update_config("DELAY", "延误未起表格定制"),
        )
        menu_info[0].add_separator()
        menu_info[0].add_command(
            label="综合效率席短信",
            command=lambda: threading.Thread(
                target=self.auto_msg,
            ).start(),
        )
        menu_info[0].add_command(
            label="当前运行概述",
            command=lambda: threading.Thread(
                target=self.stock_msg,
            ).start(),
        )
        menu_info[0].add_cascade(label="CTOT推点航班明细", menu=menu_info[3])
        menu_info[3].add_command(
            label="推点0分钟以上",
            command=lambda: threading.Thread(target=self.get_ctot).start(),
        )
        menu_info[3].add_separator()
        menu_info[3].add_command(
            label="推点-4分钟以上",
            command=lambda: threading.Thread(target=self.get_ctot, args=(-4,)).start(),
        )
        menu_info[3].add_separator()
        menu_info[3].add_command(
            label="推点3分钟以上",
            command=lambda: threading.Thread(target=self.get_ctot, args=(3,)).start(),
        )
        menu_info[3].add_command(
            label="推点5分钟以上",
            command=lambda: threading.Thread(target=self.get_ctot, args=(5,)).start(),
        )
        menu_info[3].add_command(
            label="推点1小时以上",
            command=lambda: threading.Thread(target=self.get_ctot, args=(60,)).start(),
        )
        menu_info[0].add_cascade(label="延误未起飞航班明细", menu=menu_info[2])
        menu_info[2].add_command(
            label="所有延误未起飞",
            command=lambda: threading.Thread(target=self.long_delay).start(),
        )
        menu_info[2].add_separator()
        menu_info[2].add_command(
            label="延误0.5小时以上",
            command=lambda: threading.Thread(
                target=self.long_delay, args=(0.5,)
            ).start(),
        )
        menu_info[2].add_command(
            label="延误1小时以上",
            command=lambda: threading.Thread(target=self.long_delay, args=(1,)).start(),
        )
        menu_info[2].add_command(
            label="延误1.5小时以上",
            command=lambda: threading.Thread(
                target=self.long_delay, args=(1.5,)
            ).start(),
        )
        menu_info[2].add_command(
            label="延误2小时以上",
            command=lambda: threading.Thread(target=self.long_delay, args=(2,)).start(),
        )
        menu_info[2].add_command(
            label="延误3小时以上",
            command=lambda: threading.Thread(target=self.long_delay, args=(3,)).start(),
        )
        menu_info[2].add_command(
            label="延误6小时以上",
            command=lambda: threading.Thread(target=self.long_delay, args=(6,)).start(),
        )

        # others
        def update_login():
            self.attribute_editor(
                "LOGIN",
                "用户信息编辑",
                "双击编辑，密码为加密后字段",
                key=lambda x: len(x),
                value=lambda x: len(x),
                headers=("账号", "密码"),
                width=(110, 300),
            )
            if not self.RUNNING and self.exists.get():
                users = list(self.LOGIN.keys())
                user_combobox.config(values=users)
                if self.USER.get() not in users:
                    self.login_timestamp = 0
                    self.username = ""
                    if len(users):
                        self.USER.set(users[0])
                    else:
                        self.USER.set("选择用户...")

        menu.add_cascade(label="其他功能", menu=menu_others)
        menu_others.add_checkbutton(
            label="设置导出路径...",
            variable=self.FUNC_PATH_,
            command=lambda: update_path("FUNC"),
            **ONOFFS,
        )
        update_path("FUNC", False)
        menu_others.add_separator()
        menu_others.add_command(label="航司机场查询...", command=self.search)
        menu_others.add_command(label="自动化调时...", command=self.adj)
        menu_others.add_command(label="CTOT和COBT历史查询...", command=self.history)

        # setting
        def reset():
            if messagebox.askyesno(
                TITLE, "确认重置程序？所有设置、历史记录和日志将被清除，程序将退出"
            ):
                try:
                    rmtree(CACHE)
                except Exception:
                    tb = format_exc()
                    messagebox.showerror(
                        f"{TITLE} - 重置失败",
                        "无法删除文件夹{}\n{}".format(
                            CACHE, tb[tb[:-1].rfind("\n") + 1 : -1]
                        ),
                    )
                self.reset = True
                self.destroy()
                self.quit()

        def about():
            about = tk.Toplevel(self, name="about")
            about.attributes("-topmost", self.TOPMOST.get())
            about.title("关于" + TITLE + " " + VERSION)
            self.author = "Primarie@qq.com"
            var = list(self.__dir__())
            text = tk.StringVar(about)

            def new_var(*arg):
                i = np.random.choice(var)
                j = self.__getattribute__(i)
                j = repr(j.get() if isinstance(j, tk.Variable) else j)
                if len(j) > 500:
                    j = j[:500] + "..."
                text.set("{} = {}".format(i, j))

            ttk.Label(
                about, text=f"{TITLE} 版本：{VERSION}", font=("微软雅黑", 15)
            ).pack()
            ttk.Label(about, text="开发语言：Python", font=("微软雅黑", 12)).pack()
            ttk.Label(about, text="2024-01 ~ 2024-06", font=("Consolas", 12)).pack()
            ttk.Label(about, text=f"{CACHE}", font=("Consolas", 9)).pack()
            ttk.Button(about, text="点击查看一条随机的参数吧！", command=new_var).pack(
                pady=2
            )
            label = ttk.Label(about, textvariable=text, width=300 // 8)
            label.pack()
            new_var()

            def resize(*args):
                w = about.winfo_width()
                label.config(wraplength=w - 80, width=w // 8)

            about.minsize(300, 130)
            about.bind("<Configure>", resize)
            about.bind("<Escape>", lambda *x: about.destroy())
            about.bind("<Return>", new_var)
            about.geometry(
                f"+{self.winfo_rootx() + self.winfo_width() // 3}+{self.winfo_rooty() + self.winfo_height() // 5}"
            )
            about.wait_window()
            del self.author

        def update_tags():
            self.attribute_editor(
                "TAGS",
                "字段名与标签编辑",
                "绑定A-CDM系统字段名或标签对应关系，以确保程序正常运行并导出正确表格【标签名无法更改】",
                headers=("字段名或标签", "中文名（_分割多个数据表共有字段）"),
                width=(200, 200),
                key=lambda x: len(x),
                value=lambda x: all(x in self.GET_DATA for x in x.split("_")[:-1]),
            )
            self.TAG_MAP.clear()
            for k, v in self.TAGS.items():
                v = v.split("_")
                for i in v[:-1]:
                    if i in self.TAG_MAP:
                        self.TAG_MAP[i][k] = v[-1]
                    else:
                        self.TAG_MAP[i] = {k: v[-1]}

        def update_convert():
            convert = tk.Toplevel(self, name="convert")
            convert.attributes("-topmost", self.TOPMOST.get())
            convert.title("编辑需转换为时间的字段")
            convert.resizable(False, False)
            convert.bind("<Escape>", lambda *x: convert.destroy())
            convert.geometry(
                f"+{self.winfo_rootx() + self.winfo_width() // 10}+{self.winfo_rooty() + self.winfo_height() // 10}"
            )
            if self.winfo_viewable():
                convert.transient(self)
            convert.grab_set()
            _setup_dialog(convert)
            variables = {}
            row = col = 0
            for k, v in self.TAGS.items():
                if "标签" in v:
                    continue
                variables[k] = tk.BooleanVar(convert, k in self.CONVERT)
                ttk.Checkbutton(
                    convert,
                    text=f"{v.split('_')[-1]}\n{k}",
                    variable=variables[k],
                    width=20,
                    **ONOFFS,
                ).grid(row=row, column=col, padx=2, pady=2)
                row += 1
                if not row % 14:
                    row = 0
                    col += 1

            def confirm():
                self.CONVERT.clear()
                for k, v in variables.items():
                    if v.get():
                        self.CONVERT.append(k)
                convert.destroy()

            auto = tk.StringVar(convert, "取消所有" if self.CONVERT else "自动选择")

            def auto_command():
                auto_text = {
                    "取消所有": (
                        "自动选择",
                        lambda: [v.set(False) for v in variables.values()],
                    ),
                    "自动选择": (
                        "取消所有",
                        lambda: [
                            v.set("时" in self.TAGS[k] or "T" in self.TAGS[k])
                            for k, v in variables.items()
                        ],
                    ),
                }
                auto_ = auto_text[auto.get()]
                auto.set(auto_[0]), auto_[1]()

            ttk.Button(convert, text="确定", command=confirm, width=15).grid(
                row=14, column=0, columnspan=col + 1, sticky="ws", padx=2, pady=5
            )
            ttk.Button(convert, textvariable=auto, command=auto_command, width=15).grid(
                row=14, column=0, columnspan=col + 1, sticky="s", padx=2, pady=5
            )
            ttk.Button(convert, text="取消", command=convert.destroy, width=15).grid(
                row=14, column=0, columnspan=col + 1, sticky="es", padx=2, pady=5
            )
            convert.focus_set()

        def update_datetime_format(__edit: bool = True):
            if __edit:
                self.attribute_editor(
                    "DATETIME_FORMAT",
                    "日期时间格式编辑",
                    "双击编辑",
                    headers=("日期时间格式（Excel数字格式）", "列宽"),
                    width=(250, 50),
                    key=lambda x: len(x) >= 3,
                    anchor=("center", "center"),
                    value=lambda x: float(x) >= 3 if isnumeric(x) else False,
                    key_convert=lambda x: x.lower(),
                    value_convert=lambda x: float(x) if float(x) % 1 else int(x),
                )
            if not self.RUNNING and self.exists.get():
                menu_export[2].delete(2, tk.END)
                for i in self.DATETIME_FORMAT.keys():
                    menu_export[2].add_radiobutton(
                        variable=self.CONVERT_, value=i, label=i
                    )
                s = self.CONVERT_.get()
                if s not in self.DATETIME_FORMAT and s != "标准字符串":
                    self.CONVERT_.set(
                        list(self.DATETIME_FORMAT)[0]
                        if self.DATETIME_FORMAT
                        else "标准字符串"
                    )

        menu_setting = [
            tk.Menu(menu, name="menu_setting", tearoff=False),
        ]
        for _ in range(4):
            menu_setting.append(tk.Menu(menu, tearoff=False))
        menu.add_cascade(label="高级设置", menu=menu_setting[0])
        menu_setting[0].add_cascade(label="参数编辑", menu=menu_setting[1])
        menu_setting[1].add_command(label="修改用户信息...", command=update_login)
        menu_setting[1].add_command(label="还原参数为默认值", command=set_var)
        menu_setting[1].add_separator()
        menu_setting[1].add_command(label="字段名与标签...", command=update_tags)
        menu_setting[1].add_command(label="字段转时间...", command=update_convert)
        menu_setting[1].add_command(
            label="长字符串列宽...",
            command=lambda: self.attribute_editor(
                "WRAPTEXT",
                "长字符串列宽编辑",
                "双击编辑",
                headers=("字段名", "列宽"),
                width=(200, 50),
                key=lambda x: x in self.TAGS,
                anchor=("center", "center"),
                value=lambda x: float(x) >= 3 if isnumeric(x) else False,
                value_convert=lambda x: float(x) if float(x) % 1 else int(x),
            ),
        )
        menu_setting[1].add_command(
            label="日期时间格式...", command=update_datetime_format
        )

        menu_setting[1].add_separator()
        menu_setting[1].add_command(
            label="兴快线...",
            command=lambda: self.attribute_editor(
                "EXPRESS",
                "兴快线编辑",
                "双击编辑",
                headers=("航司二字码", "目的地机场三字码（空格分隔不同机场）"),
                key=lambda x: len(x) == 2 and x.isalnum(),
                value=lambda x: all(len(x) == 3 and x.isalpha() for x in x.split(" ")),
                key_convert=lambda x: x.upper(),
                value_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="相邻登机口...",
            command=lambda: self.attribute_editor(
                "NEARGATE",
                "相邻登机口编辑",
                "双击编辑",
                headers=("登机口", "相邻登机口（空格分隔不同登机口）"),
                key=lambda x: x.isalnum(),
                value=lambda x: all(x.isalnum() for x in x.split(" ")),
                key_convert=lambda x: x.upper(),
                value_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="机场名称...",
            command=lambda: self.attribute_editor(
                "AIRPORTNAME",
                "机场名称编辑",
                "双击编辑",
                headers=("机场IATA三字码", "机场中文名称"),
                anchor=("center", "center"),
                width=(150, 150),
                key=lambda x: len(x) == 3 and x.isalpha(),
                key_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="航班监控告警...",
            command=lambda: self.attribute_editor(
                "MONITOR",
                "航班监控告警编辑",
                "双击编辑",
                headers=("类别", "告警时间（分钟）"),
                anchor=("center", "center"),
                width=(200, 150),
                value=lambda x: x.isnumeric() and int(x) >= 1,
                value_convert=int,
            ),
        )
        menu_setting[1].add_command(
            label="跑道方向对应关系...",
            command=lambda: self.attribute_editor(
                "RUNWAYDIR",
                "跑道方向对应关系编辑",
                "双击编辑",
                headers=("跑道", "方向"),
                anchor=("center", "center"),
                width=(100, 150),
                key=lambda x: (
                    (0 < int(x[:2]) <= 36 if x[:2].isnumeric() else False)
                    and x[2].upper() in "LCR"
                    if len(x) == 3
                    else (
                        (0 < int(x) <= 36 if x.isnumeric() else False)
                        if len(x) == 2
                        else False
                    )
                ),
                value=lambda x: x in "东西南北",
                key_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="离港方向...",
            command=lambda: self.attribute_editor(
                "DIR",
                "离港方向编辑",
                "双击编辑",
                headers=("航路点或机场", "方向"),
                anchor=("center", "center"),
                width=(100, 150),
                key=lambda x: (len(x) == 3 or len(x) == 5) and x.isalpha(),
                value=lambda x: x in self.RUNWAYDIR.values(),
                key_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="管制区...",
            command=lambda: self.attribute_editor(
                "AIRPORT",
                "管制区编辑",
                "双击编辑",
                headers=("管制区前两位", "管制区名称"),
                anchor=("center", "center"),
                width=(100, 150),
                key=lambda x: len(x) == 2,
                key_convert=lambda x: x.upper(),
            ),
        )
        menu_setting[1].add_command(
            label="状态码...",
            command=lambda: self.attribute_editor(
                "OPERATIONS",
                "状态码编辑",
                "双击编辑",
                headers=("状态码", "含义"),
                anchor=("center", "center"),
                width=(100, 150),
                key=true,
            ),
        )
        menu_setting[1].add_command(
            label="地服公司...",
            command=lambda: self.attribute_editor(
                "GA",
                "地服公司编辑",
                "双击编辑；N/A无地服、N/D无数据",
                key=true,
                headers=("A-CDM地服代码", "地服中文名"),
                anchor=("center", "center"),
                width=(100, 200),
            ),
        )
        menu_setting[1].add_command(
            label="地服与航司对应关系...",
            command=lambda: self.attribute_editor(
                "AG",
                "地服与航司对应关系",
                "双击编辑；N/A无地服、N/D无数据",
                headers=("航司二字码", "A-CDM地服代码"),
                anchor=("center", "center"),
                width=(100, 200),
                key=lambda x: len(x) == 2 and x.isalnum(),
                value=lambda x: x.upper() in self.GA,
                key_convert=lambda x: x.upper(),
                value_convert=lambda x: x.upper(),
            ),
        )

        menu_setting[1].add_separator()
        menu_setting[1].add_command(
            label="标头...",
            command=lambda: self.attribute_editor(
                "HEADER",
                "标头编辑",
                "【不建议修改，用于访问A-CDM系统】",
                key=true,
                anchor=("w", "w"),
                width=(180, 370),
            ),
        )
        menu_setting[1].add_command(
            label="异常提示...",
            command=lambda: self.attribute_editor(
                "RUN_EXCEPTION",
                "异常提示编辑",
                "【用于解释各类异常信息】",
                headers=("原文", "解释"),
                anchor=("w", "w"),
                width=(250, 300),
            ),
        )

        def resize_action(*args):
            w, h = self.winfo_width(), self.winfo_height()
            self.log.config(height=(h - 150) // 19 - 3, width=w // 19 - 2)
            self.status_bar.config(width=w)
            ff_map = function_frame.winfo_ismapped()
            of_map = option_frame.winfo_ismapped()
            side = self.SIDE.get()
            if side:
                frames[0].pack_configure(fill="none")
                if not rates_frames[4].winfo_ismapped():
                    rates_frames[4].pack(
                        padx=1, pady=0, side="left", before=rates_frames[0]
                    )
                for i in rates_spacer:
                    i.config(width=(w - 950) // 20 if w > 950 else 0)
                if frames[1].winfo_ismapped():
                    frames[1].pack_configure(anchor="center")
                    for frame in rates_frames:
                        if frame.winfo_ismapped():
                            frame.pack_configure(expand=True)

            else:
                frames[0].pack_configure(fill="x")
                if rates_frames[4].winfo_ismapped():
                    rates_frames[4].forget()
                if frames[1].winfo_ismapped():
                    frames[1].pack_configure(anchor="w")
                    for frame in rates_frames:
                        if frame.winfo_ismapped():
                            frame.pack_configure(expand=False)
            side = "left"
            for col, row in func_button.items():
                button = row[0].winfo_ismapped()
                w_ = 500 + col * 145 - 145
                if w > w_ and not button:
                    for row, button in enumerate(row):
                        button.grid(row=row, column=col, padx=5, pady=2, sticky="ew")
                elif w <= w_ and button:
                    for row, button in enumerate(row):
                        button.grid_forget()

            if w <= 500 and ff_map:
                function_frame.forget()
            elif w > 500 and not ff_map:
                if of_map:
                    option_frame.forget()
                function_frame.pack(padx=0, pady=0, side=side, fill="y")
                if of_map:
                    option_frame.pack(padx=0, pady=0, side=side, fill="y")
            if w <= 390 and of_map:
                option_frame.forget()
            elif w > 390 and not of_map:
                option_frame.pack(padx=0, pady=0, side=side, fill="y")

            if frames[1].winfo_ismapped():
                for i, j in (3, 500), (1, 645), (2, 790):
                    rf_map = rates_frames[i].winfo_ismapped()
                    if w < j and rf_map:
                        rates_frames[i].forget()
                    elif w >= j and not rf_map:
                        rates_frames[i].pack(
                            padx=1,
                            pady=0,
                            side="left",
                            after=rates_frames[rates_seq[rates_seq.index(i) - 1]],
                        )

        def set_wrap():
            if self.WRAP.get():
                self.log.config(wrap="char")
                log_scroll_x.forget()
            else:
                self.log.config(wrap="none")
                log_scroll_x.pack(
                    padx=0, pady=0, expand=True, fill=tk.BOTH, before=self.status_bar
                )

        menu_setting[0].add_cascade(label="窗口设置", menu=menu_setting[2])
        menu_setting[2].add_checkbutton(
            label="窗口置顶",
            variable=self.TOPMOST,
            command=lambda: self.attributes("-topmost", self.TOPMOST.get()),
            **ONOFFS,
        )
        menu_setting[2].add_separator()
        menu_setting[2].add_radiobutton(
            label="功能栏居中对齐",
            value=True,
            variable=self.SIDE,
            command=resize_action,
        )
        menu_setting[2].add_radiobutton(
            label="功能栏靠左对齐",
            value=False,
            variable=self.SIDE,
            command=resize_action,
        )
        menu_setting[2].add_separator()
        menu_setting[2].add_checkbutton(
            label="日志栏自动换行", variable=self.WRAP, command=set_wrap, **ONOFFS
        )
        menu_setting[2].add_cascade(label="日志栏最大条数", menu=menu_setting[4])
        for i in 50, 100, 200, 500, 1000:
            menu_setting[4].add_radiobutton(
                label=f"{i}条", variable=self.LOGLIMIT, value=i
            )
        menu_setting[4].add_radiobutton(label="无限制", variable=self.LOGLIMIT, value=0)

        def set_datetime():
            if self.NOW_.get():
                if messagebox.askyesno(
                    TITLE,
                    "设置当前时间将影响“导出数据”的日期时间、“导出信息”的统计尺度{}，是否确认进行更改？".format(
                        "（“运行概述”功能按当前无起飞落地时间统计，其余功能按设置后的时间节点对比统计）"
                    ),
                ):
                    s = self.datetime_selector(
                        self,
                        datetime.now(),
                        (INIT, lambda: datetime.now() + timedelta(7)),
                        "设置当前时间",
                        True,
                    )
                    if s and not self.RUNNING and self.exists.get():
                        self.NOW.set(s)
                        menu_setting[0].entryconfig(2, label=s[:16])
                        self.update_timestamp.set(0)
                        self.update_status()
                        return 0
            self.NOW_.set(False)
            menu_setting[0].entryconfig(2, label="设置当前时间")
            self.update_status()
            return 1

        menu_setting[0].add_checkbutton(
            label="设置当前时间...", command=set_datetime, variable=self.NOW_, **ONOFFS
        )

        ss, se = tk.StringVar(self), tk.StringVar(self)
        ss_, se_ = tk.BooleanVar(self), tk.BooleanVar(self)

        def schedule_set(sc: tk.StringVar, sc_: tk.BooleanVar, runs: bool):
            if sc_.get():
                if self.RUNNING and runs:
                    self.showinfo("请停止自动化后再计划启动！")
                elif not self.RUNNING and not runs:
                    self.showinfo("请启动自动化后再计划停止！")
                else:
                    s = self.datetime_selector(
                        self,
                        datetime.now() + timedelta(hours=5),
                        (datetime.now, lambda: datetime.now() + timedelta(7)),
                        "设置计划时间",
                        True,
                    )
                    if s and self.exists.get():
                        self.update_log(
                            "应用计划于{}{}自动化".format(s, "启动" if runs else "停止")
                        )
                        s = (
                            datetime.fromisoformat(s).timestamp()
                            - datetime.now().timestamp()
                        )
                        sc.set(self.after(int(s * 1000), launch if runs else terminate))
                        return 0
            else:
                self.update_log("应用计划自动{}取消".format("启动" if runs else "停止"))
            sc_.set(False)
            if sc.get():
                self.after_cancel(sc.get())
                sc.set("")
            return 1

        menu_setting[0].add_checkbutton(
            label="计划自动启动...",
            command=lambda: schedule_set(ss, ss_, True),
            variable=ss_,
            **ONOFFS,
        )
        menu_setting[0].add_checkbutton(
            label="计划自动停止...",
            command=lambda: schedule_set(se, se_, False),
            variable=se_,
            **ONOFFS,
        )

        menu_setting[0].add_cascade(label="缓存设置与时间", menu=menu_setting[3])
        menu_setting[3].add_command(label="清除所有缓存", command=self.clear_cache)
        menu_setting[3].add_separator()
        menu_setting[3].add_radiobutton(label="12小时", variable=self.CACHE, value=0.5)
        for i in 1, 2, 3, 4, 5, 6, 7:
            menu_setting[3].add_radiobutton(
                label=f"{i}天", variable=self.CACHE, value=float(i)
            )
        menu_setting[0].add_command(label="重置程序", command=reset)
        menu_setting[0].add_command(label="关于...", command=about)

        # frames
        frames, frames_index = [ttk.Frame(self) for _ in range(3)], 0

        # runtime frame
        runtime_frame = ttk.Frame(frames[frames_index])
        ttk.Label(runtime_frame, text="用  户").grid(row=0, column=0, padx=2, pady=2)
        user_combobox = ttk.Combobox(
            runtime_frame,
            name="user",
            textvariable=self.USER,
            exportselection=False,
            values=list(self.LOGIN.keys()),
            state="readonly",
            width=11,
        )
        user_combobox.grid(row=0, column=1, padx=2, pady=2, sticky="w")
        user_button = ttk.Button(
            runtime_frame, text="...", command=update_login, width=2
        )
        user_button.grid(row=0, column=2, padx=2, pady=2, sticky="e")

        ttk.Label(runtime_frame, text="任务计划").grid(row=1, column=0, padx=2, pady=4)
        schedule_combobox = ttk.Combobox(
            runtime_frame,
            name="sche",
            textvariable=self.SCHEDULE,
            exportselection=False,
            values=list(self.SCHEDULES.keys()),
            state="readonly",
            width=15,
        )
        schedule_combobox.grid(
            row=1, column=1, padx=2, pady=4, columnspan=2, sticky="w"
        )

        launch_button = ttk.Button(
            runtime_frame, text="开始", command=launch, style="Launch.TButton"
        )
        launch_button.grid(row=0, column=3, rowspan=2, sticky="nsew", padx=2, pady=2)
        refresh_button = ttk.Button(
            runtime_frame,
            text="立刻更新",
            command=lambda: threading.Thread(target=self.get_data).start(),
            width=5,
        )
        refresh_button.grid(row=3, column=3, sticky="nsew", padx=2, pady=2)

        self.progress_bar = ttk.Progressbar(
            runtime_frame, variable=self.progress, maximum=100, length=180
        )
        self.progress_bar.grid(
            row=3, column=0, columnspan=3, padx=2, pady=2, sticky="we"
        )

        runtime_frame.pack(padx=5, pady=0, side="left", fill="y")

        # function frame
        function_frame = ttk.Frame(frames[0])
        ttk.Separator(function_frame, orient="vertical").grid(
            row=0, column=0, rowspan=4, padx=2, pady=2, sticky="ns"
        )

        func_button = {}
        for i, j in enumerate(func.keys()):
            col = (i + 3) // 3
            if col in func_button:
                func_button[col].append(
                    ttk.Button(
                        function_frame,
                        text=j,
                        width=13 if col == 1 else 18,
                        command=func[j],
                    )
                )
            else:
                func_button[col] = [
                    ttk.Button(
                        function_frame,
                        text=j,
                        width=13 if col == 1 else 18,
                        command=func[j],
                    )
                ]

        # option frame
        option_frame = ttk.Frame(frames[0])
        ttk.Separator(option_frame, orient="vertical").grid(
            row=0, column=0, rowspan=4, padx=5, pady=2, sticky="ns"
        )
        ttk.Checkbutton(
            option_frame,
            text="自动信息",
            variable=self.AUTOSEL,
            **ONOFFS,
        ).grid(sticky="w", row=0, column=1, padx=2, pady=0)
        ttk.Checkbutton(
            option_frame,
            text="大面积航延",
            variable=self.msg_para[2],
            **ONOFFS,
        ).grid(sticky="w", row=1, column=1, padx=2, pady=0)
        ttk.Checkbutton(
            option_frame,
            text="霜天气",
            variable=self.msg_para[0],
            command=lambda: (
                self.msg_para[1].set(False) if self.msg_para[1].get() else None
            ),
            **ONOFFS,
        ).grid(sticky="w", row=2, column=1, padx=2, pady=0)
        ttk.Checkbutton(
            option_frame,
            text="冰雪天气",
            variable=self.msg_para[1],
            command=lambda: (
                self.msg_para[0].set(False) if self.msg_para[0].get() else None
            ),
            **ONOFFS,
        ).grid(sticky="w", row=3, column=1, padx=2, pady=0)
        frames[frames_index].pack(padx=10, pady=2)

        # rates frame
        def default_rates(*args):
            tips = (
                "始发",
                "放行",
                "起飞",
                "进港",
                {
                    "text": "取消",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "取消",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "已起",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "已落",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "未起",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "未落",
                    "width": 4,
                    "font": ("微软雅黑", 10),
                    "foreground": "dimgrey",
                },
                {
                    "text": "CTOT推点",
                    "foreground": "dimgrey",
                    "font": ("微软雅黑", 9),
                    "width": 8,
                },
                {
                    "text": "延误未起",
                    "foreground": "dimgrey",
                    "font": ("微软雅黑", 9),
                    "width": 7,
                },
                {
                    "text": "大面积航延",
                    "foreground": "dimgrey",
                    "background": "white",
                    "font": ("微软雅黑", 9),
                    "width": 9,
                },
            )
            self.rates_action(*tips, foreground="dimgrey", font=("微软雅黑", 10))

        frames_index += 1
        rates_frames = [
            ttk.Frame(frames[frames_index], style="White.TFrame") for _ in range(6)
        ]
        self.rates = [
            ttk.Label(rates_frames[0], background="white", anchor="center")
            for _ in range(4)
        ]
        self.rates.extend(
            ttk.Label(rates_frames[i % 2 + 1], background="white", anchor="center")
            for i in range(6)
        )
        self.rates.extend(
            ttk.Label(rates_frames[3], background="white", anchor="center")
            for _ in range(3)
        )
        frames[frames_index].bind("<Enter>", default_rates)
        self.rates_bind = ["", frames[frames_index].bind, frames[frames_index].unbind]

        ttk.Label(
            rates_frames[0],
            text="正常性",
            background="white",
            font=("微软雅黑", 11),
            anchor="center",
        ).grid(column=0, row=0, padx=3, pady=0, sticky="ew")
        for i in range(4):
            self.rates[i].grid(column=2 * i + 1, row=0, padx=0, pady=0, sticky="ew")
        for i in range(1, 4):
            ttk.Separator(rates_frames[0], orient="vertical").grid(
                column=2 * i, row=0, pady=2, sticky="ns"
            )

        ttk.Label(
            rates_frames[3],
            text="航延指标",
            background="white",
            font=("微软雅黑", 11),
            anchor="center",
        ).grid(column=0, row=0, padx=3, pady=0, sticky="ew")
        self.rates[10].grid(column=1, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[3], orient="vertical").grid(
            column=2, row=0, pady=2, sticky="ns"
        )
        self.rates[11].grid(column=3, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[3], orient="vertical").grid(
            column=4, row=0, pady=2, sticky="ns"
        )
        self.rates[12].grid(column=5, row=0, padx=0, pady=0, sticky="ew")

        ttk.Label(
            rates_frames[1],
            text="离港",
            background="white",
            font=("微软雅黑", 11),
            anchor="center",
        ).grid(column=0, row=0, padx=3, pady=0, sticky="ew")
        self.rates[6].grid(column=1, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[1], orient="vertical").grid(
            column=2, row=0, pady=2, sticky="ns"
        )
        self.rates[8].grid(column=3, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[1], orient="vertical").grid(
            column=4, row=0, pady=2, sticky="ns"
        )
        self.rates[4].grid(column=5, row=0, padx=0, pady=0, sticky="ew")

        ttk.Label(
            rates_frames[2],
            text="进港",
            background="white",
            font=("微软雅黑", 11),
            anchor="center",
        ).grid(column=0, row=0, padx=3, pady=0, sticky="ew")
        self.rates[7].grid(column=1, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[2], orient="vertical").grid(
            column=2, row=0, pady=2, sticky="ns"
        )
        self.rates[9].grid(column=3, row=0, padx=0, pady=0, sticky="ew")
        ttk.Separator(rates_frames[2], orient="vertical").grid(
            column=4, row=0, pady=2, sticky="ns"
        )
        self.rates[5].grid(column=5, row=0, padx=0, pady=0, sticky="ew")

        rates_spacer = [
            ttk.Label(rates_frames[i], text=" ", background="white") for i in (4, 5)
        ]
        rates_seq = (4, 0, 3, 1, 2, 5)

        for i in rates_spacer:
            i.pack(padx=0, pady=0, side="left")

        for i in rates_seq:
            rates_frames[i].pack(padx=1, pady=0, side="left")
        frames[frames_index].config(style="White.TFrame")

        # log frame
        frames_index += 1
        sub_frames = [ttk.Frame(frames[frames_index]) for _ in range(2)]
        self.log = tk.Text(
            sub_frames[0],
            name="log",
            font=("微软雅黑", 10),
            wrap="char" if self.WRAP.get() else "none",
            state=tk.DISABLED,
        )
        self.log.pack(padx=0, pady=0, expand=True, fill=tk.BOTH, side="left")
        log_scroll_x = ttk.Scrollbar(
            sub_frames[1],
            name="log_scroll_x",
            orient="horizontal",
            command=self.log.xview,
        )
        log_scroll_y = ttk.Scrollbar(
            sub_frames[0],
            name="log_scroll_y",
            orient="vertical",
            command=self.log.yview,
        )
        self.log.config(
            xscrollcommand=log_scroll_x.set, yscrollcommand=log_scroll_y.set
        )
        if not self.WRAP.get():
            log_scroll_x.pack(padx=0, pady=0, expand=True, fill=tk.BOTH)
        log_scroll_y.pack(padx=0, pady=0, expand=True, fill=tk.BOTH)

        self.log_tag = {
            "bold": {"font": ("微软雅黑", 10, "bold")},
            "time": {"font": ("Arial", 12, "bold")},
            "suffix": {"foreground": "crimson"},
            "text": {"background": "#e1fae1"},
            "file": {"background": "#c8f0f0"},
            "warn": {"foreground": "darkred"},
            "schedule": {"background": "lightgrey"},
            "monitor": {"background": "#fffacd"},
        }
        for k, v in self.log_tag.items():
            self.log.tag_config(k, **v)

        self.HISTORY = dict()
        columns = {
            "RATES": [
                "始发",
                "放行",
                "起飞",
                "进港",
                "离港取消",
                "进港取消",
                "离港已执行",
                "进港已执行",
                "离港未执行",
                "进港未执行",
                "CTOT推点",
                "延误未起飞",
                "大面积航延",
                "启动标准",
                "四地八场",
            ]
        }
        for k in ("CTOT", "json"), ("COBT", "json"), ("RATES", "xlsx"):
            v = HISTORY.format(k[0].lower(), k[1])
            if os.path.exists(v):
                try:
                    if k[1] == "json":
                        self.HISTORY[k[0]] = read_json(
                            v, orient="index", dtype="datetime64[s]", date_unit="s"
                        ).sort_index(axis=1)
                    else:
                        self.HISTORY[k[0]] = (
                            read_excel(v, index_col=0)
                            .fillna("")
                            .reindex(columns=columns[k[0]])
                        )
                    continue
                except Exception as exception:
                    exception = repr(exception)
                    self.update_log(
                        f"历史记录{v}加载失败 ({exception[: exception.find('(')]})",
                        "warn",
                    )
            self.HISTORY[k[0]] = (
                DataFrame({"guid": []}).set_index("guid", drop=True)
                if k[1] == "json"
                else DataFrame(columns=columns[k[0]])
            )

        self.status = tk.StringVar(self)
        self.ongoing, self.warning = dict(), dict()
        self.status_bar = ttk.Label(
            sub_frames[1], name="status", textvariable=self.status, width=500
        )
        self.status_bar.bind("<Double-1>", lambda *args: self.update_status(reset=True))
        self.status_bar.pack(padx=0, pady=0, fill=tk.BOTH)
        self.update_status(
            default="欢迎使用{}{}，{}".format(
                TITLE,
                (
                    ""
                    if kwargs.get("version", VERSION) >= VERSION
                    else f"（版本升级为{VERSION}）"
                ),
                "已加载设置" if os.path.exists(settings) else "已配置预设",
            )
        )

        sub_frames[0].pack(padx=0, pady=0, expand=True, fill=tk.BOTH)
        sub_frames[1].pack(padx=0, pady=0, fill=tk.BOTH)
        frames[frames_index].pack(padx=0, pady=0, expand=True, fill=tk.BOTH)

        icon = open("autoicon.ico", "wb+")
        icon.write(base64.b64decode(autoicon))
        icon.close()
        self.iconbitmap("autoicon.ico", "autoicon.ico")
        os.remove("autoicon.ico")

        # default_rates()
        # frames[1].pack(padx=0, pady=0, fill="x", before=frames[2])
        self.attributes("-topmost", self.TOPMOST.get())
        self.bind("<Configure>", resize_action)
        w, h = 800, 450
        self.geometry(
            f"{w}x{h}+{(self.winfo_screenwidth() - w) // 2}+{(self.winfo_screenheight() - h) // 2}"
        )
        self.minsize(305, 300)
        self.report_callback_exception = self.handle_exception
        threading.excepthook = self.handle_exception

        def quit(*args):
            if "sync" in self.RUNNING:
                self.showinfo("数据同步运行中，请勿关闭！")
            else:
                self.exists.set(False)
                self.destroy()

        self.wm_protocol("WM_DELETE_WINDOW", quit)
        self.focus_set()
        self.update_log(
            "{}初始化成功，{}".format(
                TITLE,
                (
                    "点击“开始”启动自动化"
                    if kwargs.get("LOGIN")
                    else "请配置用户后启动自动化"
                ),
            )
        )
        # self.get_monitor()
        self.mainloop()

        # save upon exit
        self.exists.set(False)
        if not self.reset:
            try:
                self.clear_cache(self.CACHE.get())
                if settings:
                    settings_dict, dir_list = {"version": VERSION}, self.__dir__()
                    for k in dir_list[
                        dir_list.index("EXPRESS") : dir_list.index("GET_DATA")
                    ]:
                        value = self.__getattribute__(k)
                        settings_dict[k] = (
                            value.get() if isinstance(value, tk.Variable) else value
                        )
                    with open(settings, "w") as json:
                        dump(settings_dict, json)
            except Exception:
                title, tb = f"{TITLE} - 关闭时出错", format_exc()
                try:
                    with open(LOG, "a", encoding="UTF-8") as output:
                        output.write("{:.19}  {}".format(str(datetime.now()), tb))
                except Exception:
                    title += "，日志保存失败"
                messagebox.showerror(title, tb[tb[:-1].rfind("\n") + 1 : -1])
        else:
            del self

    def clear_cache(self, day: int = 0):
        if not day and not messagebox.askyesno(TITLE, "确认清除所有日志和历史记录？"):
            return 1
        days = datetime.now() - timedelta(day)
        for i in "COBT", "CTOT":
            self.HISTORY[i] = (
                self.HISTORY[i]
                .loc[:, self.HISTORY[i].columns >= days]
                .dropna(axis=0, how="all")
                .copy()
            )
        i = self.HISTORY["RATES"].loc[self.HISTORY["RATES"].index < days].index
        self.HISTORY["RATES"].drop(index=i, inplace=True)

        if os.path.exists(LOG):
            with open(LOG, "r", encoding="UTF-8") as log:
                lines = log.readlines()
            for i, line in enumerate(lines):
                try:
                    if datetime.fromisoformat(line[:19]) >= days:
                        break
                except Exception:
                    continue
            if i > 0:
                try:
                    with open(LOG, "w", encoding="UTF-8") as log:
                        log.writelines(lines[i:])
                    status = "日志清除成功，"
                except PermissionError:
                    status = "日志清除失败，"
        if not day:
            count = self.save_history(False)
            status += (
                ("已清除{}历史记录".format("所有" if count == 3 else "部分"))
                if count
                else "历史记录清除失败"
            )
            self.update_log(status)
        return 0

    def save_history(self, clear_cache: bool = True):
        self.lock.acquire()
        if clear_cache:
            self.clear_cache(self.CACHE.get())
        count = 0
        for k, v in self.HISTORY.items():
            try:
                if k == "RATES":
                    v.to_excel(HISTORY.format(k.lower(), "xlsx"), index_label="时间")
                else:
                    output, v = {}, v.copy()
                    v.columns = v.columns.map(lambda x: x.value // 1000000000)
                    for i, v in v.iterrows():
                        output[i] = (
                            v.dropna().map(lambda x: x.value // 1000000000).to_dict()
                        )
                    with open(HISTORY.format(k.lower(), "json"), "w") as json:
                        dump(output, json)
                count += 1
            except Exception as exception:
                exception = repr(exception)
                self.update_log(
                    f"存储{k}历史记录失败 ({exception[: exception.find('(')]})", "warn"
                )
        try:
            self.MONI.reset_index().to_excel(MONI, index=False)
        except Exception as exception:
            exception = repr(exception)
            self.update_log(
                f"存储告警记录失败 ({exception[: exception.find('(')]})", "warn"
            )
        self.lock.release()
        return count

    def get_next_update(self, interval: int, offset: int):
        now = datetime.now()
        if not interval:
            i = self.UPDATE_INTERVAL.get()
            alert = (
                not self.HISTORY["RATES"]["启动标准"].empty
                and self.HISTORY["RATES"]["启动标准"].iloc[-1]
            )
            interval = (
                (5 if i == 0 else min(5, i))
                if alert
                else (
                    15
                    if now.hour <= 5
                    else 10 if now.hour <= 6 or now.hour >= 22 else 5
                )
            )
        interval *= 60
        now = now.timestamp() % interval
        interval = (
            1 if interval - now - offset >= min(60, interval // 2) else 2
        ) * interval - now
        return int((interval - offset) * 1000)

    def attribute_editor(
        self, attr: str, title: str = "参数编辑", tip: str = "", **kwargs
    ) -> int:
        """
        To edit list or dict `globals()` setted in `self`.

        Parameters
        --
        attr: `str`, attribute name setted in self, <RESET> button is enabled if found in `globals()`.
        title: `str`, editor window name.
        tip: `str`, editor tip showed at lower center.

        Kwargs
        --
        master: `tk.Misc`, default `self`.
        headers: `tuple[str]`, set column text[key, value], example: `('My Keys', 'My Values')`.
        width: `tuple[int]`, set column width[key, value] in pixels, example: `(150, 250)`.
        anchor: `tuple[str]`, set column anchor[key, value], example: `('center', 'w')`.

        key: `Callable`, a function or lambda to check key entry validity. Dict key edit is locked if not set.
        value: `Callable | dict[str, Callable]`, a function or lambda to check value entry validity.
        To check different value of key with a specific method, use dict[str, Callable].

        key_convert: `Callable`, convert key entry `str` to desired `object` upon saving.
        value_convert: `Callable | dict[str, Callable]`, convert value entry `str` to desired `object` upon saving.
        To convert different value of key with a specific method, use dict[str, Callable].
        """

        var = self.__getattribute__(attr)
        isdict = isinstance(var, dict)
        editable = (isdict and kwargs.get("key")) or not isdict
        anchor = kwargs.pop("anchor", ("center", "w"))
        width = kwargs.pop("width", (100, 450) if isdict else (20, 530))
        headers = {
            "#1": ("键名", {"anchor": anchor[0], "minwidth": 20, "width": width[0]}),
            "#2": ("键值", {"anchor": anchor[1], "minwidth": 20, "width": width[1]}),
        }
        width = sum(width)
        header = kwargs.pop("headers", (headers["#1"][0], headers["#2"][0]))
        col_to_header = {"键名": header[0], "键值": header[1]}
        master = kwargs.pop("master", self)

        editor = tk.Toplevel(master, name="editor")
        editor.attributes("-topmost", self.TOPMOST.get())
        editor.title(title)
        editor.resizable(False, False)
        editor.geometry(
            f"+{master.winfo_rootx() + master.winfo_width() // 4}+{master.winfo_rooty() + master.winfo_height() // 4}"
        )
        if master.winfo_viewable():
            editor.transient(master)
        editor.grab_set()
        _setup_dialog(editor)

        def get_edit(event):
            row = table.identify_row(event.y)
            col = table.identify_column(event.x)
            if row and col:
                if col == "#1" and not editable:
                    return 1
                else:
                    col = headers[col][0]
                edit(row, col, **kwargs)
            else:
                new_row()

        def edit(row, col, new: bool = False, **kwargs):
            editing = tk.Toplevel(editor, name="editing")
            editing.attributes("-topmost", self.TOPMOST.get())
            editing.title(f"{col_to_header[col]}{'新建' if new else '编辑'}")
            editing.resizable(False, False)
            editing.geometry(
                f"+{editor.winfo_rootx() + editor.winfo_width() // 6}+{editor.winfo_rooty() + editor.winfo_height() // 6}"
            )
            if editor.winfo_viewable():
                editing.transient(editor)

            for i in newb, delb, upb, downb:
                i.config(state=tk.DISABLED)
            table.config(selectmode="none")
            table.unbind("<<TreeviewSelect>>")
            table.unbind("<Escape>")

            item = table.item(row, "values")
            if "值" in col:
                check = kwargs.get("value", true)
                if isinstance(check, dict):
                    check = check.get(item[0], true)
                item = item[1]
            else:
                item, check = item[0], kwargs.get("key")
            scrollbar_e = ttk.Scrollbar(editing)
            entry = tk.Text(
                editing,
                width=width // 12,
                height=5,
                undo=True,
                wrap="char",
                font=("微软雅黑", 10),
                yscrollcommand=scrollbar_e.set,
            )
            scrollbar_e.config(command=entry.yview)
            entry.insert("0.0", item)

            popup = tk.Menu(editing, tearoff=False)

            def cut():
                try:
                    copy(), entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
                except Exception:
                    ...

            def copy():
                try:
                    (
                        entry.clipboard_clear(),
                        entry.clipboard_append(entry.get(tk.SEL_FIRST, tk.SEL_LAST)),
                    )
                except Exception:
                    ...

            def paste():
                try:
                    entry.insert(tk.INSERT, entry.selection_get(selection="CLIPBOARD"))
                except Exception:
                    ...

            def selall():
                entry.tag_add("sel", "0.0", tk.END)
                return "break"

            def delete():
                try:
                    entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
                except Exception:
                    ...

            popup.add_command(label="删除", command=delete)
            popup.add_command(label="剪切", command=cut)
            popup.add_command(label="复制", command=copy)
            popup.add_command(label="粘贴", command=paste)
            popup.add_separator()
            popup.add_command(label="全选", command=selall)

            entry.bind(
                "<Button-3>", lambda event: popup.post(event.x_root, event.y_root)
            )
            entry.grid(row=0, column=0, columnspan=2)
            scrollbar_e.grid(row=0, column=3, sticky="ns")

            def confirm(*args):
                v = entry.get("0.0", tk.END).strip().strip("\n").strip()
                if check(v):
                    table.set(row, col, v)
                    editing.destroy()
                    table.see(row)
                else:
                    messagebox.showinfo(
                        TITLE, f"输入{col}无效或无法匹配，请参考提示", parent=editing
                    )

            ttk.Button(
                editing, text="确认", width=width // 24 - 1, command=confirm
            ).grid(sticky="ws", row=2, column=0, columnspan=4, padx=5, pady=2)
            ttk.Button(
                editing, text="取消", width=width // 24 - 1, command=editing.destroy
            ).grid(sticky="es", row=2, column=0, columnspan=4, padx=5, pady=2)
            editing.grab_set()
            _setup_dialog(editing)
            editing.wait_window()
            if editor.winfo_exists():
                for i in newb, delb, upb, downb:
                    i.config(state=state(editable))
                table.config(selectmode="extended")
                table.bind("<<TreeviewSelect>>", button_update)
                table.bind("<Escape>", lambda x: table.selection_set([]))
                table.focus_set()
                table.selection_set([row] if row in table.children else [])
                if table.item(row)["values"][0] == "" and new:
                    table.delete(row)
            return 0

        scrollbar = ttk.Scrollbar(editor)
        table = ttk.Treeview(
            editor,
            show="headings",
            columns=list(i[0] for i in headers.values()),
            yscrollcommand=scrollbar.set,
        )
        for k, v in headers.values():
            table.column(k, **v)
            table.heading(k, text=col_to_header[k])
        scrollbar.config(command=table.yview)
        table.grid(row=0, column=0, columnspan=14, sticky="nsew")
        scrollbar.grid(row=0, column=14, sticky="ns")

        for k, v in var.items() if isdict else enumerate(var):
            table.insert("", tk.END, values=(k, str(v)))
        table.bind("<Double-1>", get_edit)

        def new_row():
            if isdict:
                row, col = table.insert("", tk.END), "键名"
                table.set(row, 0, "")
            else:
                row, col = table.insert("", tk.END), "键值"
                table.set(row, 0, int(table.item(table.prev(row), "values")[0]) + 1)
            table.set(row, 1, "")
            edit(row, col, True, **kwargs)

        def del_row():
            if messagebox.askyesno(TITLE, "确定删除？", parent=editor):
                for i in table.selection():
                    table.delete(i)

        def move_up():
            sel = [table.selection(), []]
            if table.get_children()[0] in sel[0]:
                return 0
            for i in sel[0]:
                j = table.prev(i)
                k = table.item(j, "values")
                for v in enumerate(table.item(i, "values")):
                    if not isdict and v[0] == 0:
                        continue
                    table.set(j, *v)
                for v in enumerate(k):
                    if not isdict and v[0] == 0:
                        continue
                    table.set(i, *v)
                sel[1].append(j)
            table.selection_set(sel[1])

        def move_down():
            sel = [table.selection()[::-1], []]
            if table.get_children()[-1] in sel[0]:
                return 0
            for i in sel[0]:
                j = table.next(i)
                k = table.item(j, "values")
                for v in enumerate(table.item(i, "values")):
                    if not isdict and v[0] == 0:
                        continue
                    table.set(j, *v)
                for v in enumerate(k):
                    if not isdict and v[0] == 0:
                        continue
                    table.set(i, *v)
                sel[1].append(j)
            table.selection_set(sel[1])

        def button_update(*args):
            flag = table.selection() and (not isdict or editable)
            for i in delb, upb, downb:
                i.config(state=state(flag))
            newb.config(state=state(editable))

        button_style = (
            ("EditorIcon.TButton", "▲", "▼", width // 2 - 10, "－", "＋")
            if width <= 400
            else (
                "Editor.TButton",
                "▲ 上移",
                "下移 ▼",
                (width - 80) // 2,
                "－删除",
                "新建＋",
            )
        )
        upb = ttk.Button(
            editor,
            text=button_style[1],
            style=button_style[0],
            command=move_up,
            state=tk.DISABLED,
        )
        upb.grid(sticky="ws", row=1, column=0, columnspan=2, padx=2, pady=5)
        downb = ttk.Button(
            editor,
            text=button_style[2],
            style=button_style[0],
            command=move_down,
            state=tk.DISABLED,
        )
        downb.grid(sticky="ws", row=1, column=2, columnspan=2, padx=2, pady=5)
        ttk.Label(
            editor,
            text=tip,
            font=("微软雅黑", 8),
            foreground="dimgrey",
            wraplength=button_style[3],
        ).grid(row=1, column=3, columnspan=8, padx=2, pady=2)
        delb = ttk.Button(
            editor,
            text=button_style[4],
            style=button_style[0],
            command=del_row,
            state=tk.DISABLED,
        )
        delb.grid(sticky="es", row=1, column=0, columnspan=12, padx=2, pady=5)
        newb = ttk.Button(
            editor,
            text=button_style[5],
            style=button_style[0],
            command=new_row,
            state=state(editable),
        )
        newb.grid(sticky="es", row=1, column=0, columnspan=14, padx=2, pady=5)
        table.bind("<<TreeviewSelect>>", button_update)
        table.bind("<Escape>", lambda x: table.selection_set([]))

        def confirm():
            if self.RUNNING:
                messagebox.showinfo(
                    "程序已在自动化运行中，无法保存参数，请停止自动化运行后重试",
                    parent=editor,
                )
            elif isdict:
                k, v = kwargs.get("key_convert", str), kwargs.get("value_convert", str)
                update = dict()
                for i in table.get_children():
                    value = table.item(i, "values")
                    if any(value):
                        if k(value[0]) in update:
                            messagebox.showinfo(
                                TITLE, "存在相同的键名，请确保键名唯一", parent=editor
                            )
                            try:
                                table.focus_set()
                                table.selection_set(i)
                            except Exception:
                                ...
                            return 1
                        update[k(value[0])] = (
                            v.get(k(value[0]), str)(value[1])
                            if isinstance(v, dict)
                            else v(value[1])
                        )
            else:
                update = list()
                for i in table.get_children():
                    value = table.item(i, "values")[1]
                    update.append(kwargs.get("value_convert", str)(value))
            self.__setattr__(attr, update)
            editor.destroy()

        def reset():
            if messagebox.askyesno(TITLE, "确认恢复默认？", parent=editor):
                for i in table.get_children():
                    table.delete(i)
                for k, v in (
                    globals().get(attr).items()
                    if isdict
                    else enumerate(globals().get(attr))
                ):
                    table.insert("", tk.END, values=(k, str(v)))
                table.focus_set()

        ttk.Button(editor, text="保存", width=width // 25, command=confirm).grid(
            sticky="w", row=2, column=0, padx=5, pady=5, columnspan=15
        )
        ttk.Button(
            editor,
            text="恢复默认",
            width=width // 25,
            command=reset,
            state=state(attr in globals()),
        ).grid(row=2, column=0, padx=5, pady=5, columnspan=15)
        ttk.Button(editor, text="取消", width=width // 25, command=editor.destroy).grid(
            sticky="e", row=2, column=0, padx=5, pady=5, columnspan=15
        )
        editor.wait_window()

    def datetime_selector(
        self,
        master: tk.Misc,
        default: datetime,
        datetime_range: tuple[datetime] = None,
        title: str = "日期时间选择",
        cancelable: bool = False,
    ) -> str:
        default = (
            default
            if isinstance(default, datetime)
            else datetime.fromisoformat(default)
        )
        selector = tk.Toplevel(master, name="datetimeselector")
        selector.geometry(
            f"198x60+{master.winfo_rootx() + master.winfo_width() // 3}+{master.winfo_rooty() + master.winfo_height() // 3}"
        )
        selector.attributes("-topmost", self.TOPMOST.get())
        selector.title(title)
        if master.winfo_viewable():
            selector.transient(master)
        selector.grab_set()
        _setup_dialog(selector)
        result = tk.StringVar(master, str(default)[:16])

        def get_range(index: int, by: tuple = datetime_range):
            return by[index] if isinstance(by[index], datetime) else by[index]()

        def confirm():
            try:
                r = datetime.fromisoformat(
                    "{} {:02d}:{:02d}".format(
                        date_entry.get_date(), int(hour.get()), int(minute.get())
                    )
                )
                if all(datetime_range) and not get_range(0) <= r <= get_range(1):
                    messagebox.showinfo(
                        title,
                        "日期时间范围应为{:.16}至{:.16}，请重设日期".format(
                            str(get_range(0) + timedelta(minutes=1)), str(get_range(1))
                        ),
                        parent=selector,
                    )
                    return 1
                result.set(str(r))
                selector.destroy()
            except Exception:
                messagebox.showinfo(title, "时间设置错误", parent=selector)

        def cancel():
            if cancelable:
                result.set("")
            selector.destroy()

        hour = tk.StringVar(selector, "{:02d}".format(default.hour))
        minute = tk.StringVar(selector, "{:02d}".format(default.minute))
        date_entry = DateEntry(
            selector,
            font="微软雅黑 10",
            width=10,
            background="lightgrey",
            locale="zh_CN",
            date_pattern="yyyy-MM-dd",
        )
        date_entry.set_date(default)
        date_entry.grid(row=1, column=0, padx=2, pady=2)
        ttk.Combobox(
            selector,
            values=["{:02d}".format(i) for i in range(24)],
            width=2,
            textvariable=hour,
        ).grid(row=1, column=1, padx=2, pady=2)
        ttk.Label(selector, text=":").grid(row=1, column=2, padx=0, pady=2)
        ttk.Combobox(
            selector,
            values=["{:02d}".format(i) for i in range(60)],
            width=2,
            textvariable=minute,
        ).grid(row=1, column=3, padx=2, pady=2)
        ttk.Button(selector, text="确定", command=confirm).grid(
            sticky="sw", row=2, column=0, padx=2, pady=2, columnspan=4
        )
        ttk.Button(selector, text="取消", command=cancel).grid(
            sticky="se", row=2, column=0, padx=2, pady=2, columnspan=4
        )

        selector.protocol("WM_DELETE_WINDOW", cancel)
        selector.resizable(False, False)
        selector.bind("<Escape>", lambda x: selector.destroy())
        date_entry.focus_set()
        selector.wait_window()
        return result.get()[:16]

    def rates_action(self, *args, **kwargs):
        for i, j in enumerate(args):
            if isinstance(j, dict):
                self.rates[i].config(**j)
            else:
                self.rates[i].config(text=str(j), **kwargs)

    def rates_handler(
        self, key: str, text: str, value: float, last_mean: float, **kwargs
    ):
        colors = RATES_COLOR[key]
        if value > last_mean:
            kwargs["text"] = f"{text}↑"
        elif value < last_mean:
            kwargs["text"] = f"{text}↓"
        else:
            kwargs["text"] = f"{text}-"
        if value <= colors[-1][0]:
            color = colors[-1][1]
        else:
            for i, color in colors:
                if value > i:
                    break
        kwargs["foreground"] = color
        if "font" not in kwargs:
            kwargs["font"] = ("Consolas", 11)
        return kwargs

    def ask_user(self):
        user = self.USER.get()
        if self.LOGIN.get(user):
            return user, self.LOGIN.get(user)
        elif self.LOGIN:
            self.USER.set(list(self.LOGIN.keys())[0])
            user = self.USER.get()
            return user, self.LOGIN.get(user)
        else:
            raise GuiError("请新建用户并输入账号密码")

    def ask_type(self, title: str, master: tk.Misc = None, default: list[str] = []):
        box = tk.Toplevel(master if master else self, name=title)
        box.title(title)
        box.geometry(
            f"+{master.winfo_rootx() + master.winfo_width() // 3}+{master.winfo_rooty() + master.winfo_height() // 8}"
        )
        if master.winfo_viewable():
            box.transient(master)
        box.grab_set()
        _setup_dialog(box)
        row = col = 0
        cb = {}
        confirm = tk.BooleanVar(box, False)
        for k, v in self.TYPE.items():
            cb[k] = tk.BooleanVar(box, k in default)
            ttk.Checkbutton(
                box,
                text=v + " " + k,
                variable=cb[k],
                **ONOFFS,
            ).grid(row=row, column=col, padx=5, pady=2, sticky="ew")
            row += 1
            if not row % 9:
                row = 0
                col += 1

        sel = tk.StringVar(
            box, "全不选" if all(i.get() for i in cb.values()) else "全选"
        )

        def selection():
            flag = sel.get() == "全选"
            sel.set("全不选" if flag else "全选")
            for i in cb.values():
                i.set(flag)

        ttk.Button(
            box,
            text="确定",
            command=lambda: (confirm.set(True), box.destroy()),
            width=9,
        ).grid(row=9, column=0, columnspan=col + 1, sticky="ws", padx=5, pady=5)
        ttk.Button(box, textvariable=sel, command=selection, width=9).grid(
            row=9, column=0, columnspan=col + 1, padx=5, pady=5
        )
        ttk.Button(box, text="取消", command=box.destroy, width=9).grid(
            row=9, column=0, columnspan=col + 1, sticky="es", padx=5, pady=5
        )

        box.resizable(False, False)
        box.wait_window()
        return [k for k, v in cb.items() if v.get()] if confirm.get() else default

    def ask_export(
        self,
        start: datetime,
        end: datetime,
        command: str,
        title: str = "导出表格",
        tip: str = "",
        start_range: tuple[datetime] = (None, None),
        end_range: tuple[datetime] = (None, None),
        *args,
        **kwargs,
    ):
        master = kwargs.pop("master", self)
        export = tk.Toplevel(master, name=title)
        export.attributes("-topmost", self.TOPMOST.get())
        export.geometry(
            f"+{self.winfo_rootx() + self.winfo_width() // 3}+{self.winfo_rooty() + self.winfo_height() // 3}"
        )
        export.title(title)
        export.resizable(False, False)
        export.bind("<Escape>", lambda x: export.destroy())
        if master.winfo_viewable():
            export.transient(master)
        export.grab_set()
        _setup_dialog(export)

        start_ = tk.StringVar(export, str(start)[:16])
        end_ = tk.StringVar(export, str(end)[:16])
        start_exclude = tk.BooleanVar(export)
        end_exclude = tk.BooleanVar(export, True)

        def change_dt(dt: tk.StringVar, **kwargs):
            dt.set(str(datetime.fromisoformat(dt.get()) + timedelta(**kwargs))[:16])

        def inside_range(target, range):
            if all(range):
                start = range[0] if isinstance(range[0], datetime) else range[0]()
                end = range[1] if isinstance(range[1], datetime) else range[1]()
                return not start <= target <= end
            else:
                return False

        def confirm():
            start, end = (
                datetime.fromisoformat(start_.get()),
                datetime.fromisoformat(end_.get()),
            )
            if start_exclude.get():
                start -= timedelta(minutes=1)
            if end_exclude.get():
                end -= timedelta(minutes=1)
            for i, j in (
                (start >= end, "起始至结束日期时间应大于0分钟"),
                (
                    inside_range(start, start_range),
                    "可导出的开始日期时间范围为：{:.16}至{:.16}".format(
                        *[str(i) for i in start_range]
                    ),
                ),
                (
                    inside_range(end, end_range),
                    "可导出的结束日期时间范围为：{:.16}至{:.16}".format(
                        *[str(i) for i in end_range]
                    ),
                ),
            ):
                if i:
                    messagebox.showinfo(TITLE, j + "，请重设", parent=export)
                    return 1
            if isinstance(command, str):
                self.submit_export(start, end, title, command, **kwargs)
            else:
                threading.Thread(
                    target=command, args=(start, end, *args), kwargs=kwargs
                ).start()
            export.destroy()

        row = 0
        ttk.Label(export, text="开始自", font=("微软雅黑", 10, "bold")).grid(
            row=row, column=0, padx=5, pady=1
        )
        ttk.Button(
            export,
            textvariable=start_,
            style="Datetime.TButton",
            command=lambda: start_.set(
                self.datetime_selector(
                    export, start_.get(), start_range, "设置开始日期时间"
                )
            ),
        ).grid(row=row, column=1, padx=2, pady=1, columnspan=5)
        ttk.Button(
            export,
            text="此刻",
            style="ChangeDtm.TButton",
            command=lambda: start_.set(str(datetime.now())[:16]),
        ).grid(sticky="w", row=row, column=6, padx=3, pady=1, columnspan=4)
        ttk.Button(
            export,
            text="整点",
            style="ChangeDtm.TButton",
            command=lambda: start_.set(start_.get()[:14] + "00"),
        ).grid(row=row, column=6, padx=3, pady=1, columnspan=4)
        ttk.Button(
            export,
            text="0时",
            style="ChangeDtm.TButton",
            command=lambda: start_.set(start_.get()[:11] + "00:00"),
        ).grid(sticky="e", row=row, column=6, padx=3, pady=1, columnspan=4)

        row += 1
        ttk.Checkbutton(export, text="不含", variable=start_exclude, **ONOFFS).grid(
            row=row, column=0, padx=5, pady=1
        )
        ttk.Button(
            export,
            text="-1日",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, days=-1),
        ).grid(row=row, column=1, padx=1, pady=1)
        ttk.Button(
            export,
            text="-6时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, hours=-6),
        ).grid(row=row, column=2, padx=1, pady=1)
        ttk.Button(
            export,
            text="-1时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, hours=-1),
        ).grid(row=row, column=3, padx=1, pady=1)
        ttk.Button(
            export,
            text="-30分",
            style="ChangeDtm.TButton",
            command=lambda: change_dt(start_, minutes=-30),
        ).grid(row=row, column=4, padx=1, pady=1)
        ttk.Label(export, text="  ", font=("微软雅黑", 8)).grid(
            row=row, column=5, padx=1, pady=1
        )
        ttk.Button(
            export,
            text="+30分",
            style="ChangeDtm.TButton",
            command=lambda: change_dt(start_, minutes=30),
        ).grid(row=row, column=6, padx=1, pady=1)
        ttk.Button(
            export,
            text="+1时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, hours=1),
        ).grid(row=row, column=7, padx=1, pady=1)
        ttk.Button(
            export,
            text="+6时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, hours=6),
        ).grid(row=row, column=8, padx=1, pady=1)
        ttk.Button(
            export,
            text="+1日",
            style="ChangeDt.TButton",
            command=lambda: change_dt(start_, days=1),
        ).grid(row=row, column=9, padx=1, pady=1)

        row += 1
        ttk.Label(export, text="", font=("微软雅黑", 3)).grid(
            row=row, column=0, padx=5, pady=0
        )

        row += 1
        ttk.Label(export, text="结束至", font=("微软雅黑", 10, "bold")).grid(
            row=row, column=0, padx=5, pady=1
        )
        ttk.Button(
            export,
            textvariable=end_,
            style="Datetime.TButton",
            command=lambda: end_.set(
                self.datetime_selector(
                    export, end_.get(), end_range, "设置结束日期时间"
                )
            ),
        ).grid(row=row, column=1, padx=2, pady=1, columnspan=5)
        ttk.Button(
            export,
            text="此刻",
            style="ChangeDtm.TButton",
            command=lambda: end_.set(str(datetime.now())[:16]),
        ).grid(sticky="w", row=row, column=6, padx=3, pady=1, columnspan=4)
        ttk.Button(
            export,
            text="整点",
            style="ChangeDtm.TButton",
            command=lambda: end_.set(end_.get()[:14] + "00"),
        ).grid(row=row, column=6, padx=3, pady=1, columnspan=4)
        ttk.Button(
            export,
            text="0时",
            style="ChangeDtm.TButton",
            command=lambda: end_.set(end_.get()[:11] + "00:00"),
        ).grid(sticky="e", row=row, column=6, padx=3, pady=1, columnspan=4)

        row += 1
        ttk.Checkbutton(export, text="不含", variable=end_exclude, **ONOFFS).grid(
            row=row, column=0, padx=5, pady=1
        )
        ttk.Button(
            export,
            text="-1日",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, days=-1),
        ).grid(row=row, column=1, padx=1, pady=1)
        ttk.Button(
            export,
            text="-6时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, hours=-6),
        ).grid(row=row, column=2, padx=1, pady=1)
        ttk.Button(
            export,
            text="-1时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, hours=-1),
        ).grid(row=row, column=3, padx=1, pady=1)
        ttk.Button(
            export,
            text="-30分",
            style="ChangeDtm.TButton",
            command=lambda: change_dt(end_, minutes=-30),
        ).grid(row=row, column=4, padx=1, pady=1)
        ttk.Label(export, text="  ", font=("微软雅黑", 8)).grid(
            row=row, column=5, padx=1, pady=1
        )
        ttk.Button(
            export,
            text="+30分",
            style="ChangeDtm.TButton",
            command=lambda: change_dt(end_, minutes=30),
        ).grid(row=row, column=6, padx=1, pady=1)
        ttk.Button(
            export,
            text="+1时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, hours=1),
        ).grid(row=row, column=7, padx=1, pady=1)
        ttk.Button(
            export,
            text="+6时",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, hours=6),
        ).grid(row=row, column=8, padx=1, pady=1)
        ttk.Button(
            export,
            text="+1日",
            style="ChangeDt.TButton",
            command=lambda: change_dt(end_, days=1),
        ).grid(row=row, column=9, padx=1, pady=1)

        row += 1
        ttk.Button(export, text="确定", command=confirm).grid(
            sticky="sw", row=row, column=0, padx=5, pady=5, columnspan=10
        )
        ttk.Label(
            export, text=tip, font=("微软雅黑", 8), foreground="dimgrey", wraplength=150
        ).grid(row=row, column=0, padx=5, pady=5, columnspan=10)
        ttk.Button(export, text="取消", command=export.destroy).grid(
            sticky="se", row=row, column=0, padx=5, pady=5, columnspan=10
        )

        export.focus_set()
        export.wait_window()

    def save_excel(
        self,
        parent: tk.Misc,
        title: str,
        file: str,
        func: Callable,
        enable_retry: bool = True,
        **kwargs,
    ) -> str:
        while file and parent.winfo_exists():
            try:
                if not file.strip()[-5:] == r".xlsx":
                    file += r".xlsx"
                func(file, **kwargs)
                return file
            except PermissionError:
                if enable_retry:
                    if messagebox.askretrycancel(
                        TITLE,
                        f"表格文件占用，请关闭{file}后点击重试继续保存；点击取消可另存为其他文件",
                    ):
                        continue
                    else:
                        return self.save_excel(
                            parent,
                            title,
                            filedialog.asksaveasfilename(
                                filetypes=(("Xlsx表格文件", "*.xlsx"),),
                                parent=parent,
                                confirmoverwrite=True,
                                title=title,
                                initialdir=os.path.dirname(file),
                                initialfile=(
                                    title[title.find("导出") + 2 : title.find("表格")]
                                    if "导出" in title and "表格" in title
                                    else title
                                )
                                + f"{datetime.now().strftime(r'_%m%d_%H%M%S')}.xlsx",
                            ),
                            func,
                            **kwargs,
                        )
                else:
                    return self.save_excel(
                        parent,
                        title,
                        file.replace(
                            ".xlsx", f"{datetime.now().strftime(r'_%m%d_%H%M%S')}.xlsx"
                        ),
                        func,
                        enable_retry,
                        **kwargs,
                    )
        return ""

    def save_excel_img(
        self, file: str, img: str, sheetname: str = None, sheetrange: str = None
    ):
        try:
            self.lock.acquire()
            try:
                if os.path.exists(img):
                    os.remove(img)
            except Exception:
                img = img.replace(
                    ".png", f"{datetime.now().strftime(r'_%m%d_%H%M%S')}.png"
                )
            export_img(file, img, sheetname, sheetrange)
            return img
        except Exception as exception:
            exception = repr(exception)
            self.update_log(
                f"表格图片{img} 保存失败 ({exception[: exception.find('(')]})", "warn"
            )
            return ""
        finally:
            self.lock.release()

    def submit_export(
        self, start: datetime, end: datetime, title: str, *command: str, **kwargs
    ):
        threading.Thread(
            target=self.submit,
            args=(start, end, title, *command),
            kwargs=kwargs,
            daemon=True,
        ).start()

    def submit(
        self, start: datetime, end: datetime, title: str, *command: str, **kwargs
    ):
        try:
            filename = (
                title[title.find("导出") + 2 : title.find("表格")]
                if "导出" in title and "表格" in title
                else title
            ) + f"{datetime.now().strftime(r'_%m%d_%H%M%S')}.xlsx"
            file = (
                f"{self.EXPORT_PATH.get()}/{filename}"
                if self.EXPORT_PATH_.get()
                else filedialog.asksaveasfilename(
                    filetypes=(("Xlsx表格文件", "*.xlsx"),),
                    confirmoverwrite=True,
                    parent=self,
                    title=title,
                    initialdir=self.EXPORT_PATH.get(),
                    initialfile=filename,
                )
            )
            if file:
                self.update_status({"导出表格": "获取数据"}, set_bar=True)
                kw = {}
                for name in command:
                    kw.update(
                        self.GET_DATA[name](
                            start,
                            end,
                            progress=0,
                            **kwargs,
                        )
                    )
                self.update_status({"导出表格": "保存中"})
                file = self.save_excel(self, title, file, self.format_excel, **kw)
        finally:
            self.update_status({"导出表格": ""})

    @staticmethod
    def datetime_split(
        start: datetime, end: datetime, split_hour: float
    ) -> list[tuple[datetime]]:
        delta = end - start
        delta = delta.total_seconds()
        seq = []
        assert split_hour > 0 and end > start
        _start = start
        while delta > split_hour * 3600:
            _start = _start + timedelta(hours=split_hour)
            if not 2 <= _start.hour <= 5:
                seq.append((start, _start - timedelta(minutes=1)))
                start = _start
            delta = end - _start
            delta = delta.total_seconds()
        seq.append((start, end))
        return seq

    def get_flight_data(
        self,
        start: datetime,
        end: datetime,
        progress: int = 40,
        moni: bool = True,
        **kwargs,
    ) -> dict[str, DataFrame]:
        payload = self.FLIGHT_PAYLOAD.copy()
        payload["limit"] = kwargs.pop("limit", self.UPDATE_LIMIT.get())
        payload.update(kwargs)

        if payload["limit"] >= 500 and moni:
            futures = []
            self.get_session()
            with ThreadPoolExecutor() as executor:
                for i, j in self.datetime_split(start, end, 3):
                    futures.append(
                        executor.submit(
                            self.get_flights,
                            startTime=str(i)[:16],
                            endTime=str(j)[:16],
                            **payload.copy(),
                        )
                    )
                wait(futures, max(10, payload["limit"] // 50))
                data = []
                for future in futures:
                    if exception := future.exception():
                        raise exception
                    else:
                        data.append(future.result())
            data = self.get_reindex(concat(data, ignore_index=True), "航班")
        else:
            if not moni:
                payload["loadMoniJobsLater"] = True
            data = self.get_reindex(
                self.get_flights(
                    startTime=str(start)[:16], endTime=str(end)[:16], **payload
                ),
                "航班",
            )
        if len(data):
            data[self.outStot] = data[self.outSobt] + timedelta(minutes=30)
            data[self.outMttt] = data[self.outMttt].map(
                lambda x: timedelta(minutes=x if notna(x) else 65)
            )
            data[self.sttt] = data[self.outSobt] - data[self.inSibt]
            data[self.attt] = data[self.outAobt] - data[self.inAibt]
            data[self.ttt] = data[self.sttt] - data[self.outMttt]
            data[self.outOperationStatusCodeCn] = data[self.outOperationStatusCode].map(
                self.OPERATIONS
            )
            data[self.planeNo] = data[self.planeNo].fillna("")

            if self.INITIAL.get():
                init = (
                    data["outIsinitial"]
                    .map(lambda x: x[0] == "Y", IGNORE)
                    .fillna(False)
                )
            else:
                init = data.apply(
                    lambda x: (
                        data.loc[data["planeNo"] == x["planeNo"]]
                        .loc[data["outFlightDate"] == x["outFlightDate"]]
                        .index[0]
                        == x.name
                        and x["outSobt"].hour >= 6
                        and (
                            x["inStot"] < x["outFlightDate"] + timedelta(hours=6)
                            or x["inFlightDate"] < x["outFlightDate"]
                        )
                    ),
                    axis=1,
                )
                data.loc[data["inAldt"] > data["inSldt"], "outLastTot"] = (
                    data["inAldt"] - data["inSldt"] + data["outStot"]
                )
                data.loc[data["inAldt"] <= data["inSldt"], "outLastTot"] = data[
                    "outStot"
                ]
                data.loc[init, "outLastTot"] = data["outStot"]
            data["outIsinitial"] = init

        if "get_data" in self.RUNNING:
            self.progress.set(self.progress.get() + progress)
        return {"航班_航班": data}

    def get_flights(self, **payload):
        response = self.get_session().post(
            self.URLS.get("航班查询"),
            data=dumps(payload),
            headers=self.HEADER,
            timeout=max(10, payload["limit"] // 50),
        )
        if response.status_code != 200:
            raise requests.ConnectionError(f"响应异常代码{response.status_code}")
        return DataFrame.from_records(response.json().get("data").get("rows"))

    def get_tmi_info(self, *args: datetime, progress: int = 20, **kwargs):
        limit = kwargs.get("limit", self.UPDATE_LIMIT.get())
        if len(args) == 2:
            condition = {
                "startTime#le": str(args[1] - timedelta(hours=8)),
                "endTime#ge": str(args[0] - timedelta(hours=8)),
            }
        else:
            condition = {"updateTime#ge": str(args[0])}
        payload = {
            "pageNum": 1,
            "pageSize": limit,
            "orderBy": "update_time asc",
            "condition": condition,
        }
        response = self.get_session().post(
            self.URLS.get("流控"),
            data=dumps(payload),
            headers=self.HEADER,
            timeout=max(10, limit // 50),
        )
        if response.status_code != 200:
            raise requests.ConnectionError(f"响应异常代码{response.status_code}")
        data = DataFrame.from_records(
            response.json().get("data").get("list")
        ).drop_duplicates([self.tmiContent])
        data[self.tmiRegion] = data[self.tmiContent].map(
            self.tmi_content_extract, IGNORE
        )
        data = self.get_reindex(data, "流控").reset_index(drop=True)
        for i in self.publishTime, self.startTime, self.endTime:
            data[i] += timedelta(hours=8)
        if "get_data" in self.RUNNING:
            self.progress.set(self.progress.get() + progress)
        return {"流控_流控": data}

    def get_delay_data(
        self, start: datetime, end: datetime, progress: int = 20, **kwargs
    ):
        limit = kwargs.get("limit", self.UPDATE_LIMIT.get())
        response = self.get_session().post(
            self.URLS.get("延误"),
            data=dumps(
                {
                    "pageNum": 1,
                    "pageSize": limit,
                    "orderBy": "OUT_SOBT asc",
                    "condition": {
                        "outSobt#ge": str(start)[:19],
                        "outSobt#le": str(end)[:19],
                    },
                }
            ),
            headers=self.HEADER,
            timeout=max(10, limit // 50),
        )
        if response.status_code != 200:
            raise requests.ConnectionError(f"响应异常代码{response.status_code}")
        data = self.get_reindex(
            DataFrame.from_records(response.json().get("data").get("list")), "延误"
        )
        if len(data):
            data[self.priDelayReason] = (
                data[self.rstDelayReason]
                .map(lambda x: self.PRIMARY[x[:2]], IGNORE)
                .fillna(self.PRIMARY["00"])
            )
            data[self.subDelayReason] = (
                data[self.rstDelayReason]
                .map(lambda x: x[5:], IGNORE)
                .fillna(self.PRIMARY["00"])
            )
            data[self.outLastTotDelay] = data[self.outAtot] - data[self.outLastTot]
            data[self.outStotDelay] = data[self.outAtot] - data[self.outStot]
            data[self.mttt] = data[self.mttt].map(
                lambda x: timedelta(minutes=x if notna(x) else 65)
            )
            data[self.sttt] = data[self.outSobt] - data[self.inSibt]
            data[self.attt] = data[self.outAobt] - data[self.inAibt]
            data[self.ttt] = data[self.sttt] - data[self.mttt]
        if "get_data" in self.RUNNING:
            self.progress.set(self.progress.get() + progress)
        return {"延误_延误": self.update_delay(data) if self.AUTODELAY.get() else data}

    def get_flight_info(
        self, start: datetime, end: datetime, progress: int = 20, **kwargs
    ) -> dict[str, DataFrame]:
        response = self.get_session().post(
            self.URLS.get("执行监控"),
            data=dumps(
                {
                    "minFlightDate": str(start)[:19],
                    "maxUpdateTime": "",
                    "maxFlightDate": str(end)[:19],
                }
            ),
            headers=self.HEADER,
            timeout=max(10, kwargs.get("limit", 1000) // 50),
        )
        if response.status_code != 200:
            raise requests.ConnectionError(f"响应异常代码{response.status_code}")
        data = DataFrame.from_records(response.json().get("data").get("listFlightInfo"))
        data = self.get_reindex(data, "执行")
        data[self.stot] = data[self.sobt] + timedelta(minutes=30)
        data[self.mttt] = data[self.mttt].fillna(65).map(lambda x: timedelta(minutes=x))
        data[self.operationStatusCodeCn] = data[self.operationStatusCode].map(
            self.OPERATIONS
        )
        if "get_data" in self.RUNNING:
            self.progress.set(self.progress.get() + progress)
        return {
            "执行_离港": data.loc[data[self.arrivalDepartureInd] == "D"].reset_index(
                drop=True
            ),
            "执行_进港": data.loc[data[self.arrivalDepartureInd] == "A"].reset_index(
                drop=True
            ),
        }

    def get_passenger_info(self, *args, progress: int = 20, **kwargs):
        limit = kwargs.get("limit", self.UPDATE_LIMIT.get())
        response = self.get_session().post(
            self.URLS.get("旅客保障"),
            data=dumps({"arrivalDepartureInd": "D", "pageNum": 1, "pageSize": limit}),
            headers=self.HEADER,
            timeout=max(10, limit // 50),
        )
        if response.status_code != 200:
            # raise requests.ConnectionError(f"响应异常代码{response.status_code}")
            data = self.get_reindex(DataFrame(), "旅客")
        else:
            data = DataFrame.from_records(response.json().get("data").get("list"))
            data = self.get_reindex(data, "旅客")
            data[self.parkRegion] = data[self.gateNo].map(
                lambda x: self.APRONGATE.get(x, x[0] + "指廊"), IGNORE
            )
        if "get_data" in self.RUNNING:
            self.progress.set(self.progress.get() + progress)
        return {"旅客_旅客": data.loc[data[self.atot].isna()]}

    def get_session(self) -> requests.Session:
        user, pw = self.ask_user()
        if datetime.now().timestamp() - self.login_timestamp > 660:
            if self.session:
                self.session.close()
            session = requests.Session()
            login_payload = {
                "account": user,
                "passWord": pw,
                "captcha": "",
                "pcOrApp": "pc",
            }
            response = session.post(
                self.URLS.get("登录"),
                data=dumps(login_payload),
                headers=self.HEADER,
                timeout=10,
            ).json()
            if response.get("code") != 1:
                message = response.get("message")
                raise (requests.RequestException if self.RUNNING else GuiError)(message)
            else:
                try:
                    self.username = response.get("data").get("userData").get("userName")
                except Exception:
                    pass
            self.session = session
            self.login_timestamp = datetime.now().timestamp()
        return self.session

    def update_data(self) -> dict[str, DataFrame]:
        limit = self.UPDATE_LIMIT.get()
        datetime_now = self.datetime_now()
        yesterday = self.yesterday.get()
        today = self.today(-1 if yesterday else 0)
        end = timedelta(hours=self.UPDATE_RANGE_.get())
        if not end:
            start = timedelta(
                hours=(
                    12 if datetime_now.hour <= 3 else 6 if datetime_now.hour <= 6 else 0
                )
            )
            end = timedelta(hours=2 if datetime_now.hour <= 6 else 3)
        else:
            start = timedelta(hours=self.UPDATE_RANGE.get())
        try:
            futures = []
            self.get_session()
            self.update_status({"更新数据": "进行中"})
            self.status_bar.config(background="skyblue")
            data = dict()
            with ThreadPoolExecutor() as executor:
                self.progress.set(0)
                futures.append(
                    executor.submit(
                        self.get_flight_info, today - start, datetime_now + end
                    )
                )
                futures.append(
                    executor.submit(
                        self.get_flight_data,
                        today - start,
                        datetime_now + end,
                        departureMode="outS",
                        limit=limit,
                        moni=self.msg_para[0].get() or self.msg_para[1].get(),
                    )
                )
                futures.append(
                    executor.submit(
                        self.get_delay_data,
                        today - start,
                        datetime_now + end,
                        limit=limit,
                    )
                )
                futures.append(executor.submit(self.get_passenger_info, limit=limit))
                wait(futures, max(20, limit // 50))
                self.progress.set(100)
            for future in futures:
                if exception := future.exception():
                    raise requests.RequestException(*exception.args)
                else:
                    data.update(future.result())

            """
            data.update(self.get_flight_info(today - start, datetime_now + end))
            data.update(self.get_flight_data(
                        today - start,
                        datetime_now + end,
                        departureMode="outS",
                        limit=limit,))
            data.update(self.get_delay_data(
                        today - start,
                        datetime_now + end,
                        limit=limit,))
            data.update(self.get_passenger_info(limit=limit))
            """

            self.progress.set(0)
            return data
        finally:
            self.update_status({"更新数据": ""})

    def get_all(self, *args: datetime, **kwargs) -> dict[str, DataFrame]:
        limit = kwargs.pop("limit", self.DATA_LIMIT.get())
        futures = []
        self.get_session()
        with ThreadPoolExecutor() as executor:
            self.progress.set(0)
            futures.append(executor.submit(self.get_flight_info, *args))
            futures.append(
                executor.submit(
                    self.get_flight_data,
                    *args,
                    departureMode="outS",
                    limit=limit,
                    moni=False,
                )
            )
            futures.append(executor.submit(self.get_delay_data, *args, limit=limit))
            futures.append(executor.submit(self.get_tmi_info, *args, limit=limit))
            wait(futures, max(20, limit // 50))
            self.progress.set(100)
        data = {}
        for future in futures:
            if exception := future.exception():
                raise exception
            else:
                data.update(future.result())
        self.progress.set(0)
        return data

    def get_data(
        self,
        *__name: str,
        isna: Iterable[str] = [],
        notna: Iterable[str] = [],
        **kwargs,
    ) -> DataFrame:
        self.lock.acquire()
        # data expired or get data by scheduler
        if (
            __name and datetime.now().timestamp() - self.update_timestamp.get() > 120
        ) or not __name:
            self.refresh(False)
            running = "get_data" in self.RUNNING
            if running:
                if i := self.RUNNING["get_data"][0]:
                    self.after_cancel(i)
                self.title(f"{TITLE} - 自动检测中")

            # 自动选择信息与状态（大面积航延与昨日续报）
            now = self.datetime_now()
            if self.AUTOSEL.get() and not self.HISTORY["RATES"].empty:
                if self.HISTORY["RATES"]["启动标准"].iloc[-3:].all():
                    self.msg_para[2].set(True)
                    if now.hour <= 4 and not self.yesterday.get():
                        self.yesterday.set(True)
                elif 4 <= now.hour <= 5:
                    if (
                        not self.HISTORY["RATES"].iloc[-1]["延误未起飞"]
                        and self.msg_para[2].get()
                    ):
                        self.msg_para[2].set(False)
                    if self.yesterday.get():
                        self.yesterday.set(False)

            # 数据更新
            self.DATA.update(self.update_data())
            now = datetime.now()
            self.update_timestamp.set(now.timestamp())
            self.refresh(True)
            if running:
                # 更新UI和大面积航延
                delay = "outLastTot"  # self.DELAY.get()
                alert = self.get_alert()
                alert_level, alert_by = alert.pop("大面积航延"), alert.pop("启动标准")
                self.title(
                    "{} - {} (更新于{})".format(
                        TITLE,
                        (
                            "正常运行中"
                            if alert_level == "无"
                            else f"大面积航延{alert_level}"
                        ),
                        now.strftime(r"%H:%M"),
                    )
                )

                # 计算效率指标
                rates = list(self.get_rates())
                today = self.today()
                data = self.DATA["执行_离港"]
                data = data.loc[
                    data["flightDate"].astype(str) == str(today)[:19]
                ].copy()
                data.drop(
                    data.loc[data["operationStatusCode"] == "CNCL"].index,
                    inplace=True,
                )
                rates.append(data.loc[data["atot"].notna()].__len__())
                rates.append(data.loc[data["atot"].isna()].__len__())

                data = self.DATA["执行_进港"]
                data = data.loc[
                    data["flightDate"].astype(str) == str(today)[:19]
                ].copy()
                data.drop(
                    data.loc[data["operationStatusCode"] == "CNCL"].index,
                    inplace=True,
                )
                rates.insert(-1, data.loc[data["aldt"].notna()].__len__())
                rates.append(data.loc[data["aldt"].isna()].__len__())

                data = self.DATA["航班_航班"]
                rates.append(
                    data
                    # .loc[data["inAldt"].notna() | data["inSldt"].isna()]
                    .loc[data["outAtot"].isna() & (data["outEstripStatus"] != "DEP")]
                    .loc[data["outCtot"] - data[delay] > timedelta()]
                    .__len__()
                )

                data = self.DATA["延误_延误"]
                if self.EXCLUDE.get():
                    data = data.loc[data["inAldt"].notna() | data["inSldt"].isna()]
                rates.append(
                    data.loc[data["outAtot"].isna()]
                    .loc[now - data[delay] > timedelta()]
                    .__len__()
                )
                rates += [
                    (
                        "维持"
                        if self.msg_para[2].get() and not alert_by
                        else alert_level
                    ),
                    alert_by,
                ]
                if self.rates_bind[0]:
                    self.rates_bind[2]("<Leave>", self.rates_bind[0])
                rates.append(self.fourlocation_tenvenuse())
                self.HISTORY["RATES"].loc[now] = rates

                # 计划下次更新
                next_update = self.get_next_update(
                    self.UPDATE_INTERVAL.get(),
                    self.UPDATE_OFFSET.get(),
                )
                if i := self.RUNNING.get("get_data"):
                    i[0] = self.after(
                        next_update, threading.Thread(target=self.get_data).start
                    )
                self.update_log(
                    "数据更新完毕，{}".format(
                        (
                            f"已达大面积航班延误{alert_level}：{alert_by}，"
                            if alert_by
                            else ""
                        )
                        + f"{self.min_sec_format(next_update // 1000)}后开始下次更新"
                    )
                )

                # 自动选择信息与状态（冰霜天气）
                data = self.DATA["航班_航班"]
                if (
                    self.AUTOSEL.get()
                    and not (self.msg_para[0].get() or self.msg_para[1].get())
                    and data.loc[
                        (data["outSobt"] >= now - timedelta(hours=1))
                        & (data["outSobt"] <= now + timedelta(hours=2)),
                        "outDeicingType",
                    ].any()
                ):
                    self.msg_para[0].set(True)

                # 效率指标刷新
                for i, j in enumerate(self.HISTORY["RATES"].columns[:-2]):
                    if "执行" in j or "取消" in j:
                        continue
                    elif "大面积" in j:
                        rates[i] = {
                            "text": rates[i][:2],
                            "foreground": "black" if rates[i] == "无" else "white",
                            "background": RATES_COLOR.get(rates[i][:2], "white"),
                            "font": (
                                ("微软雅黑", 10)
                                if not alert_by
                                else ("微软雅黑", 10, "bold")
                            ),
                            "width": 4,
                        }
                    else:
                        data = self.HISTORY["RATES"]
                        data = data.loc[data.index < now]
                        data = data.loc[data.index >= now - timedelta(hours=1)][j]
                        if i < 4:
                            if rates[i]:
                                data = data.map(
                                    lambda x: (
                                        float(x[:-1])
                                        if isinstance(x, str) and r"%" in x
                                        else np.nan
                                    )
                                )
                                rates[i] = self.rates_handler(
                                    "正常性",
                                    rates[i],
                                    float(rates[i][:-1]),
                                    np.round(data.mean(), 2),
                                )
                            else:
                                rates[i] = {"text": "  -  "}
                        else:
                            rates[i] = self.rates_handler(
                                j,
                                rates[i],
                                rates[i],
                                np.round(data.mean()),
                                width=4,
                            )
                self.rates_bind[0] = self.rates_bind[1](
                    "<Leave>",
                    lambda *x: self.rates_action(
                        *rates[:-2],
                        font=("Consolas", 11),
                        foreground="black",
                        width=3,
                    ),
                )
                self.rates_action(
                    *rates[:-2], font=("Consolas", 11), foreground="black", width=3
                )

                threading.Thread(target=self.save_history).start()

        self.lock.release()
        if __name:
            returns = []
            for __name in __name:
                data = self.DATA[__name if "_" in __name else __name + "_" + __name]
                for key, value in kwargs.items():
                    if isinstance(value, (tuple, list)) and key in data:
                        data = (
                            data.loc[data[key].isin(value)]
                            if value
                            else DataFrame(columns=data.columns)
                        )
                    elif key in data:
                        data = data.loc[data[key] == i]
                for key in isna:
                    data = data.loc[data[key].isna()]
                for key in notna:
                    data = data.loc[data[key].notna()]
                returns.append(data.copy())
            return returns[0] if len(returns) == 1 else tuple(returns)
        elif "sync" in self.RUNNING:
            self.PATH.set(self.AUTO_PATH.get() + "/sync")
            output = self.update_info(*self.all_schedule)
            output["数据日期时间"] = int(self.update_timestamp.get() * 1000)
            output.update(self.HISTORY["RATES"].iloc[-1].to_dict())
            output.update(alert)

            retries = 5
            while retries > 0:
                try:
                    with open(f"{self.AUTO_PATH.get()}/sync.json", "w") as json:
                        dump(output, json)
                    break
                except Exception as exception:
                    retries -= 1
                    if retries <= 0:
                        exception = repr(exception)
                        self.update_log(
                            f"信息同步失败 ({exception[: exception.find('(')]})", "warn"
                        )

    def update_info(self, *args: Callable):
        output = dict()
        if not args:
            if not self.INFO_PATH_.get():
                self.INFO_PATH_.set(True)
                self.update_path("INFO")
            if self.INFO_PATH_.get():
                args = (
                    self.auto_msg,
                    self.stock_msg,
                    (self.get_ctot, 0),
                    self.long_delay,
                    (self.long_delay, 1),
                )
            else:
                self.showinfo("选择导出路径以更新所有信息")

        """
        with ThreadPoolExecutor() as executor:
            names = [(i[0]if isinstance(i, Iterable) else i).__name__ for i in args]
            futures = [executor.submit(*i if isinstance(i, Iterable) else (i, )) for i in args]
            wait(futures, 30)

        output = dict()
        for name, future in zip(names, futures):
            if future.exception():
                try:
                    future.result()
                except Exception as exception:
                    name = self.func_names.get(name, name)
                    exception, tb = repr(exception), format_exc().split("\n", 1)[1]
                    self.update_log(
                        f"{name}自动生成出错 ({exception[: exception.find('(')]})\n{tb}", "warn"
                    )
                    self.update_status(
                        warning={f"{name}出现异常": exception[: exception.find("(")]}
                    )
            else:
                output.update(future.result())
        """

        for i in args:
            try:
                output.update(**(i[0](*i[1:]) if isinstance(i, Iterable) else i()))
            except Exception as exception:
                if isinstance(i, Iterable):
                    i = i[0]
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                name = self.func_names.get(i.__name__, i.__name__)
                self.update_log(f"{name}自动生成出错\n{tb}", "warn")
                self.update_status(
                    warning={f"{name}出现异常": exception[: exception.find("(")]}
                )
        return output

    def get_info(self, *args: Callable, **kwargs):
        if "get_info" in self.RUNNING:
            now, interval = datetime.now().timestamp(), kwargs.get("interval")
            if i := self.RUNNING.get("get_info", dict()).get(interval):
                self.after_cancel(i)

            path = self.AUTO_PATH.get()
            if self.AUTO_T.get():
                path += r"/{}"
            if self.AUTO_D.get():
                path += f"/{self.datetime_now().strftime(r'%Y-%m-%d')}"
            self.PATH.set(path)
            if now - self.update_timestamp.get() < max(interval / 2, 2) * 60:
                output = self.update_info(*args)
                interval_ = self.get_next_update(
                    interval, self.UPDATE_OFFSET.get() - 10
                )
            else:
                interval_ = 5000
            if i := self.RUNNING.get("get_info"):
                i[interval] = self.after(
                    interval_,
                    threading.Thread(
                        target=self.get_info, args=args, kwargs=kwargs
                    ).start,
                )
        return output

    def tmi_content_extract(self, __str: str):
        result = [i for i in self.AIRPORT.values() if i in __str]
        result += [
            self.AIRPORT.get(i[:2], "国际/地区")
            for i in findall(r"[A-Z\?]+", __str)
            if len(i) == 4
        ]
        return " ".join(np.unique(result)) if result else ""

    def history_summary(self, row: Series, data: Series, detailed: int = 0) -> str:
        status = {
            "outPushTime": "滑行后",
            "moniJob.tract_D.actBeginTime": "推出后",
            "outAcct": "关舱后",
            "outAebt": "登结后",
            "outAsbt": "登机前",
        }
        v_, changes = None, []
        if detailed:
            for k, v in row.items():
                if notna(v) and v != v_:
                    for i, j in status.items():
                        i = data[i]
                        if i < k:
                            i = (k - i).seconds // 60
                            break
                    changes.append(
                        "{:02d}:{:02d}{}{:02d}:{:02d}".format(
                            k.hour,
                            k.minute,
                            "为" if v_ is None else "推迟到" if v_ < v else "提前到",
                            v.hour,
                            v.minute,
                        )
                    )
                    v_ = v
            return (
                "\n".join(changes[0 if detailed < 0 else -detailed :])
                if len(changes) > 1
                else "始终" + changes[0][-6:] if changes else "未开始记录"
            )
        else:
            for k, v in reversed(status.items()):
                if notna(data[k]):
                    i = row.loc[
                        row.index < data[k] if "前" in v else row.index > data[k]
                    ].loc[row.notna()]
                    if len(i):
                        i = i.iloc[-1 if "前" in v else 0]
                        if v_ and v_ == i:
                            continue
                        changes.append(
                            "{}{}{:02d}:{:02d}".format(
                                v,
                                (
                                    (
                                        "推迟到"
                                        if v_ < i
                                        else "提前到" if v_ > i else "保持在"
                                    )
                                    if v_
                                    else "为"
                                ),
                                i.hour,
                                i.minute,
                            )
                        )
                        v_ = i
            if changes:
                if row.iloc[-1] != v_:
                    changes.append(self.history_summary(row, data, 1))
                return "\n".join(changes[0 if detailed < 0 else -detailed :])
            elif row.notna().any():
                return self.history_summary(row, data, 3)
            else:
                return "未开始记录"

    def is_exp(self, airline: str, airports: str):
        return any(r in self.EXPRESS.get(airline, "") for r in str(airports).split("-"))

    @staticmethod
    def font(**kwargs):
        if "name" not in kwargs:
            kwargs["name"] = "等线"
        if not ("size" in kwargs or "sz" in kwargs):
            kwargs["size"] = 11
        return Font(**kwargs)

    def get_reindex(self, data: DataFrame, data_type: str):
        def vget(item: dict[str, dict], v: list):
            try:
                while v:
                    item = item.get(v.pop(0))
            except Exception:
                return None
            return item

        data = data.reindex(
            columns=list(i.split(".")[0] for i in self.TAG_MAP[data_type].keys())
        )
        for v in self.TAG_MAP[data_type]:
            splitted = set()
            if "." in v:
                v_ = v.split(".")
                splitted.add(v_[0])
                data[v] = data[v_[0]].map(lambda x: vget(x, v_[1:]), IGNORE)
            if v in self.CONVERT and v in data:
                data[v] = (
                    data[v].map(datetime.fromisoformat, IGNORE).astype("datetime64[s]")
                )
            if v.endswith("AirportRegion"):
                data[v + "Cn"] = self.INTL
                i = data[v.replace("AirportRegion", "CountryCode")] == "D"
                data.loc[i, v + "Cn"] = data.loc[i, v].map(self.REGION, IGNORE)
            while splitted:
                data.drop(columns=splitted.pop(), inplace=True)

        return data

    def format_excel(self, file: str, **kwargs: DataFrame):
        writer = ExcelWriter(file, mode="w")
        datetime_format = self.CONVERT_.get()
        convert = not datetime_format == "标准字符串"
        datetime_width = self.DATETIME_FORMAT.get(datetime_format, 19)
        for dataname, dataframe in kwargs.items():
            dataname, sheetname = dataname.split("_")
            tag_map = self.TAG_MAP[dataname]
            dataframe.dropna(axis=1, how="all", inplace=True)
            if not convert:
                for i in dataframe.columns:
                    if i in self.CONVERT:
                        dataframe[i] = dataframe[i].map(str).replace("NaT", "")
                    elif i in self.TIMEDELTA:
                        dataframe[i] = (
                            dataframe[i].map(
                                lambda x: np.int64(x.total_seconds()), IGNORE
                            )
                            // 60
                        )
            dataframe.index += 1
            columns = dataframe.columns.values
            dataframe.rename(columns=tag_map).to_excel(
                writer, sheet_name=sheetname, index_label="序号", freeze_panes=(1, 1)
            )
            sheet = writer.sheets[sheetname]
            max_row = sheet.max_row + 1
            for i, j in enumerate(columns):
                col = get_column_letter(i + 2)
                if j in self.CONVERT and convert:
                    sheet.column_dimensions[col].width = datetime_width
                    for row in range(2, max_row):
                        cell = sheet[f"{col}{row}"]
                        cell.number_format = datetime_format
                        cell.border = BORDER
                elif j in self.TIMEDELTA and convert:
                    for row in range(2, max_row):
                        cell = sheet[f"{col}{row}"]
                        if isnumeric(cell.value) and cell.value >= 0:
                            cell.number_format = "H:MM"
                        else:
                            cell.value = "-"
                        cell.border = BORDER
                elif j in self.WRAPTEXT:
                    sheet.column_dimensions[col].width = self.WRAPTEXT[j]
                    for row in range(2, max_row):
                        cell = sheet[f"{col}{row}"]
                        cell.border = BORDER
                        cell.alignment = ALIGN_LEFT
                elif isinstance(j, datetime):
                    for row in range(1, max_row):
                        cell = sheet[f"{col}{row}"]
                        cell.number_format = "HH:MM"
                        cell.border = BORDER
                else:
                    for row in range(2, max_row):
                        sheet[f"{col}{row}"].border = BORDER
        writer.close()
        self.update_log(f"文件导出成功：{file}", "file")

    def handle_exception(self, *args):
        if isinstance(args[0], threading.ExceptHookArgs):
            args = args[0][:3]
        try:
            if args[0].__name__ in self.RUN_EXCEPTION:
                self.showinfo(f"{self.RUN_EXCEPTION[args[0].__name__]} ({args[1]})")
            else:
                status = "出现异常：{} ({})\n".format(args[1], args[0].__name__)
                self.update_status(warning={"出现异常": args[0].__name__})
                self.update_log(status + "".join(format_tb(args[2]))[:-1], "warn")
            if self.exists.get():
                self.refresh()
                if args[0].__module__.startswith("requests"):
                    self.login_timestamp = 0
                    self.username = ""
                    if "get_data" in self.RUNNING:
                        self.terminate(*args)
                        self.status_bar.config(background="salmon")
                elif args[0].__name__ == "Terminates":
                    self.terminate(*args)
            self.lock.release()
        except RuntimeError:
            ...
        except Exception:
            ...

    @staticmethod
    def wrap_iterstr(
        __iterstr: Iterable[str], __braces: str = "【{}】", __comma: str = "；"
    ):
        return __braces.format(__comma.join(__iterstr)) if __iterstr else ""

    def update_status(
        self,
        ongoing: dict[str, str] = {},
        warning: dict[str, str] = {},
        reset: bool = False,
        set_bar: bool = False,
        default: str = "执行完毕",
    ):
        status, exists = "", self.exists.get()
        if reset:
            self.ongoing.clear()
            self.warning.clear()
            status = "状态栏已重置｜"
        if set_bar and exists and self.progress_bar["mode"] != "indeterminate":
            self.progress_bar.config(mode="indeterminate")
            self.progress_bar.start()
        for k, v in ongoing.items():
            if v:
                self.ongoing[k] = v
            else:
                self.ongoing.pop(k, None)
        warning["应用设置改变"] = "，".join(
            i[1]
            for i in (
                (self.DELAY.get() == "outStot", "表格按起飞延误"),
                (self.yesterday.get(), "生成昨日续报信息"),
                (self.NOW_.get(), f"当前时间更改为{self.NOW.get()}"),
                ("W/Z" not in self.CTOTTYPE, "CTOT表格航班性质不含正班"),
                (
                    self.AUTODELAY.get() and self.DELAYBY["均判定为本场天气"],
                    "延误原因均自动判定为本场天气",
                ),
                (
                    self.AUTODELAY.get() and self.DELAYBY["均判定为本场军事活动"],
                    "延误原因均自动判定为本场军事活动",
                ),
            )
            if i[0]
        )
        for k, v in warning.items():
            if v:
                self.warning[k] = (
                    (
                        (i if v in i else f"{i}，{v}")
                        if (i := self.warning.get(k, ""))
                        else v
                    )
                    if "异常" in k
                    else v
                )
            else:
                self.warning.pop(k, None)
        if exists:
            self.status_bar.config(
                background="palegreen" if self.RUNNING else "#f0f0f0"
            )
            if self.ongoing:
                status += "当前执行：" + "，".join(
                    "{}（{}）".format(*i) for i in self.ongoing.items()
                )
            else:
                status += default
                self.progress_bar.stop()
                self.progress.set(0)
                self.progress_bar.config(mode="determinate")
            if self.warning:
                status += "｜" + "，".join(
                    "{}（{}）".format(*i) for i in self.warning.items()
                )
                self.status_bar.config(
                    background=(
                        ("yellow" if self.RUNNING else "gold")
                        if len(self.warning) == 1 and "应用设置改变" in self.warning
                        else "orange"
                    )
                )
            self.status.set(status)
        return tuple(self.ongoing.keys()), tuple(self.warning.keys())

    def update_log(self, __str: str, *tag: str, end: str = "\n"):
        now = str(datetime.now())[:19]
        suffix = [
            i[1]
            for i in (
                (self.yesterday.get(), "昨日续报"),
                (self.NOW_.get(), f"当前时间更改为{self.NOW.get()}"),
                ("W/Z" not in self.CTOTTYPE, "CTOT表格航班性质不含正班"),
            )
            if i[0]
        ]
        if self.exists.get():
            try:
                with open(LOG, "a", encoding="UTF-8") as output:
                    output.write(
                        "{}  {}{}{}".format(now, __str, self.wrap_iterstr(suffix), end)
                    )
            except PermissionError:
                suffix.append("无法写入日志文件")
            self.log.config(state=tk.NORMAL)
            suffix = self.wrap_iterstr(suffix)
            self.log.insert(
                "0.0",
                "{}  {}{}{}".format(now[11:], __str, suffix, end),
            )

            line = __str.count("\n")
            last = len(__str.split("\n")[-1]) if line else 10 + len(__str)
            for tag in tag:
                if "background" in self.log_tag[tag]:
                    self.log.tag_add(tag, "1.0", f"{2 + line}.0")
                else:
                    self.log.tag_add(tag, "1.10", f"{1 + line}.{last}")
                if tag == "text" and "成功" in __str:
                    self.log.tag_add("bold", "1.10", f"1.{10 + __str.find('成功') + 2}")
            if suffix:
                self.log.tag_add(
                    "suffix", f"{1 + line}.{last}", f"{1 + line}.{len(suffix) + last}"
                )
            self.log.tag_add("time", "1.0", "1.9")

            if limit := self.LOGLIMIT.get():
                self.log.delete(f"{limit}.0", tk.END)
            self.log.config(state=tk.DISABLED)
        else:
            suffix.append("程序已关闭")
            try:
                with open(LOG, "a", encoding="UTF-8") as output:
                    output.write(
                        "{}  {}{}{}".format(now, __str, self.wrap_iterstr(suffix), end)
                    )
            except Exception:
                ...

    def showinfo(self, __str: str, __title: str = TITLE):
        if self.exists.get() or self.wait_visibility():
            threading.Thread(
                target=messagebox.showinfo, args=(__title, __str), daemon=True
            ).start()

    def showwarning(self, __str: str, __title: str = TITLE):
        if self.exists.get() or self.wait_visibility():
            threading.Thread(
                target=messagebox.showwarning, args=(__title, __str), daemon=True
            ).start()

    def datetime_now(self):
        if self.NOW_.get():
            return datetime.fromisoformat(self.NOW.get())
        else:
            now = list(datetime.now().timetuple()[:5])
            if self.TIMESLOT.get():
                if now[3] == 23 and now[4] >= 58:
                    now[3], now[4] = 23, 59
                else:
                    now[4] = (now[4] + 2) // 5 * 5
                    if now[4] >= 60:
                        now[3] += 1
                        now[4] = 0
            return datetime(*now)

    def today(self, days: int = 0, **kwrags):
        return datetime(*self.datetime_now().timetuple()[:3]) + timedelta(
            days=days, **kwrags
        )

    @staticmethod
    def min_sec_format(__s, text_prefix=False, wrap_hour=False):
        __m, __s = np.divmod(__s, 60)
        prefix = ""
        if np.isnan(__m) or np.isnan(__s):
            return prefix
        if __m >= 0:
            if text_prefix:
                prefix = "提前"
        else:
            if __s:
                __m = abs(__m + 1)
                __s = 60 - __s
            else:
                __m = abs(__m)
            prefix = "晚" if text_prefix else "-"

        __m, __s = int(__m), round(__s)
        if __m >= 60 and wrap_hour:
            __h, __m = divmod(__m, 60)
            return prefix + (f"{__h}时{__m}分" if __m else f"{__h}小时")
        elif __m and not __s:
            return prefix + f"{__m}分钟"
        elif not __m and __s:
            return prefix + f"{__s}秒"
        else:
            return (
                "准时" if text_prefix and not __m + __s else prefix + f"{__m}分{__s}秒"
            )

    def airport_name(self, arg: Series):
        s, r = [], ("国际", "机场", "（")
        for i, j in zip(str(arg.iloc[0]).split("-"), str(arg.iloc[1]).split("-")):
            if i in self.AIRPORTNAME:
                j = self.AIRPORTNAME[i]
            elif j:
                for _r in r:
                    if _r in j:
                        j = j[: j.find(_r)]
                        break
            s.append(j if j else i)
        return "-".join(s)

    def save_push(self, type_: str, msg: str, img: str = ""):
        self.PUSH.loc[f"{datetime.now().strftime(r'%H:%M:%S')} {type_}"] = [
            msg.replace("\n", ""),
            img,
        ]
        if len(self.PUSH) >= 20:
            self.PUSH.drop(index=self.PUSH.index[:-20], inplace=True)

    def save_textfile(self, __string: str, path: str = None):
        if path:
            if not os.path.exists(path):
                os.makedirs(path)
            textfile = "{}/{}".format(
                path if path else self.INFO_PATH.get(),
                self.FILENAME.get("文本类型信息").format(*datetime.now().timetuple()),
            )
            suffix = [
                i[1]
                for i in (
                    (self.yesterday.get(), "昨日续报"),
                    (self.NOW_.get(), f"当前时间更改为{self.NOW.get()}"),
                )
                if i[0]
            ]
            try:
                with open(textfile, "a", encoding="UTF-8") as output:
                    output.write(
                        f"{str(datetime.now())[:19]}\n{__string}{self.wrap_iterstr(suffix)}\n\n"
                    )
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"文本类型文件{textfile}写入失败 ({exception[: exception.find('(')]})\n{tb}",
                    "text",
                )

    def get_history_cal(self, start: datetime, end: datetime):
        try:
            self.update_status({"导出历史记录": "获取数据"}, set_bar=True)
            data = (
                self.get_flight_data(start, end, 0, departureMode="outS")["航班_航班"]
                .rename(columns={"outGuid": "guid"})
                .set_index("guid", drop=True)
            )
            self.update_status({"导出历史记录": "保存中"})
            reindexer = [
                "outFlightNo",
                "inAldt",
                "inAibt",
                "inAeot",
                "ttt",
                "outRoute",
                "outRouteCn",
                "outAirportRegionCn",
                "outFlightTypeCode",
                "portNoDp",
                "outGateNo",
                "outSobt",
                "outCobt",
                "outTobt",
                "outCtot",
                "outLastTot",
                "outAtot",
                "outAsbt",
                "outAebt",
                "outAcct",
                "moniJob.tract_D.actBeginTime",
                "outPushTime",
                "outVipInd",
                "outEstripStatus",
                "outTmi",
            ]
            output = dict(
                (
                    "航班_" + i,
                    concat(
                        [
                            data.reindex(columns=reindexer),
                            self.HISTORY[i].reindex(index=data.index),
                        ],
                        axis=1,
                    )
                    .reset_index(drop=True)
                    .dropna(axis=1, how="all"),
                )
                for i in ("CTOT", "COBT")
            )
            title, filename = (
                "导出CTOT和COBT历史记录表格",
                self.FILENAME.get("CTOT和COBT历史记录").format(
                    *datetime.now().timetuple()
                ),
            )
            file = (
                f"{self.EXPORT_PATH.get()}/{filename}"
                if self.EXPORT_PATH_.get()
                else filedialog.asksaveasfilename(
                    filetypes=(("Xlsx表格文件", "*.xlsx"),),
                    confirmoverwrite=True,
                    parent=self,
                    title=title,
                    initialdir=self.EXPORT_PATH.get(),
                    initialfile=filename,
                )
            )
            self.save_excel(self, title, file, self.format_excel, **output)
        finally:
            self.update_status({"导出历史记录": ""})

    def get_history_rates(self, file: str, data: DataFrame):
        with ExcelWriter(file, mode="w") as wb:
            data["date"] = data.index.map(lambda x: x.strftime(r"%m-%d"))
            for dataname, dataframe in data.groupby("date", sort=False):
                dataframe.drop("date", axis=1).to_excel(
                    wb, sheet_name=dataname, index_label="时间"
                )
                ws = wb.sheets[dataname]
                ws.column_dimensions["A"].width = 6
                ws.column_dimensions["O"].width = 80

                for col in range(1, ws.max_column + 1):
                    cell = ws[f"{get_column_letter(col)}1"]
                    cell.font = self.font(bold=True)
                    cell.alignment = ALIGN_CENTER

                for row in range(2, ws.max_row + 1):
                    cell = ws[f"A{row}"]
                    cell.font = self.font(bold=True)
                    cell.alignment = ALIGN_CENTER
                    cell.number_format = "HH:MM"

                for col in range(2, 6):
                    col = get_column_letter(col)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f"{col}{row}"]
                        cell.border = BORDER
                        cell.font = self.font()
                        cell.alignment = ALIGN_CENTER_CENTER
                        if r"%" in cell.value:
                            cell.value = float(cell.value[:-1]) * 0.01
                            cell.number_format = r"0.00%"

                for col in range(6, ws.max_column + 1):
                    col = get_column_letter(col)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f"{col}{row}"]
                        cell.border = BORDER
                        cell.font = self.font()
                        cell.alignment = ALIGN_CENTER_CENTER
        self.update_log(f"文件导出成功：{file}", "file")

    def pivot_export(self):
        departure_mode = {
            "计划起飞STD": "outSobt",
            "计划落地STA": "inSibt",
            "实际起飞ATD": "outAtot",
            "实际落地ATA": "inAldt",
        }

        def run(
            start: datetime,
            end: datetime,
            data: str,
            mode: Callable,
            sep: int,
            condition: tuple[
                str,
                Callable,
                str,
                Callable,
                timedelta,
                Callable,
                timedelta,
                str,
                Callable,
                str,
                Callable,
                timedelta,
                Callable,
                timedelta,
            ],
            fillna: tuple[str, str, str, str],
            name: str,
            range_: str,
        ):
            def save(
                file: str,
                pivot: DataFrame,
                data: DataFrame,
                renamer: dict[str, str],
                name: str,
            ):
                renamer[name] = name
                data = data.reindex(columns=renamer.keys())
                datetime_format = self.CONVERT_.get()
                convert = not datetime_format == "标准字符串"
                datetime_width = self.DATETIME_FORMAT.get(datetime_format, 19)
                if not convert:
                    for i in data.columns:
                        if i in self.CONVERT:
                            data[i] = data[i].map(str).replace("NaT", "")
                        elif i in self.TIMEDELTA or i == name:
                            data[i] = (
                                data[i].map(
                                    lambda x: np.int64(x.total_seconds()), IGNORE
                                )
                                // 60
                            )
                with ExcelWriter(file, mode="w") as wb:
                    pivot.to_excel(wb, sheet_name="数据透视", freeze_panes=(3, 2))
                    ws = wb.sheets["数据透视"]
                    ws["A1"].value = rf"项目\{name}"
                    ws.column_dimensions["A"].width = 4.5
                    ws.column_dimensions["B"].width = 12.5
                    max_col = get_column_letter(ws.max_column)
                    max_col_ = get_column_letter(ws.max_column - 1)
                    max_row = ws.max_row + 1

                    for col in range(1, ws.max_column + 1):
                        cell = ws[f"{get_column_letter(col)}1"]
                        cell.font = self.font(bold=True)
                        cell.alignment = ALIGN_CENTER

                    for row in range(2, ws.max_row + 1):
                        for col in "A", "B":
                            cell = ws[f"{col}{row}"]
                            cell.font = self.font(bold=True)
                            cell.alignment = ALIGN_CENTER

                    for col in range(2, ws.max_column):
                        col = get_column_letter(col + 1)
                        for row in range(2, max_row):
                            cell = ws[f"{col}{row}"]
                            cell.border = BORDER
                            cell.font = self.font()

                    for i in range(3, ws.max_column):
                        col = get_column_letter(i)
                        ws[f"{col}3"].value = f"={col}2/SUM($C2:${max_col_}2)"
                        ws[f"{col}3"].number_format = r"0.00%"
                    for row in range(2, max_row):
                        cell = ws[f"{max_col}{row}"]
                        cell.value = f"=SUM(C{row}:{col}{row})"
                        cell.font = self.font(bold=True)
                    ws[f"{max_col}3"].number_format = r"0.00%"

                    for i in 1, 2, 3:
                        ws.merge_cells(f"A{i}:B{i}")

                    columns = data.columns.values
                    data.rename(columns=renamer).to_excel(
                        wb,
                        sheet_name="原始数据",
                        index_label="序号",
                        freeze_panes=(1, 1),
                    )

                    ws = wb.sheets["原始数据"]
                    max_row = ws.max_row + 1
                    for i, j in enumerate(columns):
                        col = get_column_letter(i + 2)
                        if j in self.CONVERT and convert:
                            ws.column_dimensions[col].width = datetime_width
                            for row in range(2, max_row):
                                cell = ws[f"{col}{row}"]
                                cell.number_format = datetime_format
                        elif j in self.TIMEDELTA or j == name and convert:
                            for row in range(2, max_row):
                                cell = ws[f"{col}{row}"]
                                if isnumeric(cell.value) and cell.value >= 0:
                                    cell.number_format = "H:MM"
                                else:
                                    cell.value = "-"

            try:
                self.update_status({"数据透视": "进行中"}, set_bar=True)
                range_ = departure_mode[range_]
                prefix = range_[:-4]
                renamer = self.TAG_MAP[data]
                data = self.GET_DATA[data](
                    start, end, limit=self.DATA_LIMIT.get(), departureMode=f"{prefix}S"
                )[f"{data}_{data}"]

                fillnas = {
                    "替换为当前时间": datetime.now(),
                    "替换为导出时段开始时间": start,
                    "替换为导出时段结束时间": end,
                }
                var = [data[condition[0]], data[condition[2]]]
                if fillna[0]:
                    var[0] = var[0].fillna(
                        fillnas[fillna[0]] if fillna[0] in fillnas else data[fillna[0]]
                    )
                if fillna[1]:
                    var[1] = var[1].fillna(
                        fillnas[fillna[1]] if fillna[1] in fillnas else data[fillna[1]]
                    )

                data[name] = condition[3](condition[1](*var), condition[4])
                result = condition[5](data[name], condition[6])
                if mode:
                    var = [data[condition[7]], data[condition[9]]]
                    if fillna[2]:
                        var[0] = var[0].fillna(
                            fillnas[fillna[2]]
                            if fillna[2] in fillnas
                            else data[fillna[2]]
                        )
                    if fillna[3]:
                        var[1] = var[1].fillna(
                            fillnas[fillna[3]]
                            if fillna[3] in fillnas
                            else data[fillna[3]]
                        )
                    result = mode(
                        result,
                        condition[12](
                            condition[10](condition[8](*var), condition[11]),
                            condition[13],
                        ),
                    )
                data = data.loc[result].reset_index(drop=True).copy()
                if not len(data):
                    messagebox.showinfo(
                        TITLE,
                        "数据透视筛选后无数据，请检查筛选条件和时间范围",
                        parent=box,
                    )
                    return 1
                data.index += 1
                data[name + "序数"] = (
                    data[name]
                    .map(lambda x: np.int64(x.total_seconds()))
                    .map(lambda x: x // (3600 * sep))
                    .astype(float)
                )
                data["航司"] = data["airlineIata"]
                data["区域"] = data[f"{prefix}AirportRegionCn"].map(
                    lambda x: x[:-2], IGNORE
                )
                data["方向"] = data[f"{prefix}RunwayTypeCode"].map(
                    self.RUNWAYDIR, IGNORE
                )
                data["指廊"] = (
                    data[f"{prefix}GateNo"].map(
                        lambda x: self.APRONGATE.get(x, x[0] + "指廊"), IGNORE
                    )
                    if f"{prefix}GateNo" in data.columns
                    else None
                )
                data["架次"] = "架次"
                data["占比"] = "占比"

                data_ = concat(
                    [
                        data.assign(项目=i)
                        for i in ["架次", "占比", "方向", "区域", "指廊", "航司"]
                    ]
                )

                pivot = concat(
                    [
                        pivot_table(
                            data_.loc[data_["项目"] == i],
                            name,
                            ["项目", i],
                            name + "序数",
                            len,
                        )
                        for i in ["架次", "占比", "方向", "区域", "指廊", "航司"]
                    ]
                )
                pivot = pivot.reindex(
                    columns=np.arange(pivot.columns.min(), pivot.columns.max())
                ).fillna(0)
                pivot.columns = pivot.columns.map(
                    lambda x: "{}~{}{}".format(
                        *[
                            int(i) if i == int(i) else i
                            for i in (x * sep, (x + 1) * sep)
                        ],
                        "时",
                    )
                )
                pivot["总计"] = ""

                title, filename = (
                    "导出数据透视",
                    self.FILENAME.get("数据透视").format(*datetime.now().timetuple()),
                )
                file = (
                    f"{self.EXPORT_PATH.get()}/{filename}"
                    if self.EXPORT_PATH_.get()
                    else filedialog.asksaveasfilename(
                        filetypes=(("Xlsx表格文件", "*.xlsx"),),
                        confirmoverwrite=True,
                        parent=self,
                        title=title,
                        initialdir=self.EXPORT_PATH.get(),
                        initialfile=filename,
                    )
                )
                self.update_status({"数据透视": "保存中"})
                file = self.save_excel(
                    self,
                    title,
                    file,
                    save,
                    pivot=pivot,
                    data=data,
                    renamer=renamer,
                    name=name,
                )
                if file:
                    if file[0] == ".":
                        file = "程序所在文件夹" + file[1:]
                    self.update_log("数据透视保存至" + file, "file")
                else:
                    self.update_log("数据透视保存被取消")
            finally:
                self.update_status({"数据透视": ""})

        box = tk.Toplevel(self, name="数据透视")
        box.attributes("-topmost", self.TOPMOST.get())
        box.title("设置数据透视")
        box.geometry(
            f"+{self.winfo_rootx() + self.winfo_width() // 5}+{self.winfo_rooty() + self.winfo_height() // 5}"
        )

        op = {"+": add, "-": sub, "=": eq, "≠": ne, "<": lt, "≤": le, ">": gt, "≥": ge}
        mode = {"和": series_all, "或": series_any, "无次要条件": None}
        data = tk.StringVar(box, "延误判定")
        condition = list(tk.StringVar(box) for i in range(12))
        combobox = list(
            ttk.Combobox(box, textvariable=condition[i], width=12, state="readonly")
            for i in range(12)
        )
        range_ = condition.pop()
        range_.set("计划起飞STD")
        for i in 2, 4, 5, 7, 9, 10:
            values = list(op.keys())[:2] if i % 5 else list(op.keys())[2:]
            combobox[i].config(width=1, values=values)
            condition[i].set(values[0])
        times = list(tk.IntVar(box) for i in range(4))
        timeentry = list(
            ttk.Spinbox(
                box, textvariable=times[i], from_=0, to=999, increment=5, width=3
            )
            for i in range(4)
        )
        condition[0].set("无次要条件")
        times.insert(0, tk.DoubleVar(box, 1))
        timeentry.insert(
            0,
            ttk.Spinbox(
                box, textvariable=times[0], from_=0.5, to=10, increment=0.5, width=3
            ),
        )

        row = 0
        preset = tk.StringVar(box, "选择预设..." if self.PRESET else "无预设")

        def set_preset(*args):
            kwargs = self.PRESET.get(preset.get())
            data.set(kwargs.get("data"))
            condition[0].set(kwargs.get("condition")[0])
            name.set(kwargs.get("name"))
            tag = self.TAG_MAP[kwargs.get("data")[:2]]
            range_.set(kwargs.get("range"))
            for i, j in enumerate(kwargs.get("condition")[1:]):
                if j in op:
                    condition[i + 1].set(j)
                else:
                    condition[i + 1].set(f"{tag[j]} @{j}".replace("时间", ""))
            for i, j in enumerate(kwargs.get("times")):
                times[i + 1].set(j)
            for i, j in enumerate(kwargs.get("fillna")):
                fillna[i].set(
                    f"替换为{tag[j]} @{j}".replace("时间", "") if j in tag else j
                )
            update_combobox()
            delb.config(state=tk.NORMAL)

        def update_preset():
            editing = tk.Toplevel(box, name="preset_edit")
            editing.attributes("-topmost", self.TOPMOST.get())
            editing.title("设置预设名")
            if box.winfo_viewable():
                editing.transient(box)
            editing.resizable(False, False)
            editing.geometry(f"+{box.winfo_rootx() + 100}+{box.winfo_rooty() + 50}")

            scrollbar_e = ttk.Scrollbar(editing)
            entry = tk.Text(
                editing,
                width=30,
                height=3,
                undo=True,
                wrap="char",
                font=("微软雅黑", 10),
                yscrollcommand=scrollbar_e.set,
            )
            scrollbar_e.config(command=entry.yview)
            if preset.get() in self.PRESET:
                entry.insert("0.0", preset.get())

            popup = tk.Menu(editing, tearoff=False)

            def cut():
                try:
                    copy(), entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
                except Exception:
                    ...

            def copy():
                try:
                    (
                        entry.clipboard_clear(),
                        entry.clipboard_append(entry.get(tk.SEL_FIRST, tk.SEL_LAST)),
                    )
                except Exception:
                    ...

            def paste():
                try:
                    entry.insert(tk.INSERT, entry.selection_get(selection="CLIPBOARD"))
                except Exception:
                    ...

            def selall():
                entry.tag_add("sel", "0.0", tk.END)
                return "break"

            def delete():
                try:
                    entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
                except Exception:
                    ...

            popup.add_command(label="删除", command=delete)
            popup.add_command(label="剪切", command=cut)
            popup.add_command(label="复制", command=copy)
            popup.add_command(label="粘贴", command=paste)
            popup.add_separator()
            popup.add_command(label="全选", command=selall)

            entry.bind(
                "<Button-3>", lambda event: popup.post(event.x_root, event.y_root)
            )
            entry.grid(row=0, column=0, columnspan=2)
            scrollbar_e.grid(row=0, column=3, sticky="ns")

            def confirm(*args):
                v = entry.get("0.0", tk.END).strip().strip("\n").strip()
                if v in self.PRESET:
                    if not messagebox.askyesno(
                        TITLE,
                        "该预设名称已存在，是否替换原有预设为当前预设？",
                        parent=box,
                    ):
                        return 0
                self.PRESET[v] = {
                    "data": data.get(),
                    "times": list(inside(0, i.get(), 999) for i in times[1:]),
                    "fillna": list(i.get() for i in fillna),
                    "range": range_.get(),
                    "condition": list(i.get() for i in condition),
                    "name": name.get(),
                }
                for i, j in enumerate(self.PRESET[v]["condition"]):
                    if " @" in j:
                        self.PRESET[v]["condition"][i] = j[j.find(" @") + 2 :]
                for i, j in enumerate(self.PRESET[v]["fillna"]):
                    if " @" in j:
                        self.PRESET[v]["fillna"][i] = j[j.find(" @") + 2 :]
                preset.set(v)
                set_preset()
                preset_combobox.config(values=list(self.PRESET.keys()))
                editing.destroy()
                box.focus_set()

            entry.bind("<Return>", confirm)
            ttk.Button(editing, text="确认", width=6, command=confirm).grid(
                sticky="ws", row=2, column=0, columnspan=4, padx=5, pady=2
            )
            ttk.Button(editing, text="取消", width=6, command=editing.destroy).grid(
                sticky="es", row=2, column=0, columnspan=4, padx=5, pady=2
            )
            editing.grab_set()
            _setup_dialog(editing)
            editing.wait_window()

        def del_preset():
            if messagebox.askyesno(TITLE, f"确认删除{preset.get()}预设？", parent=box):
                self.PRESET.pop(preset.get())
                values = list(self.PRESET.keys())
                preset_combobox.config(values=values)
                if values:
                    preset.set(values[0])
                else:
                    preset.set("无预设")
                    delb.config(state=tk.DISABLED)

        ttk.Label(box, text="使用预设").grid(row=row, column=0, padx=3, pady=2)
        preset_combobox = ttk.Combobox(
            box,
            width=32,
            values=list(self.PRESET.keys()),
            textvariable=preset,
            state="readonly",
        )
        preset_combobox.grid(
            sticky="w", row=row, column=1, padx=3, pady=2, columnspan=4
        )
        preset_combobox.bind("<<ComboboxSelected>>", set_preset)

        delb = ttk.Button(box, text="删除预设", command=del_preset, state=tk.DISABLED)
        delb.grid(sticky="e", row=row, column=5, padx=3, pady=2, columnspan=3)
        ttk.Button(box, text="保存为预设...", command=update_preset).grid(
            sticky="w", row=row, column=8, padx=3, pady=2, columnspan=3
        )

        row += 1
        ttk.Separator(box, orient="horizontal").grid(
            column=0, row=row, padx=5, pady=1, columnspan=11, sticky="we"
        )

        row += 1

        def update_combobox(*args):
            tags = self.TAG_MAP[data.get()[:2]]
            values = [
                f"{j} @{i}".replace("时间", "")
                for i, j in tags.items()
                if i in self.CONVERT
            ]
            for i in 1, 3, 6, 8:
                combobox[i].config(values=values)
                if condition[i].get() not in values:
                    condition[i].set(values[0])
            flag = condition[0].get() == "无次要条件"
            for i in range(6, 11):
                combobox[i].config(state=state_(not flag), show=" " if flag else "")
            for i in 3, 4:
                timeentry[i].config(state=state(not flag), show=" " if flag else "")
            for i, j in enumerate(values):
                values[i] = "替换为" + j
            values = [
                "排除数据集",
                "替换为当前时间",
                "替换为导出时段开始时间",
                "替换为导出时段结束时间",
            ] + values
            for i in range(4):
                combobox_fillna[i].config(
                    values=values, state=state_(not flag) if i > 1 else "readonly"
                )
                if flag and i > 1:
                    fillna[i].set("")
                elif fillna[i].get() not in values:
                    fillna[i].set("排除数据集")
            if data.get() == "延误判定":
                range_.set("计划起飞STD")
                combobox[-1].config(state=tk.DISABLED)
            else:
                combobox[-1].config(state="readonly")

        ttk.Label(box, text="基础数据选择").grid(row=row, column=0, padx=2, pady=4)
        data_combobox = ttk.Combobox(
            box,
            width=12,
            textvariable=data,
            state="readonly",
            values=["航班查询", "延误判定"],
        )
        data_combobox.grid(row=row, column=2, padx=2, pady=4)
        data_combobox.bind("<<ComboboxSelected>>", update_combobox)

        name = tk.StringVar(box, "目标值")
        ttk.Label(box, text="第一行计算结果作为统计主体，设置名称").grid(
            row=row, column=4, padx=2, pady=4, columnspan=5
        )
        ttk.Entry(box, textvariable=name, width=12).grid(
            row=row, column=8, padx=2, pady=4, columnspan=4, sticky="e"
        )

        row += 1
        ttk.Label(box, text="数据筛选条件").grid(
            row=row, column=0, padx=2, pady=2, columnspan=2
        )
        combobox[0].config(values=list(mode.keys()), width=8)
        combobox[0].grid(row=row + 1, column=0, padx=2, pady=2, columnspan=2)
        combobox[0].bind("<<ComboboxSelected>>", update_combobox)

        for i in range(10):
            combobox[i + 1].grid(
                row=i // 5 + row,
                column=i % 5 + (4 if i % 5 == 4 else 2),
                padx=2,
                pady=2,
            )
        for i in range(4):
            timeentry[i + 1].grid(
                row=i // 2 + row, column=6 + 3 * (i % 2), padx=2, pady=2
            )
            ttk.Label(box, text="分钟").grid(
                row=i // 2 + row, column=7 + 3 * (i % 2), padx=2, pady=2
            )

        row += 2
        fillna = list(tk.StringVar(box) for _ in range(4))
        combobox_fillna = list(
            ttk.Combobox(box, textvariable=fillna[i], width=19, state="readonly")
            for i in range(4)
        )

        ttk.Label(box, text="将第1个变量的空值").grid(
            row=row, column=0, padx=2, pady=2, columnspan=3, sticky="w"
        )
        combobox_fillna[0].grid(
            row=row, column=2, padx=32, pady=2, columnspan=3, sticky="w"
        )

        ttk.Label(box, text="将第2个变量的空值").grid(
            row=row, column=4, padx=2, pady=2, columnspan=4
        )
        combobox_fillna[1].grid(
            row=row, column=5, padx=5, pady=2, columnspan=6, sticky="e"
        )

        row += 1
        ttk.Label(box, text="将第3个变量的空值").grid(
            row=row, column=0, padx=2, pady=2, columnspan=3, sticky="w"
        )
        combobox_fillna[2].grid(
            row=row, column=2, padx=32, pady=2, columnspan=3, sticky="w"
        )

        ttk.Label(box, text="将第4个变量的空值").grid(
            row=row, column=4, padx=2, pady=2, columnspan=4
        )
        combobox_fillna[3].grid(
            row=row, column=5, padx=5, pady=2, columnspan=6, sticky="e"
        )

        row += 1
        ttk.Separator(box, orient="horizontal").grid(
            column=0, row=row, padx=5, pady=2, columnspan=11, sticky="we"
        )

        row += 3
        ttk.Label(box, text="数据获取时段标准").grid(
            row=row, column=0, padx=2, pady=2, columnspan=3, sticky="w"
        )
        combobox[11].config(width=19, values=list(departure_mode.keys()))
        combobox[11].grid(row=row, column=2, padx=32, pady=2, columnspan=3, sticky="w")

        ttk.Label(box, text="统计时段单位").grid(
            row=row, column=6, padx=2, pady=2, columnspan=3, sticky="e"
        )
        timeentry[0].grid(row=row, column=9, padx=2, pady=2)
        ttk.Label(box, text="小时").grid(row=row, column=10, padx=2, pady=2)

        row += 1
        today = datetime(*datetime.now().timetuple()[:3])

        def confirm(*args: datetime):
            conditions = list(i.get() for i in condition[1:])
            conditions = list(op.get(i, i[i.find(" @") + 2 :]) for i in conditions)
            fillnas = list(i.get() for i in fillna)
            fillnas = list(
                i[i.find(" @") + 2 :] if " @" in i else i.replace("排除数据集", "")
                for i in fillnas
            )

            times[0].set(inside(0.01, times[0].get(), 10))
            for i in range(4):
                times[i + 1].set(inside(0, times[i + 1].get(), 999))
            conditions.append(timedelta(minutes=times[4].get()))
            conditions.insert(9, timedelta(minutes=times[3].get()))
            conditions.insert(5, timedelta(minutes=times[2].get()))
            conditions.insert(4, timedelta(minutes=times[1].get()))

            kwargs = {
                "data": data.get()[:2],
                "mode": mode[condition[0].get()],
                "sep": times[0].get(),
                "condition": conditions,
                "fillna": fillnas,
                "name": name.get(),
                "range_": range_.get(),
            }
            if args:
                threading.Thread(
                    target=run,
                    args=(args[0], args[1] - timedelta(minutes=1)),
                    kwargs=kwargs,
                ).start()
            else:
                self.ask_export(
                    today,
                    today + timedelta(1),
                    run,
                    "导出数据透视表格",
                    "",
                    (INIT, today + timedelta(7)),
                    (INIT, today + timedelta(7)),
                    master=box,
                    **kwargs,
                )
            self.focus_set()

        ttk.Button(
            box,
            text="导出昨日",
            width=15,
            command=lambda: confirm(today - timedelta(1), today),
        ).grid(sticky="w", row=row, column=0, padx=5, pady=5, columnspan=11)
        ttk.Button(
            box,
            text="导出今日",
            width=15,
            command=lambda: confirm(today, today + timedelta(1)),
        ).grid(row=row, column=0, padx=5, pady=5, columnspan=11)
        ttk.Button(
            box, text="导出自定义时段...", width=18, command=lambda: confirm()
        ).grid(sticky="e", row=row, column=0, padx=5, pady=5, columnspan=11)

        update_combobox()
        box.focus_set()
        box.resizable(False, False)
        box.mainloop()

    def format_ctot(
        self,
        file: str,
        data: dict[str, tuple[DataFrame, str]],
        res: str,
        count: int,
        filename: str,
    ):
        width = {
            "A": 4,
            "B": 9,
            "C": 11,
            "D": 17.5,
            "E": 7,
            "F": 6,
            "G": 6,
            "H": 6,
            "I": 6,
            "J": 6,
            "K": 9,
            "L": 6,
            "M": 7,
            "N": 14.5,
            "O": {},
        }
        config = self.CTOTCONFIG
        if not config["延误"][1]:
            count = 0
        wb = ExcelWriter(file, mode="w")
        for name, (data, msg) in data.items():
            data.reset_index(drop=True, inplace=True)
            data.index += 1
            data.to_excel(wb, sheet_name=name, freeze_panes=(1, 1), index_label=msg)
            width["O"][name] = inside(
                9, data["流控信息"].map(len).max() * 1.3 // 1.5 / 2, 60
            )

        for ws in wb.sheets.values():
            for k, v in width.items():
                ws.column_dimensions[k].width = (
                    v[ws.title] if isinstance(v, dict) else v
                )
                ws.column_dimensions[k].hidden = config[k][1]

            for col in range(1, ws.max_column + 1):
                cell = ws[f"{get_column_letter(col)}1"]
                cell.font = self.font(bold=True)
                cell.alignment = ALIGN_CENTER

            for row in range(2, ws.max_row + 1):
                cell = ws[f"A{row}"]
                cell.font = self.font(bold=True)
                cell.alignment = ALIGN_CENTER

            for col in range(2, ws.max_column + 1):
                col = get_column_letter(col)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col}{row}"]
                    cell.border = BORDER
                    cell.font = self.font()
                    cell.alignment = ALIGN_CENTER_CENTER

            while count:
                for col in range(1, ws.max_column + 1):
                    ws[f"{get_column_letter(col)}{count + 1}"].fill = PatternFill(
                        "solid", config["延误"][2]
                    )
                count -= 1

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, 4)
                if config["D"][2]:
                    if "未到" in cell.value and "延误" in cell.value:
                        cell.fill = PatternFill("solid", config["D"][3])
                    elif ("未起" in cell.value and "延误" in cell.value) or cell.value[
                        -2:
                    ] == "未起":
                        cell.fill = PatternFill("solid", config["D"][4])
                    cell = ws.cell(row, 2)
                if config["B"][2] and cell.value.split("\n")[0] in res:
                    cell.font = self.font(bold=True)

                cell = ws.cell(row, 3)
                if len(cell.value) > 10:
                    cell.font = self.font(size=9)

                cell = ws.cell(row, 15)
                cell.alignment = ALIGN_CENTER_LEFT
                cell.font = self.font(size=9)

            for row in ws.iter_rows(2, ws.max_row, 6, 9):
                for cell in row:
                    cell.number_format = "HH:MM"

            for row in ws.iter_rows(2, ws.max_row, 10, 10):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.value = -cell.value
                        cell.number_format = "-H:MM"
                    else:
                        cell.number_format = "H:MM"

            for row in ws.iter_rows(2, ws.max_row, 11, 11):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if cell.value < 0:
                            if config["K"][2]:
                                cell.fill = PatternFill("solid", config["K"][3])
                            cell.value = -cell.value
                            cell.number_format = "-H:MM"
                        elif cell.value >= 1:
                            cell.value = "≥24h"
                        else:
                            cell.number_format = "H:MM"

            for row in ws.iter_rows(2, ws.max_row, 13, 13):
                for cell in row:
                    if cell.value and config["M"][2]:
                        cell.fill = PatternFill("solid", config["M"][3])
                        cell.font = self.font(
                            size=9 if cell.value.count("\n") >= 2 else 11,
                            bold="兴快线" in cell.value or "VIP" in cell.value,
                        )

            for row in ws.iter_rows(2, ws.max_row, 14, 14):
                for cell in row:
                    cell.alignment = ALIGN_CENTER_LEFT
                    cell.font = self.font(size=9)
                    if config["N"][2] and "推迟" in str(cell.value).split("\n")[-1]:
                        cell.fill = PatternFill("solid", config["N"][3])

            if config["J"][2]:
                ws.conditional_formatting.add(
                    f"J{2}:J{ws.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=0,
                        start_color=config["J"][3],
                        end_type="percent",
                        end_value=80,
                        end_color=config["J"][4],
                    ),
                )

            merge_range = f"A{ws.max_row + 1}:O{ws.max_row + 1}"
            ws.merge_cells(merge_range)
            cell = ws[merge_range[: merge_range.find(":")]]
            cell.value = ws["A1"].value
            cell.font = self.font()
            cell.alignment = ALIGN_CENTER_LEFT
            ws["A1"].value = "序号"

            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 36

        ws = wb.sheets["总表"]
        merge_range = f"A{ws.max_row + 1}:O{ws.max_row + 1}"
        ws.merge_cells(merge_range)
        cell = ws[merge_range[: merge_range.find(":")]]
        cell.value = res
        cell.alignment = ALIGN_CENTER_LEFT
        cell.font = self.font()
        ws.row_dimensions[ws.max_row].height = 36
        wb.close()

    def get_ctot(self, __minutes: int = 0, path: tk.StringVar = None):
        try:
            delay = self.DELAY.get()
            warning = ["按起飞延误"] if delay == "outStot" else []
            if "W/Z" not in self.CTOTTYPE:
                warning.append("不含正班")
            if path:
                path = path.get().format("CTOT推点表格")
                if not os.path.exists(path):
                    os.makedirs(path)

            datetime_now = self.datetime_now()
            self.update_status({"CTOT推点航班明细": "进行中"}, set_bar=True)
            end_time = "截至{}:{:02}，".format(datetime_now.hour, datetime_now.minute)
            data = (
                self.get_data(
                    "航班", notna=("outCtot",), outFlightTypeCode=self.CTOTTYPE
                )
                .rename(columns={"outGuid": "guid"})
                .set_index("guid", drop=True)
            )
            data = data.loc[
                data["outAtot"].isna() | (data["outAtot"] > datetime_now)
            ].copy()
            data["ctotOffset"] = data["outCtot"] - data[delay]
            target = data.loc[data["ctotOffset"] > timedelta(minutes=__minutes)].copy()

            if target.__len__():
                history = self.HISTORY["CTOT"].reindex(index=data.index)
                target["ctotHistory"] = history.apply(
                    lambda x: self.history_summary(x, data.loc[x.name]), axis=1
                )

                target["inAldtText"] = (
                    target[["inAldt", "inSldt"]]
                    .apply(
                        lambda x: (
                            "{}\n{}到达".format(
                                x["inAldt"].strftime(r"%m-%d %H:%M"),
                                timedelta_to_delay(x["inAldt"] - x["inSldt"]),
                            )
                            if np.all(notna(x))
                            else np.nan
                        ),
                        axis=1,
                    )
                    .fillna(
                        target["inAtot"]
                        .isna()
                        .map(lambda x: "前序航班未起" if x else "前序航班未到")
                        .astype(str)
                        + (target["inEldt"] - target["inSldt"])
                        .map(lambda x: f",预计{timedelta_to_delay(x)}到达", IGNORE)
                        .fillna(
                            (target["inAtot"] - target["inStot"]).map(
                                lambda x: f",前序{timedelta_to_delay(x)}起飞", IGNORE
                            )
                        )
                        .fillna("")
                    )
                )

                target["note"] = (
                    target["outVipInd"]
                    .map(lambda x: self.TAG_MAP["标签"].get(x[0], ""))
                    .map(lambda x: x if x == "" else f"{x}\n")
                )
                target["note"] += target[["airlineIata", "outRoute"]].apply(
                    lambda x: "兴快线\n" if self.is_exp(*x.values) else "", axis=1
                )
                status = {
                    "outPushTime": "滑行中",
                    "moniJob.tract_D.actBeginTime": "已推出",
                    "outAcct": "已关舱",
                    "outAebt": "已登结",
                    "outAsbt": "登机中",
                }
                target["outStatus"] = None
                for k, v in status.items():
                    i = target[k] <= datetime_now
                    target.loc[i, "outStatus"] = target.loc[i, "outStatus"].fillna(
                        f"{v}\n"
                    )
                target["note"] += target["outStatus"].fillna("")

                target["outFlightTypeCode"] = target["outFlightTypeCode"].map(self.TYPE)
                i = target["outFlightTypeCode"] != "正班"
                target.loc[i, "outFlightNo"] += (
                    "\n" + target.loc[i, "outFlightTypeCode"]
                )
                target["outGateNo"] = (
                    target["outGateNo"].fillna("-").astype(str)
                    + "\n"
                    + target["portNoDp"].fillna("-").astype(str)
                )
                target["outAirportRegionCn"] = target["outAirportRegionCn"].map(
                    lambda x: x[:-2], IGNORE
                )
                target["outRouteCn"] = (
                    target[["outRoute", "outRouteCn"]]
                    .fillna("")
                    .apply(self.airport_name, axis=1)
                )
                target["note"] = target["note"].map(lambda x: x[:-1], IGNORE)
                target["outAsbt"] = target["outAsbt"].fillna("-")

                res = []
                for i, j, k in (
                    target.loc[target["outTmi"].notna()]
                    .loc[target["ctotOffset"] + target["ttt"] >= timedelta()][
                        ["outFlightNo", "inAldtText", "ctotOffset"]
                    ]
                    .values
                ):
                    k = np.int64(k.total_seconds())
                    if "延误" in j and "未" in j:
                        if int(j[j.find("延误") + 2 : -4]) * 60 > k:
                            continue
                    res.append(i.split("\n")[0])

                target["outTmi"] = target["outTmi"].fillna("无流控信息")

                renamer = {
                    "outFlightNo": "航班号",
                    "outRouteCn": "下站",
                    "inAldtText": "前序航班落地时间",
                    "outGateNo": "登机门",
                    "outAsbt": "登机",
                    "outSobt": "STD",
                    delay: "起延" if delay == "outStot" else "最晚",
                    "outCtot": "CTOT",
                    "ctotOffset": "推点",
                    "ttt": "过站裕度",
                    "outAirportRegionCn": "区域",
                    "note": "备注",
                    "ctotHistory": "CTOT历史",
                    "outTmi": "流控信息",
                }

                data = {
                    "总表": (target.sort_values(by=delay), "CTOT推点{}航班共计{}"),
                    "后续延误": (
                        target.loc[target[delay] >= datetime_now],
                        "后续CTOT推点{}航班{}",
                    ),
                    "已延误": (
                        target.loc[target[delay] < datetime_now],
                        "已延误未起飞的CTOT推点{}航班{}",
                    ),
                }
                delayed = data["已延误"][0].__len__()
                __minutes = f"{__minutes}分钟以上" if __minutes else ""

                output = dict()
                for name, (target, count) in data.items():
                    if target.__len__():
                        msg = (
                            end_time
                            + count.format(__minutes, target.__len__())
                            + "架次（"
                            + "，".join(
                                f"{region}{_data.__len__()}架次"
                                for region, _data in target.groupby(
                                    "outAirportRegionCn", sort=False
                                )
                            )
                            + "），"
                        )
                        if target.__len__() > 1:
                            msg += "平均"
                        msg += f"推点{self.min_sec_format(target['ctotOffset'].mean().total_seconds(), wrap_hour=True)}，"
                        _data = target.loc[
                            target["ctotOffset"] > timedelta(hours=1)
                        ].copy()
                        msg += f"推点1小时以上{_data.__len__()}架次"
                        if _data.__len__():
                            msg += "【"
                            i = 1
                            while _data.__len__():
                                i += 1
                                __data = _data.loc[
                                    _data["ctotOffset"] <= timedelta(hours=i)
                                ].index
                                if __data.size:
                                    msg += f"推点{i - 1}-{i}小时{__data.size}架次，"
                                    _data.drop(__data, inplace=True)
                            msg = msg[:-1] + "】。"
                        else:
                            msg += "。"
                        output[name] = (
                            target.reindex(columns=renamer.keys()).rename(
                                columns=renamer
                            ),
                            msg,
                        )

                res = (
                    (
                        f"因流控限制推点航班{len(res)}架次："
                        + " ".join(res)
                        + (
                            ""
                            if len(res) == len(output["总表"][0])
                            else "，其余航班因过站时间不足或预计前飞晚到或TOBT较晚导致推点。"
                        )
                    )
                    if res
                    else "无流控限制推点航班，均因过站时间不足或预计前飞晚到或TOBT较晚导致推点。"
                )
                msg = output["总表"][1]  # + res

                title, filename = (
                    f"导出CTOT推点{__minutes}航班明细表格",
                    self.FILENAME.get("CTOT推点航班明细表格").format(
                        *datetime.now().timetuple(), minutes=__minutes
                    ),
                )
                file = (
                    f"{path}/{filename}"
                    if path
                    else (
                        f"{self.INFO_PATH.get()}/{filename}"
                        if self.INFO_PATH_.get()
                        else filedialog.asksaveasfilename(
                            filetypes=(("Xlsx表格文件", "*.xlsx"),),
                            confirmoverwrite=True,
                            parent=self,
                            title=title,
                            initialdir=self.INFO_PATH.get(),
                            initialfile=filename,
                        )
                    )
                )
                self.update_status({"CTOT推点航班明细": "保存中"})
                filename = os.path.basename(file)
                file = self.save_excel(
                    self,
                    title,
                    file,
                    self.format_ctot,
                    not path,
                    data=output,
                    res=res,
                    count=delayed,
                    filename=filename,
                )
                if file:
                    img = (
                        file.rsplit(r"/", 1)[0]
                        + r"/"
                        + filename.replace(".xlsx", ".png")
                    )
                    img = self.save_excel_img(
                        file, img, "总表", f"A1:O{len(output['总表'][0]) + 1}"
                    )
                    if file[0] == ".":
                        file = "程序所在文件夹" + file[1:]
                    self.update_log(f"CTOT推点{__minutes}航班明细保存至" + file, "file")
                    if not img:
                        warning.append("图片生成失败")
                else:
                    self.update_log(f"CTOT推点{__minutes}航班明细保存被取消")
                    img = ""
            else:
                msg = end_time + "无CTOT推点航班。"
                file = img = res = ""
            res += self.wrap_iterstr(warning, "（{}）", "，")
            self.update_log("CTOT推点信息生成成功：" + msg + res, "text")
            path = (
                path
                if path
                else self.INFO_PATH.get() if self.INFO_PATH_.get() else None
            )
            msg += self.wrap_iterstr(warning, "（{}）", "，")
            if "sync" not in self.RUNNING:
                self.save_textfile(msg, path)
            self.save_push("CTOT推点航班", msg, img)
            return {
                "CTOT推点航班": msg,
                "CTOT推点航班表格": file,
                "CTOT推点航班图片": img,
            }
        finally:
            self.update_status({"CTOT推点航班明细": ""})

    def format_long_delay(self, file: str, data: DataFrame, msg: str, filename: str):
        wb = ExcelWriter(file, mode="w")
        data.reset_index(drop=True, inplace=True)
        data.index += 1
        data.to_excel(
            wb, sheet_name="延误未起飞", index_label="序号", freeze_panes=(1, 1)
        )
        ws = wb.sheets["延误未起飞"]
        width = {
            "A": 4,
            "B": 9,
            "C": 13.5,
            "D": 17,
            "E": 5,
            "F": 5,
            "G": 6,
            "H": 6,
            "I": 6,
            "J": 6,
            "K": 7,
            "L": 8.5,
            "M": 7,
            "N": 10,
            "O": 16,
        }
        config = self.DELAYCONFIG
        for k, v in width.items():
            ws.column_dimensions[k].width = v
            ws.column_dimensions[k].hidden = config[k][1]

        for col in range(1, ws.max_column + 1):
            cell = ws[f"{get_column_letter(col)}1"]
            cell.font = self.font(b=True)
            cell.alignment = ALIGN_CENTER

        for col in range(1, ws.max_column + 1):
            col = get_column_letter(col)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col}{row}"]
                cell.border = BORDER
                cell.font = self.font()
                cell.alignment = ALIGN_CENTER

        for row in ws.iter_rows(2, ws.max_row, 7, 9):
            for cell in row:
                cell.number_format = "HH:MM"

        for row in ws.iter_rows(2, ws.max_row, 10, 12):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        cell.value = "-"
                    else:
                        cell.number_format = "H:MM"
                else:
                    cell.value = "-"

        for row in ws.iter_rows(2, ws.max_row, 13, 15):
            for cell in row:
                cell.font = self.font(size=10)

        for k in "J", "K", "L":
            if config[k][2]:
                ws.conditional_formatting.add(
                    f"{k}{2}:{k}{ws.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        end_type="max",
                        start_color=config[k][3],
                        end_color=config[k][4],
                    ),
                )

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            cell = ws.cell(row, 3)
            if len(cell.value) >= 8:
                cell.font = self.font(size=9)
            elif len(cell.value) >= 6:
                cell.font = self.font(size=10)

            cell = ws.cell(cell.row, 13)
            if config["M"][2] and "未" not in str(cell.value):
                cell.font = self.font(bold=True, size=10)

        merge_range = f"A{ws.max_row + 1}:O{ws.max_row + 1}"
        ws.merge_cells(merge_range)
        cell = ws[merge_range[: merge_range.find(":")]]
        cell.value = msg
        cell.alignment = ALIGN_CENTER_LEFT
        cell.font = self.font()
        ws.row_dimensions[ws.max_row].height = 36
        wb.close()

    def long_delay(self, __hours: float = 0, path: tk.StringVar = None):
        try:
            if int(__hours) == __hours:
                __hours = int(__hours)
            delay, exclude = self.DELAY.get(), self.EXCLUDE.get()
            warning = ["按起飞延误"] if delay == "outStot" else []
            if exclude:
                warning.append("不含前序未落地航班")
            if path:
                path = path.get().format("延误未起表格")
                if not os.path.exists(path):
                    os.makedirs(path)

            self.update_status({"延误未起飞航班明细": "进行中"}, set_bar=True)
            hours = f"{__hours}小时以上" if __hours else ""
            filename = self.FILENAME.get("延误未起飞航班明细表格").format(
                *datetime.now().timetuple(), hours=hours
            )
            datetime_now = self.datetime_now()

            data = self.get_data("延误", notna=["inAldt"] if exclude else [])
            data = (
                data.loc[data["outAtot"].isna() | (data["outAtot"] > datetime_now)]
                .loc[data[delay] < datetime_now]
                .copy()
            )
            data["waiting"] = datetime_now - data["outAcct"]
            data["delayed"] = datetime_now - data[delay]
            data = data.loc[data["delayed"] >= timedelta(hours=__hours)].copy()

            msg = "截至{}:{:02}，".format(datetime_now.hour, datetime_now.minute)
            if data.__len__():
                renamer = {
                    "outFlightNo": "航班号",
                    "outRouteCn": "下站",
                    "inAldt": "前序航班落地时间",
                    "outGateNo": "门",
                    "portNoDp": "机位",
                    "outSobt": "STD",
                    delay: "起延" if delay == "outStot" else "最晚",
                    "outCtot": "CTOT",
                    "delayed": "已延误",
                    "toCtot": "距CTOT",
                    "waiting": "机上等待",
                    "outStatus": "状态",
                    "outAirportRegionCn": "区域",
                    "subDelayReason": "判定延误原因",
                }

                data["outStatus"] = None
                for k, v in self.STATUS.items():
                    i = data[k] <= datetime_now
                    data.loc[i, "outStatus"] = data.loc[i, "outStatus"].fillna(v)
                data["outRouteCn"] = (
                    data[["outRoute", "outRouteCn"]]
                    .fillna("")
                    .apply(self.airport_name, axis=1)
                    .map(lambda x: x.split("-")[0], IGNORE)
                )
                data["outAirportRegionCn"] = data["outAirportRegionCn"].map(
                    lambda x: x[:-2], IGNORE
                )
                data["inAldt"] = (
                    data["inAldt"]
                    .map(lambda x: x.strftime(r"%m-%d %H:%M"), IGNORE)
                    .fillna(
                        (data["inAtot"] - data["inStot"])
                        .map(lambda x: f"前起{timedelta_to_delay(x)}", IGNORE)
                        .fillna("前站未起")
                    )
                )
                try:
                    data["portNoDp"] = data["portNoDp"].astype(int, errors="raise")
                except Exception:
                    data["portNoDp"] = data["portNoDp"].map(
                        lambda x: int(x) if str(x).isnumeric() else str(x), IGNORE
                    )
                data["toCtot"] = data["outCtot"] - datetime_now
                for i in "本场", "外站":
                    k = (data["addDelayReason"] == i) & (
                        (data["priDelayReason"] == "天气")
                        | (data["priDelayReason"] == "军事活动")
                    )
                    data.loc[k, "subDelayReason"] = data.loc[k, "subDelayReason"].map(
                        lambda x: str(x).replace("其它", i), IGNORE
                    )

                i, _data = __hours, data.copy()
                msg += (
                    f"延误{hours}未起飞航班{_data.__len__()}架次（"
                    + "，".join(
                        f"{region}{__data.__len__()}架次"
                        for region, __data in data.groupby(
                            "outAirportRegionCn", sort=False
                        )
                    )
                    + "）"
                )
                while _data.__len__():
                    i += 1
                    __data = _data.loc[_data["delayed"] <= timedelta(hours=i)].index
                    if __data.size:
                        msg += f"，延误{i - 1}-{i}小时{__data.size}架次"
                        _data.drop(__data, inplace=True)
                msg += "。"

                title = f"导出延误{hours}未起飞航班明细表格"
                file = (
                    f"{path}/{filename}"
                    if path
                    else (
                        f"{self.INFO_PATH.get()}/{filename}"
                        if self.INFO_PATH_.get()
                        else filedialog.asksaveasfilename(
                            filetypes=(("Xlsx表格文件", "*.xlsx"),),
                            confirmoverwrite=True,
                            parent=self,
                            title=title,
                            initialdir=self.INFO_PATH.get(),
                            initialfile=filename,
                        )
                    )
                )
                self.update_status({"延误未起飞航班明细": "保存中"})
                filename = os.path.basename(file)
                file = self.save_excel(
                    self,
                    title,
                    file,
                    self.format_long_delay,
                    not path,
                    data=data.reindex(columns=renamer.keys()).rename(columns=renamer),
                    msg=msg,
                    filename=filename,
                )
                if file:
                    img = (
                        file.rsplit(r"/", 1)[0]
                        + r"/"
                        + filename.replace(".xlsx", ".png")
                    )
                    img = self.save_excel_img(
                        file, img, "延误未起飞", f"A1:O{len(data) + 1}"
                    )
                    if file[0] == ".":
                        file = "程序所在文件夹" + file[1:]
                    self.update_log(f"延误{hours}未起飞航班明细保存至" + file, "file")
                    if not img:
                        warning.append("图片生成失败")
                else:
                    self.update_log(f"延误{hours}未起飞航班明细保存被取消")
                    img = ""
            else:
                msg += f"无延误{hours}未起飞航班。"
                file = img = ""

            msg += self.wrap_iterstr(warning, "（{}）", "，")

            self.update_log("延误未起飞信息生成成功：" + msg, "text")
            path = (
                path
                if path
                else self.INFO_PATH.get() if self.INFO_PATH_.get() else None
            )
            if "sync" not in self.RUNNING:
                self.save_textfile(msg, path)
            self.save_push(f"延误{hours}未起飞航班", msg, img)
            return {
                f"延误{hours}未起飞航班": msg,
                f"延误{hours}未起飞航班表格": file,
                f"延误{hours}未起飞航班图片": img,
            }
        finally:
            self.update_status({"延误未起飞航班明细": ""})

    def stock_msg(self, path: tk.StringVar = None):
        try:
            delay = "outLastTot"  # self.DELAY.get()
            yesterday = self.yesterday.get()
            datetime_now, today = (
                self.datetime_now(),
                self.today(-1 if yesterday else 0),
            )
            tommorow = today + timedelta(1)
            msg = "截至{}:{:02}，".format(datetime_now.hour, datetime_now.minute)
            self.update_status({"当前运行概述": "进行中"}, set_bar=True)

            raw_arrival, raw_depart, raw_flight, raw_pass = self.get_data(
                "执行_进港", "执行_离港", "航班", "旅客"
            )
            raw_depart.drop(
                raw_depart.loc[raw_depart["operationStatusCode"] == "CNCL"].index,
                inplace=True,
            )
            raw_arrival.drop(
                raw_arrival.loc[raw_arrival["operationStatusCode"] == "CNCL"].index,
                inplace=True,
            )

            dep_s = (
                raw_depart.loc[raw_depart["sobt"] >= today]
                .loc[raw_depart["sobt"] < min(tommorow, datetime_now)]
                .loc[raw_depart["flightTypeCode"] != "F/H"]
                .loc[raw_depart["flightTypeCode"] != "Q/B"]
            )
            arr_s = (
                raw_arrival.loc[raw_arrival["sibt"] >= today]
                .loc[raw_arrival["sibt"] < min(tommorow, datetime_now)]
                .loc[raw_arrival["flightTypeCode"] != "F/H"]
                .loc[raw_arrival["flightTypeCode"] != "Q/B"]
            )
            dep_a = dep_s.loc[dep_s["atot"].notna()].__len__()
            arr_a = arr_s.loc[arr_s["aldt"].notna()].__len__()
            msg += "{}计划执行航班{}架次（离港{}架次，进港{}架次），其中已执行{}架次（离港{}架次，进港{}架次）；".format(
                "昨日" if yesterday else "",
                dep_s.__len__() + arr_s.__len__(),
                dep_s.__len__(),
                arr_s.__len__(),
                dep_a + arr_a,
                dep_a,
                arr_a,
            )

            dep_a = (
                raw_depart.loc[raw_depart["sobt"] >= today]
                .loc[raw_depart["atot"] < datetime_now]
                .__len__()
            )
            arr_a = (
                raw_arrival.loc[raw_arrival["sibt"] >= today]
                .loc[raw_arrival["aldt"] < datetime_now]
                .__len__()
            )
            msg += "实际执行{}架次（离港{}架次，进港{}架次）；\n".format(
                dep_a + arr_a, dep_a, arr_a
            )
            if yesterday:
                msg += "昨日离港航班"
            else:
                msg += "今日计划时间在{}:{:02d}前的离港航班".format(
                    datetime_now.hour, datetime_now.minute
                )

            data = (
                raw_flight.loc[raw_flight["outAtot"].isna()]
                .loc[raw_flight["outSobt"] >= today]
                .loc[raw_flight["outSobt"] < min(tommorow, datetime_now)]
                .copy()
            )
            if data.__len__():
                data["status"] = None
                for k, v in self.STATUS.items():
                    i = data[k].notna()
                    data.loc[i, "status"] = data.loc[i, "status"].fillna(v)
                status_map = list(self.STATUS.values())[::-1]
                for i in status_map.copy():
                    if i not in data["status"].values:
                        status_map.remove(i)
                msg += (
                    "{}架次未执行（".format(data.__len__())
                    + "，".join(
                        f"{i}{len(data.loc[data['status'] == i])}架次"
                        for i in status_map
                    )
                    + "）；"
                )
            else:
                msg += "均已起飞；"

            if not (yesterday or raw_pass.empty):
                data = raw_pass.loc[raw_pass["acctCabin"].isna()].copy()
                data["parkRegion"] = (
                    data["gateNo"]
                    .map(lambda x: self.APRONGATE.get(x, x[0] + "指廊"), IGNORE)
                    .fillna("未发布登机门航班旅客")
                )
                data["unBoardNum"] = data["unBoardNum"].map(
                    lambda x: 0 if x < 0 else x, IGNORE
                )
                i = int(data["unBoardNum"].sum())
                msg += f"\n今日所有未执行的离港航班中，已过检未登机旅客约{i}人"

                if i:
                    park_regions = [
                        "A指廊",
                        "B指廊",
                        "C指廊",
                        "D指廊",
                        "E指廊",
                        "国内西远机位",
                        "国内东远机位",
                        "国际远机位",
                        "未发布登机门航班旅客",
                    ]
                    msg += "（{}）".format(
                        "，".join(
                            f"{i}约{int(data.loc[data['parkRegion'] == i, 'unBoardNum'].sum())}人"
                            for i in park_regions
                        ),
                    )
                msg += "；"

            data = raw_depart.loc[raw_depart["atot"].isna()].loc[
                raw_depart["asbt"].notna()
            ]
            data["runwayTypeCode"] = data["runwayTypeCode"].map(self.RUNWAYDIR, IGNORE)

            msg += (
                "\n已登机准备出港航班{}架次（{}）；".format(
                    data.__len__(),
                    "；".join(
                        f"{i[0]}向{len(i[1])}架次，其中{len(i[1].loc[i[1]['pushTime'].notna()])}架次滑行中"
                        for i in data.groupby("runwayTypeCode", sort=False)
                    ),
                )
                if data.__len__()
                else "\n无已登机准备出港航班"
            )

            data = self.get_data(
                "航班", notna=("outCtot",), outFlightTypeCode=self.CTOTTYPE
            )
            data = data.loc[data["outAtot"].isna() & (data["outEstripStatus"] != "DEP")]
            data["ctotOffset"] = data["outCtot"] - data[delay]
            data = data.loc[data["ctotOffset"] > timedelta()].copy()
            data_len = data.__len__()

            msg += (
                "CTOT推点航班共计{}架次（{}），{}推点时长{}，推点一小时以上{}架次。".format(
                    data_len,
                    "，".join(
                        [
                            f"{i[0][:-2]}{len(i[1])}架次"
                            for i in data.groupby("outAirportRegionCn", sort=False)
                        ]
                    ),
                    "" if data_len == 1 else "平均",
                    self.min_sec_format(
                        data["ctotOffset"].mean().total_seconds(), wrap_hour=True
                    ),
                    data.loc[data["ctotOffset"] > timedelta(hours=1)].__len__(),
                )
                if data_len
                else "无CTOT推点航班。"
            )
            msg += "未来一小时预计进港航班{}架次。".format(
                raw_arrival.loc[raw_arrival["eldt"] >= datetime_now]
                .loc[raw_arrival["eldt"] < datetime_now + timedelta(hours=1)]
                .__len__()
            )

            self.update_log("当前运行概述生成成功：" + msg, "text")
            path = (
                path.get().format("当前运行概述")
                if path
                else self.INFO_PATH.get() if self.INFO_PATH_.get() else None
            )
            if "sync" not in self.RUNNING:
                self.save_textfile(msg, path)
            self.save_push("当前运行概述", msg)
            return {"当前运行概述": msg}
        finally:
            self.update_status({"当前运行概述": ""})

    def get_monitor(self):
        if i := self.RUNNING.get("get_monitor", dict()).get(0):
            self.after_cancel(i)
        datetime_now = datetime.now()
        timetuple = datetime_now.timetuple()
        today = datetime(*timetuple[:3])
        output = {"数据日期时间": int(round(datetime_now.timestamp(), -1) * 1000)}

        raw_data = self.get_flight_info(
            today - timedelta(1),
            datetime_now + timedelta(hours=3),
            0,
        )
        for data in raw_data.values():
            data["dir"] = (
                data["runwayTypeCode"]
                .fillna(data["defaultRunwayCode"])
                .map(self.RUNWAYDIR, IGNORE)
                .fillna("西")
            )
            data["ddir"] = (
                data["defaultRunwayCode"]
                .map(self.RUNWAYDIR, IGNORE)
                .fillna(data["dir"])
            )

        # CTOT和COBT历史记录
        try:
            data = (
                raw_data["执行_离港"]
                .dropna(subset=["guid", "cobt", "ctot"])
                .set_index("guid", drop=True)
            )
            data = data.loc[data["sobt"] - datetime_now <= timedelta(hours=1)]
            self.HISTORY["CTOT"] = self.HISTORY["CTOT"].join(
                data.loc[data["atot"].isna()]["ctot"].rename(datetime_now),
                how="outer",
            )
            self.HISTORY["COBT"] = self.HISTORY["COBT"].join(
                data.loc[data["pushTime"].isna() & data["atot"].isna()]["cobt"].rename(
                    datetime_now
                ),
                how="outer",
            )
        except Exception as exception:
            tb, exception = format_exc().split("\n", 1)[1], repr(exception)
            self.update_log(
                f"COTO和COBT历史记录追加失败 ({exception[: exception.find('(')]})\n{tb}",
                "warn",
            )

        try:
            monitor = dict((k, timedelta(minutes=v)) for k, v in self.MONITOR.items())
            moni = self.MONIE.copy()
            finished = dict()

            # 进港监控
            data = raw_data["执行_进港"]
            data["portNoLd"] = data["portNoLd"].fillna("").astype(str)
            data.drop(
                data.loc[data["flightNo"].map(lambda x: str(x).endswith("C"))].index,
                inplace=True,
            )
            data["grdAgent"] = data["grdAgent"].map(self.GA, IGNORE).fillna("")

            k = "当日同航班号落地"
            data = data.loc[data["aldt"].notna()].sort_values("aldt")
            data["flightNo_"] = data["flightNo"].map(
                lambda x: x[:-1] if x.endswith("A") else x, IGNORE
            )
            for i in self.MONI.index:
                finished[i] = len(
                    data.loc[
                        (data["flightNo"] == i[0])
                        & (data["sibt"] == i[1])
                        & data["aibt"].isna()
                    ]
                )

            for v in data.loc[
                data.duplicated(subset=["flightNo_", "flightDate"])
                & data["aibt"].isna(),
                ["flightNo", "sibt", "portNoLd", "aldt", "aibt"],
            ].values:
                if (*v[:2], k) not in self.MONI.index:
                    moni.loc[(*v[:2], k)] = [
                        "",
                        "",
                        v[2],
                        "",
                        "当日同航班号第二架次已于{:%H:%M}落地".format(v[3]),
                    ]

            data = (
                raw_data["执行_进港"]
                .loc[
                    series_all(
                        *[
                            raw_data["执行_进港"][i].isna()
                            for i in ("asot", "asdt", "onBridgeTm")
                        ]
                    )
                ]
                .loc[raw_data["执行_进港"]["operationStatusCode"] != "CNCL"]
            )

            k = "进港航班备降外站"
            for i in self.MONI.index:
                finished[i] += len(
                    data.loc[
                        (data["flightNo"] == i[0])
                        & (data["sibt"] == i[1])
                        & (data["operationStatusCode"] == "DIVAL")
                        & data["aldt"].isna()
                        & data["atot"].notna()
                    ]
                )
            for v in data.loc[
                (data["operationStatusCode"] == "DIVAL")
                & data["aldt"].isna()
                & data["atot"].notna(),
                ["flightNo", "sibt"],
            ].values:
                if (*v[:2], k) not in self.MONI.index:
                    moni.loc[(*v[:2], k)] = [
                        "",
                        "",
                        "",
                        "",
                        "进港航班已备降外站",
                    ]

            data = data.loc[data["aldt"].notna()]

            k = "备降航班落地"
            for i in self.MONI.index:
                finished[i] += len(
                    data.loc[
                        (data["flightNo"] == i[0])
                        & (data["sibt"] == i[1])
                        & (data["flightTypeCode"] == "Q/B")
                    ]
                )
            for v in data.loc[
                (data["flightTypeCode"] == "Q/B"),
                ["flightNo", "sibt", "portNoLd", "aldt"],
            ].values:
                if (*v[:2], k) not in self.MONI.index:
                    moni.loc[(*v[:2], k)] = [
                        "",
                        "",
                        v[2],
                        "",
                        "该航班为备降，已于{:%H:%M}落地".format(v[3]),
                    ]

            data = data.loc[data["flightTypeCode"].isin(TYPECODE)]
            for i in self.MONI.index:
                finished[i] += len(
                    data.loc[(data["flightNo"] == i[0]) & (data["sibt"] == i[1])]
                )

            data["进港滑行{}未轮挡"] = (
                datetime_now - data.loc[data["aibt"].isna(), "aldt"]
            )
            data["进港轮挡{}后未下客"] = (
                datetime_now - data.loc[data["aibt"].notna(), "aibt"]
            )
            data["进港轮挡半小时未下客"] = (
                datetime_now - data.loc[data["aibt"].notna(), "aibt"]
            )
            for k in "进港滑行{}未轮挡", "进港轮挡{}后未下客", "进港轮挡半小时未下客":
                if k == "进港滑行{}未轮挡":
                    data["pr"] = data["portNoLd"] + "_" + data["defaultRunwayCode"]
                    data["pr"] = data["pr"].apply(lambda x: ICESNOW.get(str(x)))
                    data["pr"] = data["pr"].replace(np.NAN, 9999999)
                    data["pr"] = pd.to_timedelta(data["pr"], unit="m")
                    for v in data.loc[
                        data[k] >= data["pr"],
                        [
                            "flightNo",
                            "sibt",
                            "portNoLd",
                            "grdAgent",
                            k,
                            "aldt",
                            "runwayTypeCode",
                            "pr",
                        ],
                    ].values:
                        if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                            moni.loc[(*v[:2], k.replace("{}", ""))] = [
                                "",
                                "",
                                v[2],
                                v[3],
                                f"标准：{np.int64(v[7].total_seconds()) // 60}分钟，ATA：{v[5].strftime('%H%M')}，{f'跑道：{v[6]}，' if v[6] else ''}"
                                + k.format(
                                    f"{np.int64(v[4].total_seconds()) // 60}分钟"
                                ),
                            ]
                else:
                    for v in data.loc[
                        data[k] >= monitor[k],
                        ["flightNo", "sibt", "portNoLd", "grdAgent", k],
                    ].values:
                        if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                            moni.loc[(*v[:2], k.replace("{}", ""))] = [
                                "",
                                "",
                                v[2],
                                v[3],
                                k.replace("半小时", "{}").format(
                                    f"{np.int64(v[4].total_seconds()) // 60}分钟"
                                ),
                            ]

            data = raw_data["执行_进港"]
            data = (
                data.loc[data["flightTypeCode"].isin(TYPECODE)]
                .loc[data["aibt"].notna() & data["baggageLastTm"].isna()]
                .loc[data["operationStatusCode"] != "CNCL"]
            ).copy()
            for i in self.MONI.index:
                finished[i] += len(
                    data.loc[(data["flightNo"] == i[0]) & (data["sibt"] == i[1])]
                )
            for k in "进港轮挡{}无首件行李", "进港轮挡{}仍无首件行李":
                data[k] = datetime_now - data["aibt"]
                for v in data.loc[
                    data[k] >= monitor[k],
                    ["flightNo", "sibt", "portNoLd", "grdAgent", k],
                ].values:
                    if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                        moni.loc[(*v[:2], k.replace("{}", ""))] = [
                            "",
                            "",
                            v[2],
                            v[3],
                            k.format(
                                f"{np.int64(v[4].total_seconds()) // 60}分钟".replace(
                                    "首件", "末件"
                                )
                            ),
                        ]

            # 离港监控
            data = raw_data["执行_离港"]
            data = (
                data.loc[data["operationStatusCode"] != "CNCL"]
                .loc[data["atot"].isna()]
                .set_index("guid", drop=True)
            )
            data["routeFst"] = (
                data[["route", "routeCn"]]
                .fillna("")
                .apply(self.airport_name, axis=1)
                .map(lambda x: x.split("-")[0], IGNORE)
            )
            status = {
                "pushTime": "滑行中",
                "acct": "已关舱",
                "aebt": "已登结",
                "asbt": "登机中",
                "sobt": "未登机",
            }
            data["status"] = None
            for k, v in status.items():
                i = data[k] <= datetime_now
                data.loc[i, "status"] = data.loc[i, "status"].fillna(v)
            data["status"] = data["status"].fillna("")
            data["grdAgent"] = data["grdAgent"].map(self.GA, IGNORE).fillna("")
            data["portNoTo"] = data["portNoTo"].fillna("").astype(str)

            for i in self.MONI.index:
                finished[i] += len(
                    data.loc[(data["flightNo"] == i[0]) & (data["sobt"] == i[1])]
                )
            for k, v in finished.items():
                if not v:
                    moni.loc[(*k[:2], "解除")] = ["", "", "", "", ""]

            k = "离港滑行{}未起飞"
            data[k] = datetime_now - data["pushTime"]
            for v in data.loc[
                data[k] >= monitor[k], ["flightNo", "sobt", "routeFst", k]
            ].values:
                if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                    moni.loc[(*v[:2], k.replace("{}", ""))] = [
                        v[2],
                        "",
                        "",
                        "",
                        k.format(f"{np.int64(v[3].total_seconds()) // 60}分钟"),
                    ]

            data = data.loc[data["flightTypeCode"].isin(TYPECODE)].copy()
            data["起飞延误{}"] = datetime_now - data["stot"]
            data["旅客机上等待{}"] = datetime_now - data["acct"]
            for k in "起飞延误{}", "旅客机上等待{}":
                for v in data.loc[
                    data[k] >= monitor[k],
                    [
                        "flightNo",
                        "sobt",
                        "routeFst",
                        "gateNo",
                        "portNoTo",
                        "grdAgent",
                        k,
                        "status",
                    ],
                ].values:
                    if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                        moni.loc[(*v[:2], k.replace("{}", ""))] = [
                            *v[2:6],
                            k.format(f"{np.int64(v[6].total_seconds()) // 60}分钟")
                            + (f"（{v[7]}）" if v[7] else ""),
                        ]

            k = "同航点同时刻相邻登机门"
            data_copy = data.copy()
            for i in data_copy.loc[data_copy["asbt"].isna()].index:
                data_copy[k] = (data_copy["sobt"] - data_copy.loc[i, "sobt"]).apply(
                    lambda x: x.total_seconds() / 60
                )
                data_ = data_copy.loc[
                    (data_copy["routeFst"] == data_copy.loc[i, "routeFst"])
                    & data_copy["gateNo"].map(
                        lambda x: str(x)
                        in self.NEARGATE.get(data_copy.loc[i, "gateNo"], ""),
                        IGNORE,
                    )
                    & (abs(data_copy[k]) < 40)
                    & (data_copy.index != i)
                ]
                v = (
                    data_copy.loc[i, "flightNo"],
                    data_copy.loc[i, "sobt"],
                    k.replace("{}", ""),
                )
                if not data_.empty and v not in self.MONI.index:
                    # text = "航班musdf（STD 1340/机位 127/地点）与同航点航班CZ1000分配至相邻机位且起飞时间接近，请重点关注！"
                    # for num in range(data_.shape[0]):
                    #     text = text + (f"STD {data_["sobt"].iloc[num].strftime("%H%M")} {data_["routeFst"].iloc[num]} " + \
                    #                    f"{data_["gateNo"].iloc[num]} {data_["flightNo"].iloc[num]}，请提示航站楼与地服重点关注")
                    moni.loc[v] = [
                        data_copy.loc[i, "routeFst"],
                        "",
                        data_copy.loc[i, "portNoTo"],
                        "",
                        "与同航点航班{}分配至相邻机位且起飞时间接近，请重点关注".format(
                            "、".join(data_["flightNo"].values)
                        ),
                    ]
                    data_copy = data_copy[data_copy.index != i]

            k = "CTOT逆序{}"
            for i in data.loc[
                data["asbt"].isna() & ((datetime_now - data["sobt"]) > timedelta())
            ].index:
                data[k] = data.loc[i, "ctot"] - data["ctot"]
                data_ = data.loc[
                    (data["sobt"] > data.loc[i, "sobt"])
                    & (data["routeFst"] == data.loc[i, "routeFst"])
                    & data["gateNo"].map(
                        lambda x: str(x)
                        in self.NEARGATE.get(data.loc[i, "gateNo"], ""),
                        IGNORE,
                    )
                    & (data[k] < monitor[k])
                    & (data[k] > timedelta())
                ]
                v = (
                    data.loc[i, "flightNo"],
                    data.loc[i, "sobt"],
                    k.replace("{}", ""),
                )
                if not data_.empty and v not in self.MONI.index:
                    moni.loc[v] = [
                        *data.loc[
                            i, ["routeFst", "gateNo", "portNoTo", "grdAgent"]
                        ].values,
                        "当前CTOT晚于同航线相邻登机口航班{}，请提示航站楼与地服重点关注".format(
                            "、".join(data_["flightNo"].values),
                        ),
                    ]

            data["未上客航班COBT提前{}"] = (
                self.HISTORY["COBT"].iloc[:, -4:-1].max(axis=1) - data["cobt"]
            )
            data["已上客航班CTOT推迟{}"] = data["ctot"] - self.HISTORY["CTOT"].iloc[
                :, -4:-1
            ].min(axis=1)
            data.loc[data["asbt"].notna(), "未上客航班COBT提前{}"] = None
            data.loc[data["asbt"].isna(), "已上客航班CTOT推迟{}"] = None
            for k in "未上客航班COBT提前{}", "已上客航班CTOT推迟{}":
                for v in data.loc[
                    data[k] >= monitor[k],
                    [
                        "flightNo",
                        "sobt",
                        "routeFst",
                        "gateNo",
                        "portNoTo",
                        "grdAgent",
                        k,
                        "status",
                    ],
                ].values:
                    if (*v[:2], k.replace("{}", "")) not in self.MONI.index:
                        moni.loc[(*v[:2], k.replace("{}", ""))] = [
                            *v[2:6],
                            v[7]
                            + k.format(f"{np.int64(v[6].total_seconds()) // 60}分钟")[
                                3:
                            ],
                        ]
                for v in data.loc[data[k] < monitor[k], ["flightNo", "sobt"]].values:
                    if (*v[:2], k.replace("{}", "")) in self.MONI.index and len(
                        self.MONI.loc[(*v[:2], slice(None))]
                    ) == 1:
                        moni.loc[(*v[:2], "解除")] = ["", "", "", "", ""]

            # 更新与同步告警
            if not moni.empty:
                i = [
                    i for i in moni.index if i not in self.MONI.index and i[2] != "解除"
                ]
                if i:
                    self.update_log(
                        "新增{}条航班监控告警：{}".format(
                            len(i),
                            (
                                ", ".join(i[0] for i in i)
                                if "sync" in self.RUNNING
                                or self.RUNNING.get("get_info", 1) == 0
                                else ("\n" if len(i) > 1 else "")
                                + "\n".join(
                                    "{0}/{st} {1:%H%M}{gate}{port}{ga}{dst}，{note}".format(
                                        *i[0],
                                        st=(
                                            "STA"
                                            if "进港" in i[0][2] or "落地" in i[0][2]
                                            else "STD"
                                        ),
                                        gate=(
                                            f"/登机口{i[1]['登机门']}"
                                            if i[1]["登机门"]
                                            else ""
                                        ),
                                        port=(
                                            f"/机位{i[1]['机位']}"
                                            if i[1]["机位"]
                                            else ""
                                        ),
                                        ga=(f"/{i[1]['地服']}" if i[1]["地服"] else ""),
                                        dst=(
                                            f"/{i[1]['目的地']}"
                                            if i[1]["目的地"]
                                            else ""
                                        ),
                                        note=i[1]["描述"],
                                    )
                                    for i in moni.loc[i].iterrows()
                                )
                            ),
                        ),
                        "monitor",
                    )
                self.MONI = (
                    moni.drop(moni.loc[(slice(None), slice(None), "解除")].index)
                    if "解除" in moni.index.get_level_values(2)
                    else moni
                ).combine_first(
                    self.MONI.drop(
                        [
                            i[:2]
                            for i in moni.loc[(slice(None), slice(None), "解除")].index
                        ]
                    )
                    if "解除" in moni.index.get_level_values(2)
                    else self.MONI
                )

                moni = moni.reset_index()
                moni["日期"] = (moni["日期"] - timedelta(hours=8)).map(
                    lambda x: int(x.timestamp() * 1000), IGNORE
                )
                output["告警"] = moni.to_dict("records")

        except Exception as exception:
            tb, exception = format_exc().split("\n", 1)[1], repr(exception)
            self.update_log(
                f"监控告警追加失败 ({exception[: exception.find('(')]})\n{tb}",
                "warn",
            )

        # 运行态势监控
        if "sync" in self.RUNNING:
            hour, hhour = (
                datetime_now - timedelta(hours=1),
                datetime_now - timedelta(minutes=30),
            )
            dirs = ["东", "西"]

            try:
                key = "跑道运行态势"
                dep = raw_data["执行_离港"]
                dep = dep.loc[dep["operationStatusCode"] != "CNCL"].set_index(
                    "guid", drop=True
                )

                arr = raw_data["执行_进港"]
                arr = arr.loc[arr["operationStatusCode"] != "CNCL"].set_index(
                    "guid", drop=True
                )

                output.setdefault("态势", dict())[key] = [
                    {
                        "dir": i,
                        "lt": (
                            "{:%H:%M}".format(dep.loc[(dep["dir"] == i), "atot"].max())
                            if notna(dep.loc[(dep["dir"] == i), "atot"].max())
                            else "-"
                        ),
                        "sep": self.min_sec_format(
                            dep.loc[(dep["dir"] == i) & (dep["atot"] >= hhour), "atot"]
                            .map(
                                lambda x: x
                                - dep.loc[
                                    (dep["dir"] == i) & (dep["atot"] < x), "atot"
                                ].max()
                            )
                            .mean()
                            .total_seconds()
                        ),
                        "ll": (
                            "{:%H:%M}".format(arr.loc[(arr["dir"] == i), "aldt"].max())
                            if notna(arr.loc[(arr["dir"] == i), "aldt"].max())
                            else "-"
                        ),
                        "e": "{} / {}".format(
                            dep.loc[
                                (dep["dir"] == i)
                                & dep["atot"].isna()
                                & (
                                    (dep["pushTime"].notna())
                                    | (dep["ctot"] <= datetime_now + timedelta(hours=1))
                                )
                            ].__len__(),
                            arr.loc[
                                (arr["dir"] == i)
                                & arr["aldt"].isna()
                                & (arr["eldt"] <= datetime_now + timedelta(hours=1))
                            ].__len__(),
                        ),
                    }
                    for i in dirs
                ]
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            try:
                key = "待离港航班态势"
                data = raw_data["执行_离港"]
                data = (
                    data.loc[data["operationStatusCode"] != "CNCL"]
                    .loc[data["atot"].isna()]
                    .set_index("guid", drop=True)
                )
                status = {
                    "pushTime": "push",
                    "acct": "acct",
                    "aebt": "aebt",
                    "asbt": "asbt",
                }
                data["status"] = None
                for k, v in status.items():
                    i = data[k] <= datetime_now
                    data.loc[i, "status"] = data.loc[i, "status"].fillna(v)
                pivot = pivot_table(data, "flightNo", "dir", "status", len).reindex(
                    index=dirs + ["合计"]
                )

                data["operationStatusCode"] = (
                    data["operationStatusCode"]
                    .fillna("")
                    .map(lambda x: 1 if "SLIBK" in x else 0)
                )
                if data["operationStatusCode"].sum():
                    pivot["slibk"] = None
                    for k, v in data.groupby("dir", sort=False):
                        pivot.loc[k, "slibk"] = v["operationStatusCode"].sum()

                if not pivot.empty:
                    pivot.loc["合计"] = pivot.sum(axis=0).values
                    pivot = pivot.astype("Int64").fillna(0)

                v = data.loc[data["deicingType"] == "ICP"]
                if not v.empty:
                    pivot["icp"] = None
                    pivot.loc["合计", "icp"] = "{} / {}".format(
                        len(v.loc[v["aezt"].isna()]), len(v.loc[v["aezt"].notna()])
                    )
                    for i in dirs:
                        k = v.loc[v["dir"] == i]
                        pivot.loc[i, "icp"] = "{} / {}".format(
                            len(k.loc[k["aezt"].isna()]), len(k.loc[k["aezt"].notna()])
                        )

                output.setdefault("态势", dict())[key] = pivot.reset_index().to_dict(
                    "records"
                )
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            try:
                key = "实际/计划 进离港态势"
                h = lambda x: 1 if x <= datetime_now and x > hour else 0
                hh = lambda x: 1 if x <= datetime_now and x > hhour else 0
                pivot = dict()
                for i in dirs:
                    pivot.setdefault(i, dict())["deps"] = (
                        dep.loc[dep["ddir"] == i, "stot"].map(h).sum()
                    )
                    pivot[i]["arrs"] = arr.loc[arr["ddir"] == i, "sldt"].map(h).sum()
                    pivot[i]["hdeps"] = dep.loc[dep["ddir"] == i, "stot"].map(hh).sum()
                    pivot[i]["harrs"] = arr.loc[arr["ddir"] == i, "sldt"].map(hh).sum()
                    pivot[i]["depa"] = dep.loc[dep["dir"] == i, "atot"].map(h).sum()
                    pivot[i]["arra"] = arr.loc[arr["dir"] == i, "aldt"].map(h).sum()
                    pivot[i]["hdepa"] = dep.loc[dep["dir"] == i, "atot"].map(hh).sum()
                    pivot[i]["harra"] = arr.loc[arr["dir"] == i, "aldt"].map(hh).sum()

                for i in pivot[dirs[0]].keys():
                    pivot.setdefault("合计", dict())[i] = sum(
                        [pivot[j][i] for j in dirs]
                    )

                output.setdefault("态势", dict())[key] = [
                    {
                        "dir": k,
                        "dep": f"{v['depa']} / {v['deps']}",
                        "hdep": f"{v['hdepa']} / {v['hdeps']}",
                        "arr": f"{v['arra']} / {v['arrs']}",
                        "harr": f"{v['harra']} / {v['harrs']}",
                    }
                    for k, v in pivot.items()
                ]
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            try:
                key = "机上等待态势"
                today, yesterday = str(today)[:19], str(today - timedelta(1))[:19]
                dep = dep.loc[dep["flightDate"].astype(str) == today].copy()
                dep["waiting"] = dep["atot"].fillna(datetime_now) - dep["acct"]
                k = (
                    (dep["waiting"] >= timedelta(minutes=30)) & dep["atot"].notna()
                ) | ((dep["waiting"] >= timedelta()) & dep["atot"].isna())
                pivot = dict()
                for i in dirs:
                    pivot.setdefault(i, dict())["n"] = dep.loc[
                        (dep["dir"] == i)
                        & (dep["waiting"] >= timedelta())
                        & dep["atot"].isna()
                    ].__len__()
                    pivot[i]["a"] = dep.loc[
                        (dep["dir"] == i)
                        & (dep["waiting"] >= timedelta(minutes=30))
                        & dep["atot"].notna()
                    ].__len__()
                    pivot[i]["s"] = pivot[i]["a"] + pivot[i]["n"]
                    pivot[i]["mean"] = self.min_sec_format(
                        dep.loc[(dep["dir"] == i) & k, "waiting"].mean().total_seconds()
                    )
                    pivot[i]["max"] = self.min_sec_format(
                        dep.loc[dep["dir"] == i, "waiting"].max().total_seconds()
                    )
                    pivot[i]["dir"] = i

                for i in pivot[dirs[0]].keys():
                    if len(i) == 1:
                        pivot.setdefault("整体", dict())[i] = sum(
                            [pivot[j][i] for j in dirs]
                        )
                pivot["整体"]["dir"] = "整体"
                pivot["整体"]["mean"] = self.min_sec_format(
                    dep.loc[k, "waiting"].mean().total_seconds()
                )
                if not dep["waiting"].isna().all():
                    pivot["整体"]["max"] = dep.loc[dep["waiting"].idxmax(), "flightNo"]
                output.setdefault("态势", dict())[key] = list(pivot.values())
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            try:
                key = "当日执行态势"
                dep, arr = raw_data["执行_离港"], raw_data["执行_进港"]
                pivot = DataFrame(columns=["dir", "t", "a", "s", "c", "n"]).set_index(
                    "dir"
                )
                data = dep.loc[dep["flightDate"].astype(str) == today]
                pivot.loc["离港"] = [
                    data.__len__(),
                    data.loc[
                        data["atot"].notna() & (data["operationStatusCode"] != "CNCL")
                    ].__len__(),
                    data.loc[
                        data["atot"].isna() & (data["operationStatusCode"] != "CNCL")
                    ].__len__(),
                    data.loc[data["operationStatusCode"] == "CNCL"].__len__(),
                    dep.loc[
                        (dep["flightDate"].astype(str) == yesterday)
                        & dep["atot"].isna()
                        & (dep["operationStatusCode"] != "CNCL")
                    ].__len__(),
                ]
                data = arr.loc[arr["flightDate"].astype(str) == today]
                pivot.loc["进港"] = [
                    data.__len__(),
                    data.loc[
                        data["aldt"].notna() & (data["operationStatusCode"] != "CNCL")
                    ].__len__(),
                    data.loc[
                        data["aldt"].isna() & (data["operationStatusCode"] != "CNCL")
                    ].__len__(),
                    data.loc[data["operationStatusCode"] == "CNCL"].__len__(),
                    arr.loc[
                        (arr["flightDate"].astype(str) == yesterday)
                        & arr["aldt"].isna()
                        & (arr["operationStatusCode"] != "CNCL")
                    ].__len__(),
                ]
                pivot.loc["合计"] = pivot.sum(axis=0).values
                output.setdefault("态势", dict())[key] = (
                    pivot.astype("Int64").reset_index().to_dict("records")
                )
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            try:
                key = "截至当前执行态势"
                pivot = DataFrame(
                    columns=["dir", "wd", "ed", "d", "wa", "ea", "a"]
                ).set_index("dir")
                deps = dep.loc[
                    (dep["flightDate"].astype(str) == today)
                    & (dep["sobt"] <= datetime_now)
                    & (dep["operationStatusCode"] != "CNCL")
                ]
                arrs = arr.loc[
                    (arr["flightDate"].astype(str) == today)
                    & (arr["sibt"] <= datetime_now)
                    & (arr["operationStatusCode"] != "CNCL")
                ]
                pivot.loc["计划量"] = [
                    deps.loc[deps["ddir"] == "西"].__len__(),
                    deps.loc[deps["ddir"] == "东"].__len__(),
                    deps.__len__(),
                    arrs.loc[arrs["ddir"] == "西"].__len__(),
                    arrs.loc[arrs["ddir"] == "东"].__len__(),
                    arrs.__len__(),
                ]
                pivot.loc["已完成"] = [
                    deps.loc[(deps["ddir"] == "西") & deps["atot"].notna()].__len__(),
                    deps.loc[(deps["ddir"] == "东") & deps["atot"].notna()].__len__(),
                    deps.loc[deps["atot"].notna()].__len__(),
                    arrs.loc[(arrs["ddir"] == "西") & arrs["aldt"].notna()].__len__(),
                    arrs.loc[(arrs["ddir"] == "东") & arrs["aldt"].notna()].__len__(),
                    arrs.loc[arrs["aldt"].notna()].__len__(),
                ]

                deps = dep.loc[
                    (dep["flightDate"].astype(str) == today)
                    & (dep["operationStatusCode"] != "CNCL")
                    & dep["atot"].isna()
                ]
                arrs = arr.loc[
                    (arr["flightDate"].astype(str) == today)
                    & (arr["operationStatusCode"] != "CNCL")
                    & arr["aldt"].isna()
                ]
                pivot.loc["当日余"] = [
                    deps.loc[deps["ddir"] == "西"].__len__(),
                    deps.loc[deps["ddir"] == "东"].__len__(),
                    deps.__len__(),
                    arrs.loc[arrs["ddir"] == "西"].__len__(),
                    arrs.loc[arrs["ddir"] == "东"].__len__(),
                    arrs.__len__(),
                ]

                output.setdefault("态势", dict())[key] = (
                    pivot.astype("Int64").reset_index().to_dict("records")
                )
            except Exception as exception:
                tb, exception = format_exc().split("\n", 1)[1], repr(exception)
                self.update_log(
                    f"{key}运行失败 ({exception[: exception.find('(')]})\n{tb}",
                    "warn",
                )

            retries = 5
            while retries > 0 and "sync" in self.RUNNING and len(output) > 1:
                try:
                    with open(f"{self.AUTO_PATH.get()}/moni.json", "w") as json:
                        dump(output, json)
                    break
                except Exception as exception:
                    retries -= 1
                    if retries <= 0:
                        exception = repr(exception)
                        self.update_log(
                            f"监控告警与运行态势同步失败 ({exception[: exception.find('(')]})",
                            "warn",
                        )

        next_update = self.get_next_update(self.MONI_INTERVAL.get(), 0)
        if i := self.RUNNING.get("get_monitor"):
            i[0] = self.after(
                next_update, threading.Thread(target=self.get_monitor).start
            )
        return output

    def get_alert1(self) -> dict:
        datetime_now = self.datetime_now()
        alert = [-1]
        alert_map = {
            -1: ["无", ""],
            0: ["预警标准"],
            1: ["黄色响应"],
            2: ["橙色响应"],
            3: ["红色响应"],
        }
        alerts = dict()

        try:
            session = self.get_session()
            response = session.post(
                self.URLS["大面积航延"],
                headers=self.HEADER,
                timeout=10,
            )
            if response.status_code != 200:
                raise requests.ConnectionError(f"响应异常代码{response.status_code}")
            else:
                response = response.json().get("data")
                alerts["预警条件"] = response.get("data1")
                alerts["响应条件一"] = response.get("data2")
                alerts["响应条件二"] = response.get("data5")
                alerts["响应条件三"] = response.get("data6")
                for i in alerts.values():
                    assert isinstance(i, int)

        except Exception:
            self.update_log("大面积航延接口数据获取失败，采用计算值", "warn")

            # 大面积航班延误预警
            data = self.get_data("航班", outFlightTypeCode=TYPECODE)
            data = data.loc[data["outAtot"].isna() & (data["outEstripStatus"] != "DEP")]
            alerts["预警条件"] = data.loc[
                data["outCtot"]
                .fillna(datetime_now)
                .map(lambda x: x if x >= datetime_now else datetime_now)
                - data["outSobt"]
                >= timedelta(hours=1.5)
            ].__len__()

            # 大面积航班延误响应检测条件一
            alerts["响应条件一"] = data.loc[
                datetime_now - data["outStot"] > timedelta(hours=1)
            ].__len__()

            # 大面积航班延误响应检测条件二、三
            data = self.get_data("执行_离港", flightTypeCode=TYPECODE)
            alerts["响应条件二"] = (
                data.loc[data["operationStatusCode"] == "CNCL"]
                .loc[data["sobt"] >= datetime_now]
                .loc[data["sobt"] < datetime_now + timedelta(hours=2)]
                .__len__()
            )
            alerts["响应条件三"] = (
                data.loc[data["operationStatusCode"] == "CNCL"]
                .loc[data["sobt"] >= datetime_now]
                .loc[data["sobt"] < datetime_now + timedelta(hours=1)]
                .__len__()
            )
            data = self.get_data("执行_进港", flightTypeCode=TYPECODE)
            alerts["响应条件三"] += (
                data.loc[data["operationStatusCode"] != "CNCL"]
                .loc[data["eldt"] >= datetime_now]
                .loc[data["eldt"] < datetime_now + timedelta(hours=1)]
                .__len__()
            )

        if alerts["预警条件"] >= 10:
            alert_map[0].append(
                f"离港航班CTOT≥STD+1.5小时达10架次及以上（当前{alerts['预警条件']}架次）"
            )
            alert.append(0)

        if alerts["响应条件一"] >= 25:
            for i, j in (50, 3), (40, 2), (25, 1):
                if alerts["响应条件一"] >= i:
                    alert_map[j].append(
                        f"离港航班不正常时长超1小时以上且未起飞客运航班达{i}架次（按起飞延误，当前{alerts['响应条件一']}架次）"
                    )
                    alert.append(j)
                    break

        if alerts["响应条件二"] >= 15:
            for i, j in (30, 3), (25, 2), (15, 1):
                if alerts["响应条件二"] >= i:
                    alert_map[j].append(
                        f"未来2小时内出港客运航班航司决策临时取消达{i}架次（当前{alerts['响应条件二']}架次）"
                    )
                    alert.append(j)
                    break
        if datetime_now.hour >= 22 or datetime_now.hour < 6:
            if alerts["响应条件三"] >= 60:
                for i, j in (75, 3), (70, 2), (60, 1):
                    if alerts["响应条件三"] >= i:
                        alert_map[j].append(
                            f"22时至次日6时任一小时进港航班和离港取消航班合计达{i}架次（当前{alerts['响应条件三']}架次）"
                        )
                        alert.append(j)
                        break
        else:
            alerts.pop("响应条件三", 0)

        alert = alert_map[max(alert)]
        alerts["大面积航延"] = alert[0]
        alerts["启动标准"] = "；".join(alert[1:])
        return alerts

    def get_alert(self) -> dict:
        datetime_now = self.datetime_now()
        timetuple = datetime_now.timetuple()
        today = datetime(*timetuple[:3])
        last_hour = datetime_now - timedelta(hours=1)
        one_hour = datetime_now + timedelta(hours=1)
        two_hour = datetime_now + timedelta(hours=2)
        alert = [-1]
        alert_map = {
            -1: ["无", ""],
            0: [
                "视情况启动调时",
                "【视情况启动调时】兴效能已自动检测到满足立即启动调时标准（先前为无）：启动条件（",
            ],
            1: [
                "立即启动调时",
                "【立即启动调时】兴效能已自动检测到满足立即启动调时标准（先前为无/视情况启动调时）：启动条件（",
            ],
        }
        alerts = dict()

        try:
            session = self.get_session()
            response = session.post(
                self.URLS["大面积航延"],
                headers=self.HEADER,
                timeout=10,
            )
            if response.status_code != 200:
                raise requests.ConnectionError(f"响应异常代码{response.status_code}")
            else:
                response = response.json().get("data")
                alerts["预警条件"] = response.get("data1")
                alerts["响应条件一"] = response.get("data2")
                alerts["响应条件二"] = response.get("data5")
                alerts["响应条件三"] = response.get("data6")
                for i in alerts.values():
                    assert isinstance(i, int)

        except Exception:
            self.update_log("大面积航延接口数据获取失败，采用计算值", "warn")

            # 大面积航班延误预警
            data = self.get_data("航班", outFlightTypeCode=TYPECODE)
            data = data.loc[data["outAtot"].isna() & (data["outEstripStatus"] != "DEP")]
            alerts["预警条件"] = data.loc[
                data["outCtot"]
                .fillna(datetime_now)
                .map(lambda x: x if x >= datetime_now else datetime_now)
                - data["outSobt"]
                >= timedelta(hours=1.5)
            ].__len__()

            # 大面积航班延误响应检测条件一
            alerts["响应条件一"] = data.loc[
                datetime_now - data["outStot"] > timedelta(hours=1)
            ].__len__()

            # 大面积航班延误响应检测条件二、三
            data = self.get_data("执行_离港", flightTypeCode=TYPECODE)
            alerts["响应条件二"] = (
                data.loc[data["operationStatusCode"] == "CNCL"]
                .loc[data["sobt"] >= datetime_now]
                .loc[data["sobt"] < datetime_now + timedelta(hours=2)]
                .__len__()
            )
            alerts["响应条件三"] = (
                data.loc[data["operationStatusCode"] == "CNCL"]
                .loc[data["sobt"] >= datetime_now]
                .loc[data["sobt"] < datetime_now + timedelta(hours=1)]
                .__len__()
            )
            data = self.get_data("执行_进港", flightTypeCode=TYPECODE)
            alerts["响应条件三"] += (
                data.loc[data["operationStatusCode"] != "CNCL"]
                .loc[data["eldt"] >= datetime_now]
                .loc[data["eldt"] < datetime_now + timedelta(hours=1)]
                .__len__()
            )
        for i in range(1, 6):
            match i:
                case 1:
                    if alerts["预警条件"] >= 10:
                        alert_map[0].append("1. 大兴机场启动大面积航班延误预警")

                    if (
                        (alerts["响应条件一"] >= 50)
                        | (alerts["响应条件二"] >= 30)
                        | (
                            (datetime_now.hour >= 22 or datetime_now.hour < 6)
                            & alerts["响应条件三"]
                            >= 75
                        )
                    ):
                        alert_map[1].append("1. 大兴机场启动大面积航班延误红色响应")
                case 2:
                    try:
                        if datetime_now.hour >= 7 or datetime_now.hour < 24:

                            def neizhi(*args):
                                data = self.get_data("航班", outFlightTypeCode=TYPECODE)

                                # 进港的情况
                                inaldt_data = data.loc[
                                    (data[args[0]] - datetime.now())
                                    < timedelta(minutes=0)
                                ].sort_values(by=args[0], ascending=False)
                                inaldt_time = inaldt_data.iloc[0][args[0]]

                                inaldt_now_sldt = data.loc[
                                    data[args[2]] > inaldt_time
                                ].__len__()
                                inaldt_now_stot = data.loc[
                                    data[args[3]] > inaldt_time
                                ].__len__()

                                # 出港的情况
                                outatot_data = data.loc[
                                    (data[args[1]] - datetime.now())
                                    < timedelta(minutes=0)
                                ].sort_values(by=args[1], ascending=False)
                                outatot_time = outatot_data.iloc[0][args[1]]

                                outatot_now_sldt = data.loc[
                                    data[args[2]] > outatot_time
                                ].__len__()
                                outatot_now_stot = data.loc[
                                    data[args[3]] > outatot_time
                                ].__len__()
                                return [
                                    inaldt_time,
                                    inaldt_now_sldt,
                                    inaldt_now_stot,
                                    outatot_time,
                                    outatot_now_sldt,
                                    outatot_now_stot,
                                ]

                            re_list = neizhi(
                                *["inAldt", "outAtot", "inSldt", "outStot"]
                            )

                            if (
                                (re_list[0] < datetime_now - timedelta(minutes=20))
                                & ((re_list[1] > 9) | (re_list[1] + re_list[2] > 12))
                            ) | (
                                (re_list[3] < datetime_now - timedelta(minutes=20))
                                & ((re_list[5] > 9) | (re_list[4] + re_list[5] > 12))
                            ):
                                alert_map[0].append(
                                    f"2.本场预计或目前已停止起飞或落地20分钟以上，且直接影响航班总量大于其高峰小时容量的20%;"
                                )

                            if (
                                (re_list[0] < datetime_now - timedelta(minutes=60))
                                & ((re_list[1] > 24) | (re_list[1] + re_list[2] > 31))
                            ) | (
                                (re_list[3] < datetime_now - timedelta(minutes=60))
                                & ((re_list[5] > 24) | (re_list[4] + re_list[5] > 31))
                            ):
                                alert_map[1].append(
                                    f"2.本场预计或目前已停止起飞或落地1小时以上，且直接影响航班总量大于其高峰小时容量的50%"
                                )
                    except Exception:
                        self.update_log(
                            "2.本场预计或目前已停止起飞或落地1小时以上，且直接影响航班总量大于其高峰小时容量的50%",
                            "warn",
                        )
                case 3:
                    data = self.get_data("执行_离港", flightTypeCode=TYPECODE)
                    tis = {}

                    for ti in ["stot", "atot"]:
                        tis[f"{ti}_last_data"] = (
                            data.loc[data[ti] < datetime_now]
                            .loc[data[ti] > last_hour]
                            .__len__()
                        )
                        tis[f"{ti}_one_data"] = (
                            data.loc[data[ti] < one_hour]
                            .loc[data[ti] > datetime_now]
                            .__len__()
                        )
                        tis[f"{ti}_two_data"] = (
                            data.loc[data[ti] < two_hour]
                            .loc[data[ti] > one_hour]
                            .__len__()
                        )

                    for i in tis.values():
                        assert isinstance(i, int)

                    if (
                        tis["atot_last_data"]
                        < tis["stot_last_data"]
                        & (tis["stot_one_data"] - tis["atot_last_data"])
                        > 0 & (tis["stot_two_data"] - tis["atot_last_data"])
                        > 0
                        & (tis["stot_one_data"] + tis["stot_two_data"])
                        - tis["atot_last_data"] * 2
                        > 9
                    ):
                        alert_map[1].append(
                            f"3.本场全向或特定方向，预计后续航班计划连续2小时超出保障能力上限，且溢出航班总量大于其高峰小时容量的20%"
                        )

                    if (
                        tis["atot_last_data"]
                        < tis["stot_last_data"]
                        & (tis["stot_one_data"] - tis["atot_last_data"])
                        / tis["atot_last_data"]
                        > 0.2
                        & (tis["stot_two_data"] - tis["atot_last_data"])
                        / tis["atot_last_data"]
                        > 0.2
                        & (tis["stot_one_data"] + tis["stot_two_data"])
                        - tis["atot_last_data"] * 2
                        > 24
                    ):
                        alert_map[1].append(
                            f"3.本场全向或特定方向，预计后续航班计划连续2小时超出保障能力上限20%以上，且溢出航班总量大于其高峰小时容量的50%"
                        )
                case 4:
                    last_two_hour = datetime_now - timedelta(minutes=2)
                    data = self.get_data(
                        "航班", outFlightTypeCode=TYPECODE[:3]
                    )  # 筛选航班_性质为W/Z,C/B,L/W的航班
                    takeoffs = {}
                    count = 0
                    for start_time, end_time in {
                        last_two_hour: last_hour,
                        last_hour: datetime_now,
                    }.items():
                        data = (
                            data.loc[data["outSobt"] >= start_time]
                            .loc[data["outSobt"] < end_time]
                            .copy()
                        )
                        data = self.get_data("执行_进港", outFlightTypeCode=TYPECODE)

                        for route in ["ALL", "CAN,SZX"]:
                            if route == "ALL":
                                re_data = data
                            else:
                                for fly in route.slplit(","):
                                    re_data += data[data["outRoute"] == route]
                            if re_data.__len__():
                                departing = re_data.loc[
                                    re_data["outAtot"].isna()
                                    | (re_data["outAtot"] > end_time)
                                ]
                                takeoff_departing_delayed = departing.loc[
                                    departing["outStot"] < end_time
                                ].__len__()
                                re_data.drop(departing.index, inplace=True)
                                # 起飞正常性
                                try:
                                    takeoff_departed = re_data.loc[
                                        re_data["outAtot"] <= re_data["outStot"]
                                    ].__len__()
                                    takeoff_delayed = (
                                        re_data.__len__() - takeoff_departed
                                    )
                                    takeoff = takeoff_departed / (
                                        takeoff_departed
                                        + takeoff_departing_delayed
                                        + takeoff_delayed
                                    )
                                    takeoffs[f"{route}{count}"] = takeoff
                                except ZeroDivisionError:
                                    takeoffs[f"{route}{count}"] = 0
                            else:
                                takeoffs[f"{route}{count}"] = 0
                        count += 1

                    for i in takeoffs.values():
                        assert isinstance(i, int)

                    if (
                        (
                            takeoffs["ALL0"]
                            < 0.5 & takeoffs["ALL1"]
                            < 0.5 & takeoffs["ALL0"]
                            != 0 & takeoffs["ALL1"]
                            != 0
                        )
                        | takeoffs["CAN,SZX0"]
                        < 0.5
                        & takeoffs[
                            "CAN,SZX1" & takeoffs["CAN,SZX0"] != 0,
                            takeoffs["CAN,SZX0"] != 0,
                        ]
                        < 0.5
                    ):
                        alert_map[0].append(
                            f"4.本场全向或特定方向，预计离港航班起飞正常率连续2小时低于50%"
                        )

                    data = self.get_data("执行_离港", flightTypeCode=TYPECODE)
                    last_atot_len = (
                        data.loc[data["atot"] > last_two_hour]
                        .loc[data["atot"] < datetime_now]
                        .__len__()
                    )
                    last_stot_atot_len = (
                        data.loc[data["atot"] > last_two_hour]
                        .loc[data["atot"] < datetime_now]
                        .loc[data["atot"].isna()]
                        .__len__()
                    )
                    three_hour = (
                        data.loc[data["stot"] > datetime_now]
                        .loc[data["stot"] > datetime_now + timedelta(hours=3)]
                        .__len__()
                    )
                    for i in [last_atot_len, last_stot_atot_len, three_hour]:
                        assert isinstance(i, int)

                    if last_atot_len < (last_stot_atot_len + three_hour) / 3:
                        alert_map[1].append(
                            f"4.本场全向或特定方向，预计离港积压航班（判断逻辑）在3小时内无法消化完毕"
                        )

                case 5:
                    # 累计备降航班架次
                    data = self.get_data("执行_进港", outFlightTypeCode=TYPECODE)
                    dival_count = (
                        data.loc[data["operationStatusCode"] == "DIVAL"]
                        .loc[data["operationStatusCode"] != "CNCL"]
                        .__len__()
                    )
                    if dival_count > 10:
                        alert_map[0].append(f"5.本场进港航班当日累计备降超过10班")

                    qb_data = self.get_data("执行_进港", flightTypeCode=["Q/B"])
                    qb_count = 0
                    if qb_data.__len__():
                        qb_list = qb_data["sibt"].dt.hour.to_list()
                        qb_count = max([qb_list.count(item) for item in set(qb_list)])

                    if dival_count > 20 or qb_count > 10:
                        alert_map[1].append(
                            f"5.本场进港航班1小时内备降超过10班，或当日累计备降超过20班"
                        )

        alert = alert_map[max(alert)]
        alerts["大面积航延"] = alert[0]
        alerts["启动标准"] = "；".join(alert[1:])
        return alerts

        # 四地十场

    def fourlocation_tenvenuse(
        self, days: int = 0
    ) -> tuple[str, str, str, str, int, int]:
        today, datetime_now = self.today(days), self.datetime_now()
        start_time = datetime_now - timedelta(minutes=1)
        end_time = datetime_now + timedelta(minutes=1)
        data = self.get_data(
            "航班", outFlightTypeCode=TYPECODE[:3]
        )  # 筛选航班_性质为W/Z,C/B,L/W的航班
        data = (
            data.loc[data["outSobt"] >= start_time]
            .loc[data["outSobt"] < end_time]
            .copy()
        )  #
        takeoffs = {}
        # 广州、重庆、成都双流、成都天府、杭州、上海浦东、上海虹桥、深圳
        for route in ["CAN", "CKG", "CTU", "TFU", "HGH", "PVG", "SHA", "SZX"]:
            data_route = data[data["outRoute"] == route]
            if data_route.__len__():
                departing = data_route.loc[
                    data_route["outAtot"].isna()
                    | (data_route["outAtot"] > datetime_now)
                ]
                takeoff_departing_delayed = departing.loc[
                    departing["outStot"] < datetime_now
                ].__len__()
                data_route.drop(departing.index, inplace=True)
                # 起飞正常性
                try:
                    takeoff_departed = data_route.loc[
                        data_route["outAtot"] <= data_route["outStot"]
                    ].__len__()
                    takeoff_delayed = data_route.__len__() - takeoff_departed
                    takeoff = takeoff_departed / (
                        takeoff_departed + takeoff_departing_delayed + takeoff_delayed
                    )
                    takeoff = (
                        "100%"
                        if takeoff == 1
                        else "{:.2f}%".format(round(takeoff * 100, 2))
                    )
                    takeoffs[route] = takeoff
                except ZeroDivisionError:
                    takeoffs[route] = ""
            else:
                takeoffs[route] = ""
        return takeoffs

    def get_rates(self, days: int = 0) -> tuple[str, str, str, str, int, int]:
        today, datetime_now = self.today(days), self.datetime_now()
        tommorow = today + timedelta(1)
        data = self.get_data("航班", outFlightTypeCode=TYPECODE[:3])
        data = data.loc[data["outSobt"] >= today].loc[data["outSobt"] < tommorow].copy()

        # 始发航班正常性
        initial_data = data.loc[data["outIsinitial"]].copy()
        if len(initial_data):
            try:
                initial_departing = initial_data.loc[
                    initial_data["outAtot"].isna()
                    | (initial_data["outAtot"] > datetime_now)
                ]
                initial_departing_delayed = initial_departing.loc[
                    initial_data["outStot"] < datetime_now
                ].__len__()
                initial_data.drop(initial_departing.index, inplace=True)
                initial_departed = initial_data.loc[
                    initial_data["outAtot"] <= initial_data["outStot"]
                ].__len__()
                initial_delayed = initial_data.__len__() - initial_departed
                initial = initial_departed / (
                    initial_departed + initial_departing_delayed + initial_delayed
                )
                initial = (
                    "100%"
                    if initial == 1
                    else "{:.2f}%".format(round(initial * 100, 2))
                )
            except ZeroDivisionError:
                initial = ""
        else:
            initial = ""

        # 放行正常性
        departing = data.loc[data["outAtot"].isna() | (data["outAtot"] > datetime_now)]
        clearance_departing_delayed = (
            departing.loc[departing["inAldt"].notna() | departing["inSldt"].isna()]
            .loc[departing["outLastTot"] < datetime_now]
            .__len__()
        )
        takeoff_departing_delayed = departing.loc[
            departing["outStot"] < datetime_now
        ].__len__()
        data.drop(departing.index, inplace=True)

        try:
            clearance_departed = data.loc[
                data["outAtot"] <= data["outLastTot"]
            ].__len__()
            clearance_delayed = (
                data.loc[data["inAldt"].notna() | data["inSldt"].isna()].__len__()
                - clearance_departed
            )
            clearance = clearance_departed / (
                clearance_departed + clearance_departing_delayed + clearance_delayed
            )
            clearance = (
                "100%"
                if clearance == 1
                else "{:.2f}%".format(round(clearance * 100, 2))
            )
        except ZeroDivisionError:
            clearance = ""

        # 起飞正常性
        try:
            takeoff_departed = data.loc[data["outAtot"] <= data["outStot"]].__len__()
            takeoff_delayed = data.__len__() - takeoff_departed
            takeoff = takeoff_departed / (
                takeoff_departed + takeoff_departing_delayed + takeoff_delayed
            )
            takeoff = (
                "100%" if takeoff == 1 else "{:.2f}%".format(round(takeoff * 100, 2))
            )
        except ZeroDivisionError:
            takeoff = ""

        # 进港正常性
        data = self.get_data("执行_进港", flightTypeCode=TYPECODE[:3])
        data = data.loc[data["flightDate"].astype(str) == str(today)[:19]].copy()
        incncl = data.loc[data["operationStatusCode"] == "CNCL"].index
        data.drop(incncl, inplace=True)

        arriving = data.loc[data["aldt"].isna() | (data["aldt"] > datetime_now)]
        landing_arriving_delayed = arriving.loc[
            data["sibt"] + timedelta(minutes=10) < datetime_now
        ].__len__()
        data.drop(arriving.index, inplace=True)
        try:
            landing_arrived = data.loc[
                data["aldt"] <= data["sibt"] + timedelta(minutes=10)
            ].__len__()
            landing_delayed = data.__len__() - landing_arrived
            landing = landing_arrived / (
                landing_arrived + landing_arriving_delayed + landing_delayed
            )
            landing = (
                "100%" if landing == 1 else "{:.2f}%".format(round(landing * 100, 2))
            )
        except ZeroDivisionError:
            landing = ""

        # 离港取消
        data = self.get_data("执行_离港")
        data = data.loc[data["flightDate"].astype(str) == str(today)[:19]]
        outcncl = data.loc[data["operationStatusCode"] == "CNCL"].index

        return initial, clearance, takeoff, landing, outcncl.size, incncl.size

    def auto_msg(self, path: tk.StringVar = None):
        try:
            warnings = []
            i, j, mass_delay = (para.get() for para in self.msg_para)
            delay = "outLastTot"  # self.DELAY.get()
            if i:
                prefix = ["除霜", "【霜天气保障专项通报】"]
            if j:
                prefix = ["除冰", "【冰雪天气保障专项通报】"]
            deice = i or j
            yesterday = self.yesterday.get()
            day_prefix = "昨日" if yesterday else "当日"

            today, datetime_now = (
                self.today(-1 if yesterday else 0),
                self.datetime_now(),
            )
            tommorow = today + timedelta(1)
            msg = end_time = "截至{}:{:02}，".format(
                datetime_now.hour, datetime_now.minute
            )
            self.update_status({"综合效率席短信": "进行中"}, set_bar=True)

            raw_flight, raw_delay, raw_depart, raw_arrival = self.get_data(
                "航班", "延误", "执行_离港", "执行_进港"
            )
            raw_flight = (
                raw_flight.loc[raw_flight["outSobt"] >= today]
                .loc[raw_flight["outSobt"] < tommorow]
                .copy()
            )
            raw_delay["delayed"] = (
                raw_delay["outAtot"].fillna(datetime_now) - raw_delay[delay]
            )
            raw_delay = (
                raw_delay.loc[raw_delay["outSobt"] >= today]
                .loc[raw_delay["outSobt"] < tommorow]
                .loc[raw_delay[delay] < datetime_now]
                .loc[raw_delay["delayed"] > timedelta()]
                .copy()
            )
            raw_depart = raw_depart.loc[
                raw_depart["operationStatusCode"] != "CNCL"
            ].copy()
            raw_arrival = raw_arrival.loc[
                raw_arrival["operationStatusCode"] != "CNCL"
            ].copy()

            if raw_flight.__len__():
                (
                    initial,
                    clearance,
                    takeoff,
                    _,
                    outCanceled,
                    _,
                ) = self.get_rates(-1 if yesterday else 0)

                # 短信编排
                if initial == clearance and clearance == takeoff:
                    msg_rates = (
                        f"{day_prefix}始发、放行、起飞正常率均为" + takeoff
                        if initial
                        else f"{day_prefix}无客班正常率数据"
                    )
                elif initial == clearance:
                    msg_rates = (
                        f"{day_prefix}始发、放行正常率均为{initial}，起飞正常率为"
                        + takeoff
                    )
                elif clearance == takeoff:
                    msg_rates = (
                        (
                            f"{day_prefix}始发正常率为" + initial
                            if initial
                            else f"{day_prefix}无始发航班"
                        )
                        + "，放行、起飞正常率均为"
                        + takeoff
                    )
                else:
                    msg_rates = (
                        (
                            f"{day_prefix}始发正常率为" + initial
                            if initial
                            else f"{day_prefix}无始发航班"
                        )
                        + f"，放行正常率为{clearance}，起飞正常率为"
                        + takeoff
                    )
                msg_rates += "。"
                msg += msg_rates

                if len(raw_delay):
                    delayed = raw_delay.copy()
                    delayed.drop(
                        delayed.loc[
                            delayed["inAldt"].isna() & delayed["inSldt"].notna()
                        ]
                        .loc[(delayed["outSobt"] - delayed["inSibt"]) > timedelta()]
                        .index,
                        inplace=True,
                    )
                    delay_exclude = delayed.loc[
                        delayed["subDelayReason"] == self.PRIMARY["00"]
                    ].index
                    delayed.drop(delay_exclude, inplace=True)
                    delay_exclude = delay_exclude.size
                    delayed_departing = delayed.loc[
                        delayed["outAtot"].isna() | (delayed["outAtot"] > datetime_now)
                    ].index
                    delay_error = len(delayed)
                else:
                    delay_error = delay_exclude = 0

                if delay_error:
                    msg += f"放行延误{len(delayed)}架次（"
                    if delayed_departing.size:
                        i = len(delayed) - delayed_departing.size
                        msg += (
                            "未起飞），"
                            if len(delayed) == 1
                            else (
                                f"其中{i}架次已起飞），平均"
                                if i
                                else "均未起飞），平均"
                            )
                        )
                    else:
                        msg += "已起飞），" if len(delayed) == 1 else "均已起飞），平均"
                    msg += "延误时间{}，".format(
                        self.min_sec_format(
                            delayed[("delayed" if "未起飞" in msg else delay + "Delay")]
                            .mean()
                            .total_seconds()
                        )
                    )
                    msg_finalized = msg

                    # 筛选军事活动与天气
                    military = delayed.loc[delayed["priDelayReason"] == "军事活动"]
                    weather = delayed.loc[delayed["priDelayReason"] == "天气"]
                    military_other = military.loc[military["addDelayReason"] == "外站"]
                    weather_other = weather.loc[weather["addDelayReason"] == "外站"]
                    military_local = military.loc[
                        military["addDelayReason"] == "本场"
                    ].__len__()
                    weather_local = weather.loc[
                        weather["addDelayReason"] == "本场"
                    ].__len__()
                    count_mo, count_wo = (
                        military_other.__len__(),
                        weather_other.__len__(),
                    )

                    i = military.__len__() - count_mo - military_local
                    if i:
                        status = (
                            f"军事活动原因延误的航班中，{i}架次三级原因未判定本场或外站"
                        )
                        warnings.append(status)
                    i = weather.__len__() - count_wo - weather_local
                    if i:
                        status = (
                            f"天气原因延误的航班中，{i}架次三级原因未判定本场或外站"
                        )
                        warnings.append(status)

                    delay_error -= count_mo + count_wo
                    msg_mo = msg_wo = msg_a = ""

                    msg += f"因外站/航路降效导致延误{count_mo + count_wo}架次【"
                    msg += f"其他用户活动原因延误{count_mo}架次（"
                    for i in self.REGION.values():
                        count = military_other.loc[
                            military_other["outAirportRegionCn"] == i
                        ].__len__()
                        msg += f"{i}延误{count}架次，"
                        if count:
                            msg_mo += f"{i}延误{count}架次，"
                    msg = msg[:-1] + "），"
                    msg += f"外站天气延误{count_wo}架次（"
                    for i in self.REGION.values():
                        count = weather_other.loc[
                            weather_other["outAirportRegionCn"] == i
                        ].__len__()
                        msg += f"{i}延误{count}架次，"
                        if count:
                            msg_wo += f"{i}延误{count}架次，"
                    msg = msg[:-1] + "）】，"

                    if count_mo + count_wo:
                        msg_finalized += (
                            f"因外站/航路降效导致延误{count_mo + count_wo}架次【"
                        )
                        if count_mo:
                            msg_finalized += (
                                f"其他用户活动原因延误{count_mo}架次（" + msg_mo
                            )
                            msg_finalized = msg_finalized[:-1] + "），"
                        if count_wo:
                            msg_finalized += f"外站天气延误{count_wo}架次（" + msg_wo
                            msg_finalized = msg_finalized[:-1] + "），"
                        msg_finalized = msg_finalized[:-1] + "】，"

                    # 本场天气
                    delay_error -= weather_local
                    msg += f"因本场/终端区特殊天气延误{weather_local}架次，"
                    if weather_local:
                        msg_finalized += (
                            f"因本场/终端区特殊天气延误{weather_local}架次，"
                        )

                    # 本场军事活动
                    delay_error -= military_local
                    msg += f"因本场其他用户活动原因延误{military_local}架次，"
                    if military_local:
                        msg_finalized += (
                            f"因本场其他用户活动原因延误{military_local}架次，"
                        )

                    # 航空公司
                    airline = delayed.loc[delayed["priDelayReason"] == "航空公司"]
                    count_a = len(airline)
                    msg += f"因航空公司原因延误{count_a}架次（"

                    for i in self.AIRLINE:
                        count = airline.loc[airline["subDelayReason"] == i].__len__()
                        i = i.replace("原因", "")
                        msg += f"{i}原因延误{count}架次，"
                        if count:
                            delay_error -= count
                            msg_a += f"{i}原因延误{count}架次，"
                    msg = msg[:-1] + "），"

                    if count_a:
                        msg_finalized += f"因航空公司原因延误{count_a}架次（" + msg_a
                        msg_finalized = msg_finalized[:-1] + "），"

                    # 旅客
                    passenger = len(delayed.loc[delayed["priDelayReason"] == "旅客"])
                    delay_error -= passenger
                    msg += f"因旅客原因延误{passenger}架次，"
                    if passenger:
                        msg_finalized += f"因旅客原因延误{passenger}架次，"

                    # 公共安全
                    security = len(delayed.loc[delayed["priDelayReason"] == "公共安全"])
                    delay_error -= security
                    msg += f"因公共安全原因延误{security}架次。"
                    if security:
                        msg_finalized += f"因公共安全原因延误{security}架次，"

                    msg_finalized = msg_finalized[:-1] + "。"

                    msg_delayed = msg_finalized[msg_finalized.find("放行延误") :]

                    if delay_error:
                        status = f"{delay_error}架次延误原因判定出错，请检查信息"
                        warnings.append(status)

                    if delay_exclude:
                        status = f"{delay_exclude}架次延误原因未判定，已在信息中排除"
                        warnings.append(status)

                elif delay_exclude:
                    msg_delayed = f"放行延误{delay_exclude}架次（延误原因未判定）。"
                    msg += msg_delayed
                    msg_finalized = msg
                    status = (
                        f"{delay_exclude}架次延误原因未判定，已在信息中排除，请检查"
                    )
                    warnings.append(status)
                else:
                    msg_delayed = "无放行延误航班。"
                    msg += msg_delayed
                    msg_finalized = msg

            else:
                msg_delayed = msg_rates = ""
                msg += f"{day_prefix}无客运航班计划起飞。"
                msg_finalized = msg

            extra_msg = ""

            if mass_delay and len(raw_delay):
                extra_msg += end_time + msg_rates
                start, end = today, min(datetime_now, today + timedelta(1))

                dep_s = (
                    raw_depart.loc[raw_depart["sobt"] >= start]
                    .loc[raw_depart["sobt"] < end]
                    .loc[raw_depart["flightTypeCode"] != "F/H"]
                    .loc[raw_depart["flightTypeCode"] != "Q/B"]
                )
                arr_s = (
                    raw_arrival.loc[raw_arrival["sibt"] >= start]
                    .loc[raw_arrival["sibt"] < end]
                    .loc[raw_arrival["flightTypeCode"] != "F/H"]
                    .loc[raw_arrival["flightTypeCode"] != "Q/B"]
                )
                dep_a = dep_s.loc[dep_s["atot"] <= datetime_now].__len__()
                arr_a = arr_s.loc[arr_s["aldt"] <= datetime_now].__len__()

                extra_msg += (
                    "计划执行{}架次，已执行{}架次（出港{}架次，进港{}架次）{}".format(
                        dep_s.__len__() + arr_s.__len__(),
                        dep_a + arr_a,
                        dep_a,
                        arr_a,
                        (
                            f"，取消{outCanceled}架次。"
                            if deice and len(raw_flight)
                            else "。"
                        ),
                    )
                )

                # 机上等待时间
                waiting = (
                    datetime_now
                    - raw_delay.loc[
                        raw_delay["outAtot"].isna()
                        | (raw_delay["outAtot"] > datetime_now)
                    ]
                    .loc[raw_delay["outAcct"] <= datetime_now]
                    .loc[raw_delay["delayed"] > timedelta()]["outAcct"]
                )
                extra_msg += (
                    msg_delayed
                    + "客齐等待航班{}架次，旅客机上等待1-2小时航班{}架次，旅客机上等待2小时以上航班{}架次。".format(
                        raw_flight.loc[
                            raw_flight["outAobt"].isna()
                            | (raw_flight["outAobt"] > datetime_now)
                        ]
                        .loc[raw_flight["outAebt"] <= datetime_now]
                        .__len__(),
                        waiting.loc[waiting > timedelta(hours=1)]
                        .loc[waiting <= timedelta(hours=2)]
                        .__len__(),
                        waiting.loc[waiting > timedelta(hours=2)].__len__(),
                    )
                )

                extra_msg += "平均机上等待时长{}。".format(
                    self.min_sec_format(
                        (
                            raw_flight["outAtot"].fillna(datetime_now)
                            - raw_flight["outAcct"]
                        )
                        .mean()
                        .total_seconds()
                    )
                )

                start = datetime_now - timedelta(hours=1)
                dep_s = (
                    raw_depart.loc[raw_depart["sobt"] >= start]
                    .loc[raw_depart["sobt"] < datetime_now]
                    .loc[raw_depart["flightTypeCode"] != "F/H"]
                    .loc[raw_depart["flightTypeCode"] != "Q/B"]
                )
                dep_a = raw_depart.loc[raw_depart["atot"] >= start].loc[
                    raw_depart["atot"] < datetime_now
                ]
                arr_s = (
                    raw_arrival.loc[raw_arrival["sibt"] >= start]
                    .loc[raw_arrival["sibt"] < datetime_now]
                    .loc[raw_arrival["flightTypeCode"] != "F/H"]
                    .loc[raw_arrival["flightTypeCode"] != "Q/B"]
                )
                arr_a = raw_arrival.loc[raw_arrival["aldt"] >= start].loc[
                    raw_arrival["aldt"] < datetime_now
                ]

                extra_msg += (
                    "上一小时计划出港{}架次，实际出港{}架次，计划进港{}架次，实际进港{}架次。".format(
                        dep_s.__len__(),
                        dep_a.__len__(),
                        arr_s.__len__(),
                        arr_a.__len__(),
                    )
                    + "楼内设备设施与旅客秩序总体正常。"
                )

            if deice and len(raw_flight):
                if not extra_msg:
                    extra_msg += (
                        prefix.pop()
                        + end_time
                        + "出港计划执行{}架次，已执行{}架次，取消{}架次，".format(
                            raw_depart.loc[
                                (
                                    raw_depart["flightDate"].astype(str)
                                    == str(today)[:19]
                                )
                                & (raw_depart["sobt"] <= datetime_now)
                                & (raw_depart["flightTypeCode"] != "F/H")
                                & (raw_depart["flightTypeCode"] != "Q/B")
                            ].__len__(),
                            raw_flight.loc[
                                raw_flight["outAtot"] <= datetime_now
                            ].__len__(),
                            outCanceled,
                        )
                        + msg_rates
                        + msg_delayed
                    )
                prefix = prefix[0]
                deicing = raw_flight.loc[raw_flight["outDeicingType"] == "ICP"].copy()

                canceled = raw_flight.loc[
                    raw_flight["outDeicingType"] == "NOI"
                ].__len__()
                if deicing.__len__():
                    deicing["outCcct"] = deicing["outSobt"] - timedelta(minutes=5)

                    # 计划关舱 实际关舱 计划撤轮挡 实际撤轮挡
                    handle_time = ["outCcct", "outAcct", "outSobt", "outAobt"]

                    # 实际推出 开始滑行 进入除冰坪等待 除冰坪入位 开始除冰 结束除冰 离开除冰坪 实际起飞
                    deicing_time = [
                        "moniJob.tract_D.actBeginTime",
                        "outPushTime",
                        "enterWaitingAreaTime",
                        "rightIntoTime",
                        "outAczt",
                        "outAezt",
                        "awayApronTime",
                        "outAtot",
                    ]

                    handle_time_tuple = [
                        (deicing[i] - deicing[j]).mean().total_seconds()
                        for i, j in zip(handle_time[::2], handle_time[1::2])
                    ]

                    deicing_time_msg = ""
                    handle_time = ("平均{}关舱门", "平均{}推出")
                    for handle_ms, handle_time in zip(handle_time_tuple, handle_time):
                        handle_ms = self.min_sec_format(handle_ms, True)
                        if handle_ms:
                            deicing_time_msg += handle_time.format(handle_ms) + "，"

                    extra_msg += deicing_time_msg

                    deicing_time_msg = ""
                    deicing_time_tuple = [
                        (deicing[i] - deicing[j]).mean().total_seconds()
                        for i, j in zip(deicing_time[1:], deicing_time[:-1])
                    ]
                    deicing_time = (
                        "平均推出时间",
                        "滑至除冰坪用时",
                        "平均入位用时",
                        prefix + "前准备",
                        prefix + "作业",
                        prefix + "后开车及检查",
                        "滑出至起飞",
                    )
                    for deicing_ms, deicing_time in zip(
                        deicing_time_tuple, deicing_time
                    ):
                        deicing_ms = self.min_sec_format(deicing_ms)
                        if deicing_ms:
                            deicing_time_msg += deicing_time + deicing_ms + "，"

                    extra_msg += deicing_time_msg

                    deicing_time_msg = (
                        (deicing["outAtot"] - deicing["outAobt"]).mean().total_seconds()
                    )
                    if np.isnan(deicing_time_msg):
                        extra_msg = extra_msg[:-1] + "。"
                    else:
                        deicing_time_msg = f"推出至起飞全流程平均用时{self.min_sec_format(deicing_time_msg)}（"
                        for aircraft_size, deicing_ in deicing.groupby(
                            "aircraftSizeCategory"
                        ):
                            deicing_ms = (
                                (deicing_["outAtot"] - deicing_["outAobt"])
                                .mean()
                                .total_seconds()
                            )
                            if not np.isnan(deicing_ms):
                                deicing_ms = self.min_sec_format(deicing_ms)
                                deicing_time_msg += (
                                    f"{aircraft_size}类平均用时" + deicing_ms + "，"
                                )
                        deicing_time_msg = deicing_time_msg[:-1] + "）。"
                        extra_msg += deicing_time_msg

                    extra_msg += "目前共收到航空器{0}需求{1}架次，正在{0}{2}架次，已完成{0}{3}架次，取消{0}{4}架次，未出现航空器{0}排队等待情况。".format(
                        prefix,
                        deicing.__len__() + canceled,
                        deicing.loc[
                            deicing["outAezt"].isna()
                            | (deicing["outAezt"] > datetime_now)
                        ]
                        .loc[deicing["outAczt"] <= datetime_now]
                        .__len__(),
                        deicing.loc[deicing["outAezt"] <= datetime_now].__len__(),
                        canceled,
                    )
                else:
                    extra_msg += (
                        f"收到航空器{prefix}需求{canceled}架次，均已取消。"
                        if canceled
                        else f"未收到航空器{prefix}需求。"
                    )

            path = (
                path.get().format("综合效率席短信")
                if path
                else self.INFO_PATH.get() if self.INFO_PATH_.get() else None
            )
            output = {
                "航班正常性与延误详情": msg_finalized
                + self.wrap_iterstr(warnings, "（{}）"),
            }
            if "sync" not in self.RUNNING:
                self.save_textfile(msg_finalized + self.wrap_iterstr(warnings), path)
            self.update_log(f"航班正常性与延误详情生成成功：{msg_finalized}", "text")
            self.save_push("航班正常性与延误详情", msg_finalized)
            if extra_msg:
                output["综合效率席短信"] = extra_msg + self.wrap_iterstr(
                    warnings, "（{}）"
                )
                if "sync" not in self.RUNNING:
                    self.save_textfile(extra_msg + self.wrap_iterstr(warnings), path)
                self.update_log(f"综合效率席短信生成成功：{extra_msg}", "text")
                self.save_push("综合效率席短信", extra_msg)
            if warnings:
                self.update_log("警告：" + "；".join(warnings), "warn")
            return output
        finally:
            self.update_status(
                {"综合效率席短信": ""}, {"延误原因判定问题": "，".join(warnings)}
            )

    def get_search(self, name: str, code: str, insert=None, get_ga: bool = False):
        lc = len(code)
        payload = {"pageNum": 1, "pageSize": 100, "condition": {"name#like": name}}

        result = []
        session = self.get_session()
        if 2 <= lc <= 3 or lc == 0:
            payload["condition"]["iata#like"] = code if lc == 2 else ""
            payload["condition"]["icao#like"] = code if lc == 3 else ""
            response = session.post(
                self.URLS["航司查询"],
                data=dumps(payload),
                headers=self.HEADER,
                timeout=10,
            )
            for r in response.json().get("data").get("list"):
                iata = r.get("iata")
                grdAgent = self.AG.get(iata, "")
                grdAgent = (
                    self.get_ga(iata)
                    if get_ga and grdAgent == ""
                    else self.GA.get(grdAgent, "双击更新")
                )
                s = r.get("countryCode")
                result.append(
                    (
                        r.get("name"),
                        iata,
                        r.get("icao"),
                        s if s else "",
                        "双击更新" if "无" in grdAgent else grdAgent,
                    )
                )

        if 3 <= lc <= 4 or lc == 0:
            payload["condition"]["iata#like"] = code if lc == 3 else ""
            payload["condition"]["icao#like"] = code if lc == 4 else ""
            response = session.post(
                self.URLS["机场查询"],
                data=dumps(payload),
                headers=self.HEADER,
                timeout=10,
            )
            for r in response.json().get("data").get("list"):
                dir = r.get("depRunway")
                dir = self.RUNWAYDIR.get(dir, dir if dir else "")
                s = r.get("abbr")
                result.append(
                    (r.get("name"), r.get("iata"), r.get("icao"), s if s else "", dir)
                )

        if insert:
            for r in result:
                insert("", tk.END, values=r)
        return result

    def get_ga(self, iata: str):
        data = self.get_flight_data(
            datetime.today() - timedelta(days=1),
            datetime.today() + timedelta(days=3),
            0,
            airLine=iata,
            limit=2,
            sortOrder="descending",
            sortName="outSobt",
            departureMode="outS",
            moni=False,
        )["航班_航班"]
        grdAgent = data.get("grdAgent")[0] if len(data) else "N/D"
        if not isinstance(grdAgent, str):
            grdAgent = "N/A"
        self.AG[iata] = grdAgent
        return self.GA.get(grdAgent)

    def search(self, *args):
        box = tk.Toplevel(self, name="航司机场查询")
        box.attributes("-topmost", self.TOPMOST.get())
        box.title("航司机场查询")

        upper = ttk.Frame(box)
        lower = ttk.Frame(box)
        mid = ttk.Frame(box)
        upper.pack(side="top")
        mid.pack(side="top")
        lower.pack(side="bottom")
        note = tk.StringVar(
            box,
            "中文名支持模糊查询，代码仅可精确匹配，点击搜索或回车同时查找符合条件的航司或机场。"
            + "下表中，额外信息为机场进离港方向或航司地服代理，如需查找新入场航司地服或更新地服公司请双击对应数据行。",
        )
        search_time = tk.DoubleVar(box)

        def search(*args):
            now = datetime.now().timestamp()
            if now - search_time.get() > 1:
                search_time.set(now)
                name_, code_ = name.get().strip(), code.get().strip().upper()
                lc = len(code_)
                if 1 < lc < 5 or (lc == 0 and not name_ == ""):
                    if note.get():
                        mid.destroy()
                        box.update()
                        note.set("")
                    table.delete(*table.get_children())
                    threading.Thread(
                        target=self.get_search, args=(name_, code_, table.insert)
                    ).start()

        ttk.Label(upper, text="中文名").grid(row=0, column=0, padx=5, pady=5)
        name = ttk.Entry(upper, textvariable=tk.StringVar(box), width=30)
        name.grid(row=0, column=1, padx=5, pady=5, columnspan=3)
        name.bind("<Return>", search)
        name.focus_set()

        ttk.Label(upper, text="代码").grid(row=0, column=5, padx=5, pady=5)
        code = ttk.Entry(upper, textvariable=tk.StringVar(box), width=10)
        code.grid(row=0, column=6, padx=5, pady=5)
        code.bind("<Return>", search)

        ttk.Button(upper, text="搜索", command=search, width=10).grid(
            row=0, column=7, padx=5, pady=5
        )
        box.bind("<Escape>", lambda x: box.destroy())

        nl = ttk.Label(
            mid, textvariable=note, font=("微软雅黑", 9), justify="left", wraplength=520
        )
        nl.pack(padx=1, pady=5, fill=tk.BOTH)

        columns = ("中文名", "IATA", "ICAO", "城市/国家", "额外信息")
        widths = (200, 60, 60, 120, 80)

        sby = ttk.Scrollbar(lower)
        table = ttk.Treeview(
            lower,
            height=box.winfo_screenheight(),
            show="headings",
            columns=columns,
            yscrollcommand=sby.set,
            selectmode="browse",
        )
        for column, width in zip(columns, widths):
            table.column(
                column, width=width, minwidth=width, anchor="center", stretch=True
            )
            table.heading(column, text=column)
        sby.pack(side=tk.RIGHT, fill=tk.BOTH)
        sby.config(command=table.yview)

        width = sum(widths)
        table.pack(side=tk.LEFT, fill=tk.BOTH)

        def get_ga(event):
            r = table.identify_row(event.y)
            now = datetime.now().timestamp()
            if now - search_time.get() > 0.5:
                search_time.set(now)
                if r:
                    i, j = table.item(r)["values"][1:3]
                    if len(j) == 3:
                        try:
                            s = "更新失败"
                            table.set(r, "额外信息", "更新中...")
                            s = self.get_ga(i)
                        finally:
                            table.set(r, "额外信息", s)

        table.bind("<Double-1>", get_ga)
        box.geometry(
            f"538x300+{self.winfo_rootx() + self.winfo_width() // 5}+{self.winfo_rooty() + self.winfo_height() // 5}"
        )
        box.minsize(width=538, height=160)
        box.resizable(False, True)
        box.mainloop()

    def adj(self, *args):
        try:
            adj = self.nametowidget("自动化调时")
            adj.focus_set()
            return 0
        except Exception:
            ...

        adj = tk.Toplevel(self, name="自动化调时")
        adj.title("自动化调时")
        adj.geometry(
            f"+{self.winfo_rootx() + self.winfo_width() // 3}+{self.winfo_rooty() + self.winfo_height() // 8}"
        )

        cb_ls = tk.BooleanVar(adj, True)
        limit_start = tk.IntVar(adj, 6)
        limit_end = tk.IntVar(adj, 9)
        hour_limit = tk.IntVar(adj, 22)
        slot_limit = tk.IntVar(adj, 4)
        oversize_cb = tk.BooleanVar(adj, True)
        oversize_sep = tk.IntVar(adj, 10)
        hour_max = tk.IntVar(adj, 28)
        slot_max = tk.IntVar(adj, 6)
        to_acdm = tk.BooleanVar(adj, True)
        iteration = tk.DoubleVar(adj, 2)
        adj_limit = tk.IntVar(adj, 120)
        adj_file = tk.StringVar(adj)
        cb_file = tk.BooleanVar(adj)

        def update_adj_type():
            self.ADJ_TYPE = self.ask_type("配置调时航班性质", adj, self.ADJ_TYPE)
            atl.set(f"调时航班性质 / {len(self.ADJ_TYPE)}类")

        row_f = 0
        ttk.Label(adj, text="调时限制时间段内", font=("微软雅黑", 10, "bold")).grid(
            column=0, row=row_f, padx=5, pady=2, columnspan=3, sticky="w"
        )

        row_f += 1
        ttk.Checkbutton(adj, text="次日", **ONOFFS, variable=cb_ls).grid(
            column=0, row=row_f, padx=10, pady=1, columnspan=1, sticky="e"
        )
        ttk.Spinbox(adj, textvariable=limit_start, from_=0, to=23, width=2).grid(
            column=0, row=row_f, padx=5, pady=1, columnspan=3, sticky="e"
        )
        ttk.Label(adj, text="至").grid(
            column=3, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        ttk.Checkbutton(adj, text="次日", **ONOFFS, variable=cb_ls).grid(
            column=3, row=row_f, padx=5, pady=1, columnspan=2, sticky="e"
        )
        ttk.Spinbox(adj, textvariable=limit_end, from_=0, to=23, width=2).grid(
            column=5, row=row_f, padx=3, pady=1, columnspan=2, sticky="w"
        )
        ttk.Label(adj, text="时").grid(column=5, row=row_f, padx=5, pady=1, sticky="e")

        row_f += 1
        ttk.Label(adj, text="小时离港总量限制").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(adj, textvariable=hour_limit, from_=0, to=48, width=2).grid(
            column=3, row=row_f, padx=5, pady=1, sticky="w"
        )
        ttk.Label(adj, text="架次/方向/小时").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="时刻离港总量限制").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(adj, textvariable=slot_limit, from_=0, to=8, width=2).grid(
            column=3, row=row_f, padx=5, pady=1, sticky="w"
        )
        ttk.Label(adj, text="架次/方向/时刻").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        scb = ttk.Checkbutton(adj, text="大型机间隔", **ONOFFS, variable=oversize_cb)
        scb.grid(column=0, row=row_f, padx=5, pady=1, columnspan=3, sticky="e")
        ssb = ttk.Spinbox(
            adj, textvariable=oversize_sep, from_=5, to=60, width=2, increment=5
        )
        ssb.grid(column=3, row=row_f, padx=5, pady=1, sticky="w")
        scb.config(
            command=lambda: (
                (oversize_sep.set(10), ssb.config(state=tk.NORMAL, show=""))
                if oversize_cb.get()
                else (oversize_sep.set(0), ssb.config(state=tk.DISABLED, show=" "))
            )
        )
        ttk.Label(adj, text="分钟/方向/架次").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="调时限制时间段外", font=("微软雅黑", 10, "bold")).grid(
            column=0, row=row_f, padx=5, pady=2, columnspan=3, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="最大小时离港总量").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(adj, textvariable=hour_max, from_=20, to=48, width=2).grid(
            column=3, row=row_f, padx=5, pady=1, sticky="w"
        )
        ttk.Label(adj, text="架次/方向/小时").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="最大时刻离港总量").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(adj, textvariable=slot_max, from_=4, to=8, width=2).grid(
            column=3, row=row_f, padx=5, pady=1, sticky="w"
        )
        ttk.Label(adj, text="架次/方向/时刻").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="更多设置", font=("微软雅黑", 10, "bold")).grid(
            column=0, row=row_f, padx=5, pady=2, columnspan=3, sticky="w"
        )

        row_f += 1
        ttk.Checkbutton(
            adj,
            text="在A-CDM系统创建调时后方案",
            variable=to_acdm,
            command=lambda: (
                None
                if to_acdm.get()
                else (
                    cb_file.set(True),
                    entry_file.config(state=tk.NORMAL),
                    adj_file.set(
                        self.FILENAME.get("调时结果").format(
                            *datetime.now().timetuple()
                        )
                    ),
                )
            ),
            **ONOFFS,
        ).grid(column=0, row=row_f, columnspan=6, padx=5, pady=1)

        row_f += 1
        ttk.Label(adj, text="迭代次数倍数").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(
            adj, textvariable=iteration, from_=1, to=5, width=3, increment=0.1
        ).grid(column=3, row=row_f, padx=5, pady=1, columnspan=2, sticky="w")
        ttk.Label(adj, text="倍").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Label(adj, text="最大调时时长").grid(
            column=0, row=row_f, padx=5, pady=1, sticky="e", columnspan=3
        )
        ttk.Spinbox(
            adj, textvariable=adj_limit, from_=60, to=180, width=3, increment=5
        ).grid(column=3, row=row_f, padx=5, pady=1, columnspan=2, sticky="w")
        ttk.Label(adj, text="分钟").grid(
            column=4, row=row_f, padx=5, pady=1, columnspan=2, sticky="w"
        )

        row_f += 1
        ttk.Checkbutton(
            adj,
            text="生成表格文件",
            variable=cb_file,
            command=lambda: (
                (
                    entry_file.config(state=tk.NORMAL),
                    adj_file.set(
                        self.FILENAME.get("调时结果").format(
                            *datetime.now().timetuple()
                        )
                    ),
                )
                if cb_file.get()
                else (
                    adj_file.set(""),
                    entry_file.config(state=tk.DISABLED),
                    to_acdm.set(True),
                )
            ),
            **ONOFFS,
        ).grid(sticky="e", row=row_f, column=0, padx=5, pady=1, columnspan=3)
        entry_file = ttk.Entry(adj, textvariable=adj_file, width=16, state=tk.DISABLED)
        entry_file.grid(sticky="w", row=row_f, column=3, padx=5, pady=1, columnspan=4)

        row_f += 1
        pal = tk.StringVar(adj, f"预调时航班 / {len(self.PRE_ADJ)}班")
        ttk.Label(adj, textvariable=pal).grid(
            sticky="e", row=row_f, column=0, padx=5, pady=1, columnspan=2
        )
        ttk.Button(
            adj,
            text="配置预调时航班...",
            width=18,
            style="Settings.TButton",
            command=lambda: (
                self.attribute_editor(
                    "PRE_ADJ",
                    "配置预调时航班",
                    "航班号为IATA格式（如CZ3000）；目标时间为'1020'的时分格式，留空优先调整",
                    master=adj,
                    headers=("航班号", "目标时间"),
                    width=(120, 120),
                    anchor=("center", "center"),
                    key=lambda x: len(x) > 2 and x.isalnum() and x[2].isnumeric(),
                    key_convert=lambda x: x.upper(),
                    value=lambda x: (
                        (
                            int(x[:-2]) < 24
                            and int(x[-2]) < 6
                            and (x[-1] == "5" or x[-1] == "0")
                        )
                        if x.isnumeric() and 3 <= len(x) <= 4
                        else True if x == "" else False
                    ),
                ),
                pal.set(f"预调时航班 / {len(self.PRE_ADJ)}班"),
            ),
        ).grid(sticky="w", row=row_f, column=3, padx=5, pady=1, columnspan=3)

        row_f += 1
        atl = tk.StringVar(adj, f"调时航班性质 / {len(self.ADJ_TYPE)}类")
        ttk.Label(adj, textvariable=atl).grid(
            sticky="e", row=row_f, column=0, padx=5, pady=1, columnspan=2
        )
        ttk.Button(
            adj,
            text="配置调时航班性质...",
            width=18,
            style="Settings.TButton",
            command=update_adj_type,
        ).grid(sticky="w", row=row_f, column=3, padx=5, pady=1, columnspan=3)

        row_f += 1
        lb = ttk.Button(
            adj, text="开始自动化调时", name="开始自动化调时", style="Main.TButton"
        )
        lb.grid(sticky="ew", row=row_f, column=0, padx=10, pady=10, columnspan=7)

        iter_ = tk.DoubleVar(self)

        def get_adj(
            datetime_start: datetime,
            datetime_end: datetime,
            hour_limit: int,
            slot_limit: int,
            oversize_sep: int,
            hour_max: int,
            slot_max: int,
            iteration: float,
            adj_limit: int,
            to_acdm: bool,
            file: str,
        ):
            try:
                iter_.set(0)
                pb = ttk.Progressbar(
                    adj,
                    maximum=100,
                    mode="determinate",
                    name="自动化调时进度",
                    orient="horizontal",
                    variable=iter_,
                )
                pb.grid(
                    sticky="ew", row=row_f, column=0, padx=10, pady=10, columnspan=7
                )

                if file.strip()[-5:] != r".xlsx" and file.strip() != "":
                    file += r".xlsx"
                if not adj.winfo_exists():
                    raise InterruptedError

                start = -1
                adj_limit_ = divmod(adj_limit, 60)
                adj_limit_ = 2 * (adj_limit_[0] + 1 if adj_limit_[1] else adj_limit_[0])
                datetime_i = (datetime_end - datetime_start).seconds / 300
                start_time = datetime_start + timedelta(hours=start)
                end_time = datetime_end + timedelta(hours=adj_limit_)
                limit_fi = np.arange(
                    -start * 12 + 6, -start * 12 + datetime_i - 6
                ).astype(int)
                oversize_sepi = oversize_sep // 5

                payload = self.FLIGHT_PAYLOAD.copy()
                raw_flight = self.get_flights(
                    limit=self.DATA_LIMIT.get(),
                    departureMode="outS",
                    startTime=str(start_time)[:16],
                    endTime=str(end_time)[:16],
                    **payload,
                )
                adj_limit = adj_limit // 5

                pre_adj = self.PRE_ADJ.copy()
                for k in list(pre_adj.keys()):
                    if pre_adj[k]:
                        pre_adj[k] = datetime(
                            *datetime_start.timetuple()[:3],
                            int(pre_adj[k][:-2]),
                            int(pre_adj[k][-2:]),
                        )

                if not adj.winfo_exists():
                    raise InterruptedError
                iter_.set(5)

                def next_slot(h: int, m: int, slot: int = 5):
                    m += slot
                    if m >= 60:
                        m -= 60
                        h += 1
                        if h >= 24:
                            h -= 24
                    elif m < 0:
                        m += 60
                        h -= 1
                        if h < 0:
                            h += 24
                    return h, m

                def near_hour(data):
                    data_len = data.__len__()
                    return [
                        np.sum(data[max(0, i - 6) : min(data_len - 1, i + 5)])
                        for i in range(data_len)
                    ]

                raw_flight = concat(
                    [
                        raw_flight.loc[raw_flight["outFlightTypeCode"] == i]
                        for i in self.ADJ_TYPE
                    ],
                    ignore_index=True,
                )
                raw_flight["outDir"] = (
                    raw_flight["outToPath"]
                    .astype(str)
                    .map(lambda x: x[:5], IGNORE)
                    .map(self.DIR, IGNORE)
                )
                raw_flight["outDir"] = raw_flight["outDir"].fillna(
                    raw_flight["outRoute"].map(lambda x: self.DIR.get(x[:3]), IGNORE)
                )
                raw_flight["outSobt"] = (
                    raw_flight["outSobt"]
                    .astype(str)
                    .map(datetime.fromisoformat, IGNORE)
                )
                raw_flight["outSobtStamp"] = (
                    raw_flight["outSobt"]
                    .astype(str)
                    .map(lambda x: datetime.fromisoformat(x).timestamp(), IGNORE)
                )
                raw_flight["outSobtHour"] = raw_flight["outSobt"].map(
                    lambda x: x.hour, IGNORE
                )
                raw_flight["outSobtMin"] = raw_flight["outSobt"].map(
                    lambda x: x.minute, IGNORE
                )
                raw_flight["aircraftSizeCategory"] = raw_flight[
                    "aircraftSizeCategory"
                ].map(self.AIRCRAFTSIZE, IGNORE)
                raw_flight["outDefaultRunwayCode"] = raw_flight[
                    "outDefaultRunwayCode"
                ].map(self.RUNWAYDIR, IGNORE)
                raw_flight["outRouteCn"] = (
                    raw_flight[["outRoute", "outRouteCn"]]
                    .fillna("")
                    .apply(self.airport_name, axis=1)
                )

                raw_flight_ = raw_flight.loc[raw_flight["outDir"].isna()].copy()
                for i in raw_flight_.index:
                    if messagebox.askyesno(
                        TITLE,
                        "{}: {} ({}) 无法确定离港跑道方向，是否为西向？\n".format(
                            *raw_flight_.loc[
                                i, ["outFlightNo", "outRouteCn", "outRoute"]
                            ].values
                        ),
                        parent=adj,
                    ):
                        self.DIR[raw_flight_.loc[i, "outRoute"][:3]] = "西"
                        raw_flight.loc[i, "outDir"] = "西"
                    else:
                        self.DIR[raw_flight_.loc[i, "outRoute"][:3]] = "东"
                        raw_flight.loc[i, "outDir"] = "东"

                def conflict(data: DataFrame):
                    if len(data):
                        if messagebox.askyesno(
                            TITLE,
                            "以下航班离港跑道方向与系统冲突，程序替换结果如下，是否另存为表格？"
                            + str().join(
                                "\n{0}: {2}向离港至{3} ({4})".format(*x)
                                for x in data.values
                            ),
                            parent=adj,
                        ):

                            def save_conflict(file: str, data: DataFrame, **kwargs):
                                with ExcelWriter(file, mode="w") as wb:
                                    data.to_excel(wb, sheet_name="方向冲突", **kwargs)
                                    ws = wb.sheets["方向冲突"]
                                    for row in range(1, ws.max_row):
                                        ws.cell(row + 1, 2).number_format = "HH:MM"
                                    ws.column_dimensions["C"].width = 4.5
                                    ws.column_dimensions["D"].width = 20

                            title = "导出方向冲突明细表格"
                            filename = self.FILENAME.get("方向冲突").format(
                                *datetime.now().timetuple()
                            )
                            s = self.save_excel(
                                adj,
                                title,
                                (
                                    f"{self.FUNC_PATH.get()}/{filename}"
                                    if self.FUNC_PATH_.get()
                                    else filedialog.asksaveasfilename(
                                        filetypes=(("Xlsx表格文件", "*.xlsx"),),
                                        confirmoverwrite=True,
                                        parent=adj,
                                        title=title,
                                        initialdir=self.FUNC_PATH.get(),
                                        initialfile=filename,
                                    )
                                ),
                                save_conflict,
                                data=data,
                                index=False,
                                header=["航班号", "STD", "方向", "航线中文", "航线"],
                            )
                            if s:
                                self.update_log(f"方向冲突表格保存至{s}", "file")

                threading.Thread(
                    target=conflict,
                    args=(
                        raw_flight.loc[
                            raw_flight["outDefaultRunwayCode"] != raw_flight["outDir"],
                            [
                                "outFlightNo",
                                "outSobt",
                                "outDir",
                                "outRouteCn",
                                "outRoute",
                            ],
                        ],
                    ),
                ).start()
                full_slots = "以下时刻航班密集，避免调入："
                result = DataFrame(
                    columns=["航班日期", "航班号", "方向", "原时刻", "新时刻"]
                )
                iter_.set(10)

                for direction, adj_flight in raw_flight.groupby("outDir"):
                    if not adj.winfo_exists():
                        raise InterruptedError

                    origin_count = pivot_table(
                        adj_flight,
                        "outFlightNo",
                        "outSobtMin",
                        "outSobtHour",
                        lambda x: len(x),
                    ).fillna(0)
                    oversize_count = pivot_table(
                        adj_flight,
                        "aircraftSizeCategory",
                        "outSobtMin",
                        "outSobtHour",
                        "sum",
                    ).fillna(0)
                    adj_flight.set_index(["outSobtHour", "outSobtMin"], inplace=True)
                    adj_flight.sort_index(inplace=True)

                    h, m = start_time.hour, start_time.minute
                    flight_flow = DataFrame(
                        columns=[
                            "时",
                            "分",
                            "小时上限",
                            "时刻上限",
                            "大型机上限",
                            "航班量",
                            "大型机",
                        ]
                    )
                    for i in range(int((end_time - start_time).seconds / 300)):
                        hour_max_, slot_max_, os_max = (
                            (hour_limit, slot_limit, 1)
                            if -start * 12 <= i < datetime_i - start * 12
                            else (hour_max, slot_max, slot_max)
                        )
                        try:
                            oc = (
                                int(origin_count.at[m, h]),
                                int(oversize_count.at[m, h]),
                            )
                        except KeyError:
                            oc = 0, 0
                        flight_flow.loc[i] = [h, m, hour_max_, slot_max_, os_max, *oc]
                        h, m = next_slot(h, m)

                    del origin_count
                    del oversize_count

                    flight_flow_ = flight_flow.copy()
                    flow_count = flight_flow.__len__()
                    zeros = np.zeros(flow_count)
                    flight_flow["时刻超量"] = (
                        flight_flow["航班量"] - flight_flow["时刻上限"]
                    )
                    flight_flow["小时超量"] = (
                        near_hour(flight_flow["航班量"].values)
                        - flight_flow["小时上限"]
                    )

                    overlay = flight_flow.loc[
                        flight_flow["时刻超量"] >= 0, ["时刻超量", "时", "分"]
                    ].values
                    if len(overlay):
                        full_slots += f"{direction}向"
                        for i in overlay:
                            full_slots += " {1:02}{2:02}".format(*i)
                        full_slots += "；"

                    overlay, index = (
                        DataFrame(columns=["航班流", "大型机", "调时幅度"]),
                        0,
                    )
                    overlay_os = overlay.copy()

                    # 优先调整航班
                    adj_override = DataFrame(columns=["航班号", "航班流", "大型机"])
                    for k, v in pre_adj.items():
                        if k in adj_flight["outFlightNo"].values:
                            j = adj_flight.loc[adj_flight["outFlightNo"] == k].index[0]
                            si = adj_flight.loc[j, "aircraftSizeCategory"].iloc[0]
                            i = (
                                flight_flow.loc[flight_flow["时"] == j[0]]
                                .loc[flight_flow["分"] == j[1]]
                                .index[0]
                            )
                            if v:
                                i_ = (
                                    flight_flow.loc[flight_flow["时"] == v.hour]
                                    .loc[flight_flow["分"] == v.minute]
                                    .index
                                )
                                if len(i_):
                                    i_ = i_[0]
                                    flight_flow.loc[i, "航班量"] -= 1
                                    flight_flow_.loc[i, "航班量"] -= 1
                                    flight_flow.loc[i_, "航班量"] += 1
                                    flight_flow_.loc[i_, "航班量"] += 1
                                    flight_flow.loc[i, "大型机"] -= si
                                    flight_flow_.loc[i, "大型机"] -= si
                                    flight_flow.loc[i_, "大型机"] += si
                                    flight_flow_.loc[i_, "大型机"] += si
                                    overlay_os.loc[overlay_os.__len__()] = [
                                        i,
                                        si,
                                        i_ - i,
                                    ]
                                    result.loc[result.__len__()] = [
                                        v.date().isoformat(),
                                        k,
                                        direction,
                                        "{:02}:{:02}".format(j[0], j[1]),
                                        "{:02}:{:02}".format(v.hour, v.minute),
                                    ]
                            else:
                                adj_override.loc[adj_override.__len__()] = [k, i, si]
                    overlay_os["强制"] = 1

                    # 大型机若要求有间隔，则优先直接TOBT提前5分钟，若不满足间隔或时刻超量则加入后续调整队列
                    if oversize_sepi > 0:
                        for i in flight_flow_.index[oversize_sepi : -oversize_sepi - 1]:
                            adj_limit_ = min(adj_limit, flow_count - i)
                            while flight_flow_.loc[i, "大型机"] - overlay.loc[
                                overlay["航班流"] + overlay["调时幅度"] == i
                            ].__len__() - overlay_os.loc[overlay_os["大型机"] == 1].loc[
                                overlay_os["航班流"] + overlay_os["调时幅度"] == i
                            ].__len__() > 1 or (
                                flight_flow_.loc[i, "大型机"]
                                - overlay.loc[
                                    overlay["航班流"] + overlay["调时幅度"] == i
                                ].__len__()
                                - overlay_os.loc[overlay_os["大型机"] == 1]
                                .loc[overlay_os["航班流"] + overlay_os["调时幅度"] == i]
                                .__len__()
                                == 1
                                and (
                                    (
                                        flight_flow_.loc[
                                            i - oversize_sepi + 1 : i - 1, "大型机"
                                        ]
                                        >= flight_flow.loc[
                                            i - oversize_sepi + 1 : i - 1, "大型机上限"
                                        ]
                                    ).any()
                                    or (
                                        flight_flow_.loc[
                                            i + 1 : i + oversize_sepi - 1, "大型机"
                                        ]
                                        >= flight_flow.loc[
                                            i + 1 : i + oversize_sepi - 1, "大型机上限"
                                        ]
                                    ).any()
                                )
                            ):
                                flight_flow_.loc[i, "大型机"] -= 1
                                flight_flow_.loc[i, "航班量"] -= 1
                                if (
                                    flight_flow_.loc[i - 1, "航班量"]
                                    < flight_flow.loc[i - 1, "时刻上限"]
                                    and (
                                        flight_flow_.loc[
                                            i - oversize_sepi : i - 1, "大型机"
                                        ]
                                        < flight_flow.loc[
                                            i - oversize_sepi : i - 1, "大型机上限"
                                        ]
                                    ).all()
                                ):
                                    flight_flow.loc[i, "航班量"] -= 1
                                    flight_flow.loc[i, "大型机"] -= 1
                                    flight_flow.loc[i - 1, "航班量"] += 1
                                    flight_flow.loc[i - 1, "大型机"] += 1
                                    flight_flow_.loc[i - 1, "航班量"] += 1
                                    flight_flow_.loc[i - 1, "大型机"] += 1
                                    overlay_os.loc[overlay_os.__len__()] = [i, 1, -1, 0]
                                    continue

                                for j in range(7, adj_limit_):
                                    if (
                                        flight_flow_.loc[i + j, "航班量"]
                                        < flight_flow.loc[i + j, "时刻上限"]
                                        and (
                                            flight_flow_.loc[
                                                i
                                                + j
                                                - oversize_sepi
                                                + 1 : i
                                                + j
                                                + oversize_sepi
                                                - 1,
                                                "大型机",
                                            ]
                                            < flight_flow.loc[
                                                i
                                                + j
                                                - oversize_sepi
                                                + 1 : i
                                                + j
                                                + oversize_sepi
                                                - 1,
                                                "大型机上限",
                                            ]
                                        ).all()
                                    ):
                                        flight_flow_.loc[i + j, "航班量"] += 1
                                        flight_flow_.loc[i + j, "大型机"] += 1
                                        overlay.loc[index] = [i, 1, j]
                                        index += 1
                                        break
                                else:
                                    flight_flow_.loc[i, "大型机"] += 1
                                    flight_flow_.loc[i, "航班量"] += 1

                    # 调整队列转换时刻超量，排除已添加大型机
                    for i in flight_flow_.index:
                        adj_limit_ = min(adj_limit, flow_count - i)
                        while (
                            flight_flow_.loc[i, "航班量"]
                            - flight_flow.loc[i, "时刻上限"]
                            > 0
                            and adj_limit_ > 7
                        ):
                            for j in range(7, adj_limit_):
                                if (
                                    flight_flow_.loc[i + j, "航班量"]
                                    < flight_flow.loc[i + j, "时刻上限"]
                                ):
                                    flight_flow_.loc[i, "航班量"] -= 1
                                    flight_flow_.loc[i + j, "航班量"] += 1
                                    overlay.loc[index] = [i, 0, j]
                                    index += 1
                                    break

                    # 若时刻超量小于小时超量，增加超量小时内时刻超量最多航班作为时刻超量
                    s = (
                        lambda x, i: x.loc[x["航班流"] < i - adj_limit / 2].__len__()
                        - x.loc[x["航班流"] >= i - 6].loc[x["航班流"] < i + 6].__len__()
                    )
                    for i in limit_fi[::6]:
                        ss = flight_flow.loc[i, "小时超量"] + s(overlay, i)
                        if ss <= 0:
                            continue

                        k = (
                            adj_override.loc[adj_override["航班流"] < i + 6]
                            .loc[adj_override["航班流"] >= i - 6]
                            .loc[adj_override["大型机"] == 0]["航班流"]
                            .to_list()
                        )
                        k.extend(
                            (
                                flight_flow.loc[i - 6 : i + 5, "航班量"]
                                - flight_flow.loc[i - 6 : i + 5, "大型机"]
                            )
                            .sort_values(ascending=False)
                            .index.to_list()
                        )
                        for i in k:
                            if ss <= 0:
                                break
                            if (
                                flight_flow_.loc[i, "航班量"]
                                <= flight_flow_.loc[i, "大型机"]
                                or flight_flow_.loc[i, "航班量"]
                                <= overlay.loc[
                                    (overlay["航班流"] + overlay["调时幅度"]) == i
                                ].__len__()
                                + overlay_os.loc[
                                    (overlay_os["航班流"] + overlay_os["调时幅度"]) == i
                                ].__len__()
                            ):
                                continue
                            adj_limit_ = min(adj_limit, flow_count - i)
                            for j in range(7, adj_limit_):
                                if (
                                    flight_flow_.loc[i + j, "航班量"]
                                    < flight_flow.loc[i + j, "时刻上限"]
                                ):
                                    flight_flow_.loc[i, "航班量"] -= 1
                                    flight_flow_.loc[i + j, "航班量"] += 1
                                    overlay.loc[index] = [i, 0, j]
                                    ss -= 1
                                    index += 1
                                    break
                    del adj_override

                    # 迭代控制时刻超量，得到满足时刻限制的最优结果，最小化小时超量和调时幅度
                    def hour_oc():
                        return Series(
                            np.max(
                                [
                                    zeros,
                                    near_hour(flight_flow_["航班量"])
                                    - flight_flow["小时上限"],
                                ],
                                axis=0,
                            )
                        )

                    s = hour_oc()
                    ol, ff, ss, ss_ = (
                        overlay.copy(),
                        flight_flow_.copy(),
                        0.5
                        * (
                            s.loc[-start * 12 : datetime_i - start * 12 - 1].sum()
                            + s.sum()
                        )
                        + 12 * s.loc[limit_fi[::6]].sum() ** 2,
                        0,
                    )
                    max_iteration = int(iteration * len(overlay) ** 2)

                    for i in range(1, max_iteration + 1):
                        if not adj.winfo_exists():
                            raise InterruptedError
                        if hour_oc().loc[limit_fi[::6]].max == 0:
                            iter_.set(50 if iter_.get() <= 50 else 90)
                            break
                        iter_.set(
                            40 * i / max_iteration + (10 if iter_.get() < 50 else 50)
                        )

                        # 查找索引顺序按：调时幅度越大、航班时刻越早，越优先寻找最优解
                        i, index = overlay.index.to_list(), []
                        for _ in overlay.index:
                            p = overlay.loc[i, "调时幅度"] - (
                                overlay.loc[i, "航班流"] - flow_count
                            )
                            index.append(np.random.choice(i, p=p / p.sum()))
                            i.remove(index[-1])
                        for i in index:
                            fi, si, oi = overlay.loc[i].values
                            flight_flow_.loc[fi + oi, "大型机"] -= si
                            flight_flow_.loc[fi + oi, "航班量"] -= 1

                        # 按查找索引开始寻找最优解
                        for i in index:
                            fi, si, oi = overlay.loc[i].values
                            flight_flow_["时刻超量"] = (
                                flight_flow_["航班量"] - flight_flow["时刻上限"]
                            )
                            target = flight_flow_.loc[
                                fi + 7 : min(fi + adj_limit, flow_count) - 1
                            ]
                            target = target.loc[target["时刻超量"] < 0].index.tolist()
                            if si:
                                for j in target:
                                    if (
                                        flight_flow.loc[
                                            max(0, j - oversize_sepi + 1) : min(
                                                j + oversize_sepi, flow_count
                                            )
                                            - 1,
                                            "大型机上限",
                                        ]
                                        >= flight_flow_.loc[
                                            max(0, j - oversize_sepi + 1) : min(
                                                j + oversize_sepi, flow_count
                                            )
                                            - 1,
                                            "大型机",
                                        ]
                                    ).any():
                                        target.remove(j)
                            if target.__len__():
                                s = hour_oc().loc[target]
                                s = -s + s.max()
                                j = np.random.choice(
                                    target, p=s.values / s.sum() if s.sum() else None
                                )
                                flight_flow_.loc[j, "大型机"] += si
                                flight_flow_.loc[j, "航班量"] += 1
                                overlay.loc[i, "调时幅度"] = j - fi
                            else:
                                flight_flow_.loc[fi + oi, "大型机"] += si
                                flight_flow_.loc[fi + oi, "航班量"] += 1

                        s = hour_oc()
                        ss_ = (
                            0.5
                            * (
                                s.loc[-start * 12 : datetime_i - start * 12 - 1].sum()
                                + s.sum()
                            )
                            + 12 * s.loc[limit_fi[::6]].sum() ** 2
                        )
                        if ss <= ss_:
                            overlay, flight_flow_ = ol, ff
                        else:
                            ol, ff = overlay.copy(), flight_flow_.copy()
                            ss = ss_

                    flight_flow_["时刻超量"] = (
                        flight_flow_["航班量"] - flight_flow["时刻上限"]
                    )
                    flight_flow_["小时超量"] = (
                        near_hour(flight_flow_["航班量"].values)
                        - flight_flow["小时上限"]
                    )
                    overlay = concat(
                        [
                            overlay_os.loc[overlay_os["强制"] == 0].drop(
                                "强制", axis=1
                            ),
                            overlay,
                        ],
                        ignore_index=True,
                    ).sort_values("航班流", ignore_index=True)

                    FlightDate = datetime_start.date().isoformat()
                    for fi, si, oi in overlay.values:
                        h, m = flight_flow.loc[fi, ["时", "分"]].values
                        h_, m_ = flight_flow.loc[fi + oi, ["时", "分"]].values
                        alts = DataFrame(columns=["FlightNo", "Count"])

                        for FlightNo, CountryCode, airlineIata, ss in adj_flight.loc[
                            (h, m),
                            [
                                "outFlightNo",
                                "outCountryCode",
                                "airlineIata",
                                "aircraftSizeCategory",
                            ],
                        ].values:
                            if (
                                pre_adj.get(FlightNo, None) == ""
                                and ss == si
                                and FlightNo not in result["航班号"].values
                            ):
                                result.loc[result.__len__()] = [
                                    FlightDate,
                                    FlightNo,
                                    direction,
                                    "{:02}:{:02}".format(h, m),
                                    "{:02}:{:02}".format(h_, m_),
                                ]
                                break
                            elif CountryCode.strip() == "D" and ss == si:
                                h__, m__ = h_, m_
                                while (h__, m__) not in adj_flight.index:
                                    h__, m__ = next_slot(h__, m__, -5)
                                ff = adj_flight.loc[(h, m):(h__, m__)]
                                alts.loc[alts.__len__()] = [
                                    FlightNo,
                                    ff.loc[ff["airlineIata"] == airlineIata].__len__(),
                                ]
                            else:
                                continue
                        else:
                            while alts.__len__():
                                i = np.random.choice(
                                    alts.index,
                                    p=(
                                        None
                                        if np.isnan(alts["Count"].sum())
                                        or alts["Count"].sum() == 0
                                        else alts["Count"].values / alts["Count"].sum()
                                    ),
                                )
                                if (
                                    alts.loc[i, "FlightNo"]
                                    not in result["航班号"].values
                                ):
                                    result.loc[result.__len__()] = [
                                        FlightDate,
                                        alts.loc[i, "FlightNo"],
                                        direction,
                                        "{:02}:{:02}".format(h, m),
                                        "{:02}:{:02}".format(h_, m_),
                                    ]
                                    break
                                else:
                                    alts.drop(i, inplace=True)
                        del alts

                if not adj.winfo_exists():
                    raise InterruptedError
                if "向" in full_slots:
                    full_slots = full_slots[:-1]
                    self.update_log(full_slots, "text")
                    self.save_textfile(
                        full_slots,
                        self.FUNC_PATH.get() if self.FUNC_PATH_.get() else None,
                    )

                def save(file: str, data: DataFrame, **kwargs):
                    with ExcelWriter(file, mode="w") as wb:
                        data.to_excel(wb, sheet_name="调时结果", **kwargs)
                        ws = wb.sheets["调时结果"]
                        for row in range(1, ws.max_row):
                            ws.cell(row + 1, 4).number_format = "HH:MM"
                            ws.cell(row + 1, 5).number_format = "HH:MM"
                        ws.column_dimensions["A"].width = 11
                        ws.column_dimensions["C"].width = 4.5

                if file:
                    file = self.save_excel(
                        adj,
                        "导出调时结果表格",
                        (
                            f"{self.FUNC_PATH.get()}/{file}"
                            if self.FUNC_PATH_.get()
                            else filedialog.asksaveasfilename(
                                confirmoverwrite=True,
                                parent=adj,
                                filetypes=(("Xlsx表格文件", "*.xlsx"),),
                                title="导出调时结果表格",
                                initialdir=self.FUNC_PATH.get(),
                                initialfile=file,
                            )
                        ),
                        save,
                        data=result,
                        index=False,
                    )
                    if file:
                        if file[0] == ".":
                            file = "程序所在文件夹" + file[1:]
                        self.update_log(f"调时结果表格保存至{file}", "file")
                    else:
                        self.update_log("调时结果表格保存被取消")
                if not to_acdm:
                    iter_.set(100)
                    return 0

                if not adj.winfo_exists():
                    raise InterruptedError
                session = self.get_session()

                payload = {
                    "id": None,
                    "adjFlights": 36,
                    "limitFlightsEast": None,
                    "limitFlightsWest": None,
                    "limitFlights": None,
                    "cancelFlights": None,
                    "status": 0,
                    "chgCount": 0,
                    "updatePer": self.ask_user()[0],
                    "updateTime": None,
                    "remark": None,
                    "adjAsPercent": None,
                    "roundMode": 0,
                    "topAirlines": 5,
                    "westLimit": 1,
                    "eastLimit": 1,
                    "limitReason": "常态化调时",
                    "adjName": f"程序生成调时计划_{datetime_start.strftime(r'%Y%m%d')}",
                    "cfgHourCapa": hour_limit,
                    "cfgSlotCapa": slot_limit,
                    "limitStartTm": start_time.isoformat(" ", "seconds"),
                    "limitEndTm": end_time.isoformat(" ", "seconds"),
                    "limitFlightType": str().join(i + "," for i in self.ADJ_TYPE)[:-1],
                    "westSlotCapa": slot_limit,
                    "eastSlotCapa": slot_limit,
                    "westHourCapa": hour_limit,
                    "eastHourCapa": hour_limit,
                }

                # 调时数据
                response = session.post(
                    self.URLS["新建调时"],
                    data=dumps(payload),
                    headers=self.HEADER,
                    timeout=30,
                )

                response = response.json().get("data").get("adjFlightVersionList")[0]
                id, version = response.get("adjMainId"), response.get("adjVersionId")
                payload = {
                    "adjMainId": id,
                    "adjType": 1,
                    "adjVersionId": version,
                    "chgCount": 0,
                }

                response = session.get(
                    f"http://acdm.bdia.com.cn/acdm/adjFlightVerDetailV/info/{id}/{version}",
                    headers=self.HEADER,
                    timeout=10,
                )
                flights = DataFrame.from_records(response.json().get("data"))

                i = ss = result.__len__()
                ss_ = 0
                s = ""
                for FlightNo, j, k, FlightTime in result[
                    ["航班号", "方向", "原时刻", "新时刻"]
                ].values:
                    iter_.set(10 * ss_ / ss + 90)
                    pb.update()
                    flights.loc[flights["flightNo"] == FlightNo, ["id", "adjId"]].values
                    payload["adjSchTm"] = f"{FlightDate} {FlightTime}:00"
                    payload["flightId"], payload["id"] = flights.loc[
                        flights["flightNo"] == FlightNo, ["id", "adjId"]
                    ].values[0]
                    payload["flightId"], payload["id"] = (
                        int(payload["flightId"]),
                        int(payload["id"]),
                    )
                    response = session.post(
                        self.URLS["调时更新"],
                        data=dumps(payload),
                        headers=self.HEADER,
                        timeout=10,
                    )
                    ss_ += 1
                    if response.json().get("code") != 1:
                        i -= 1
                        s += f"\n{FlightNo} ({j}向): 由{k}调整至{FlightTime} 调整失败"
                if not adj.winfo_exists():
                    raise InterruptedError

                self.update_log(f"创建调时方案{i}/{ss}班成功：{s}", "text")
            except InterruptedError:
                ...
            finally:
                if adj.winfo_exists():
                    pb.destroy()
                    lb.grid(
                        sticky="ew", row=row_f, column=0, padx=10, pady=10, columnspan=7
                    )

        def launch():
            today = datetime.today()
            day = (today + timedelta(1) if cb_ls.get() else today).timetuple()[:3]
            limit_start.set(inside(0, limit_start.get(), 23))
            limit_end.set(inside(0, limit_end.get(), 23))
            hour_limit.set(inside(0, hour_limit.get(), 48))
            slot_limit.set(inside(0, slot_limit.get(), 8))
            oversize_sep.set(inside(0, oversize_sep.get(), 60))
            hour_max.set(inside(20, hour_max.get(), 48))
            slot_max.set(inside(4, slot_max.get(), 8))
            iteration.set(inside(1, iteration.get(), 5))
            adj_limit.set(inside(60, adj_limit.get(), 180))
            start = datetime(*day, limit_start.get())
            end = datetime(*day, limit_end.get())
            if start >= end:
                messagebox.showinfo(TITLE, "时间范围错误", parent=adj)
                lb.grid(
                    sticky="ew", row=row_f, column=0, padx=10, pady=10, columnspan=7
                )
            else:
                thread = threading.Thread(
                    target=get_adj,
                    args=(
                        start,
                        end,
                        hour_limit.get(),
                        slot_limit.get(),
                        oversize_sep.get(),
                        hour_max.get(),
                        slot_max.get(),
                        iteration.get(),
                        adj_limit.get(),
                        to_acdm.get(),
                        adj_file.get().strip() if adj_file.get().strip() else "",
                    ),
                )
                thread.start()

        lb.config(command=lambda: [lb.grid_forget(), launch()])
        adj.resizable(False, False)
        adj.bind("<Escape>", lambda x: adj.destroy())
        adj.focus_set()
        adj.mainloop()

    def history(self):
        history = tk.Toplevel(self, name="history")
        history.title("CTOT和COBT历史查询")
        history.attributes("-topmost", self.TOPMOST.get())

        upper, mid, lower = [ttk.Frame(history) for _ in range(3)]
        search_time = tk.DoubleVar(history)

        def save(file: str, data: DataFrame):
            with ExcelWriter(file, mode="w", engine="openpyxl") as wb:
                data.to_excel(wb, sheet_name="历史记录", index_label="时间")

                ws = wb.sheets["历史记录"]
                ws.column_dimensions["A"].width = 6
                for col in range(1, ws.max_column + 1):
                    cell = ws[f"{get_column_letter(col)}1"]
                    cell.font = self.font(bold=True)
                    cell.alignment = ALIGN_CENTER

                for row in range(2, ws.max_row + 1):
                    for col in "A", "B", "C", "D":
                        cell = ws[f"{col}{row}"]
                        cell.font = self.font()
                        cell.border = BORDER
                        cell.alignment = ALIGN_CENTER
                        if isinstance(cell.value, (int, float)):
                            if cell.value < 0:
                                cell.value = -cell.value
                                cell.number_format = "-H:MM"
                            else:
                                cell.number_format = "H:MM"
                        else:
                            cell.number_format = "HH:MM"
                    cell = ws[f"E{row}"]
                    cell.font = self.font()
                    cell.border = BORDER
                    cell.alignment = ALIGN_CENTER

            if file[0] == ".":
                file = "程序所在文件夹" + file[1:]
            file = f"文件导出成功：{file}"
            self.update_log(file, "file")
            if self.FUNC_PATH_.get():
                messagebox.showinfo(TITLE, file, parent=history)

        def search(*args):
            now = datetime.now().timestamp()
            if now - search_time.get() <= 1:
                return 1
            search_time.set(now)
            fno_ = fno.get().strip().upper()
            fdate_ = datetime.fromisoformat(fdate.get())
            if not (len(fno_) > 2 and fno_.isalnum() and fno_[2].isnumeric()):
                return messagebox.showinfo(
                    TITLE, "请输入IATA航班号以精确匹配", parent=history
                )

            data = self.get_flight_data(
                fdate_,
                fdate_ + timedelta(1, minutes=-1),
                0,
                limit=1,
                sortOrder="descending",
                sortName="outSobt",
                flightNo=fno_,
                departureMode="outS",
                moni=False,
            )["航班_航班"].rename(columns={"outGuid": "guid"})
            if data.empty:
                return messagebox.showinfo(
                    TITLE, f"未在{fdate_.date()}查询到{fno_}", parent=history
                )
            index = data.loc[0, "guid"]
            inno, outno = data.loc[0, ["inFlightNo", "outFlightNo"]]
            if index in self.HISTORY["CTOT"].index:
                if fno_ in inno and fno_ not in outno:
                    return messagebox.showinfo(
                        TITLE,
                        f"{fno_}匹配到的航班为进港{inno}、离港{outno}，无法查询进港航班的历史记录",
                        parent=history,
                    )
            else:
                return messagebox.showinfo(
                    TITLE,
                    f"{fdate_.date()}的{fno_}暂无CTOT和COBT历史记录",
                    parent=history,
                )
            export.config(text="导出", state=tk.DISABLED)
            table.delete(*table.get_children())
            for i in info:
                i.config(text="-")
            cobt = (
                self.HISTORY["COBT"].loc[index].dropna()
                if index in self.HISTORY["COBT"].index
                else Series()
            )
            ctot = self.HISTORY["CTOT"].loc[index].dropna()
            last = data.loc[0, "outLastTot"]
            info[0].config(text=outno)
            info[1].config(text=data.loc[0, "outSobt"].strftime(r"%H:%M"))
            info[2].config(text=last.strftime(r"%H:%M"))
            info[3].config(text=ctot.nunique() - 1)

            result = concat([cobt, ctot], keys=["cobt", "ctot"], axis=1).assign(
                last=lambda x: x.ctot - last
            )
            result["status"] = result.index.map(
                lambda x: [v for k, v in self.STATUS.items() if x >= data.loc[0, k]][0]
            )

            info[5].config(
                text=(
                    self.min_sec_format(
                        result["last"].min().total_seconds(), wrap_hour=True
                    )
                )
            )
            info[4].config(
                text=(
                    self.min_sec_format(
                        result["last"].max().total_seconds(), wrap_hour=True
                    )
                )
            )

            formatted = result.reset_index()
            for i in "ctot", "cobt", "index":
                formatted[i] = formatted[i].map(lambda x: x.strftime(r"%H:%M"), IGNORE)
            formatted["last"] = formatted["last"].map(
                lambda x: "{}{}".format(
                    "+" if x > timedelta() else "", int(x.total_seconds() // 60)
                ),
                IGNORE,
            )
            for i in formatted.fillna("").itertuples(index=False):
                table.insert("", tk.END, values=tuple(i))

            renamer = {"cobt": "COBT", "ctot": "CTOT", "last": "推点", "status": "状态"}
            export.config(
                state=tk.NORMAL,
                command=lambda: (
                    export.config(text="已导出", state=tk.DISABLED)
                    if self.save_excel(
                        history,
                        "导出调时结果表格",
                        (
                            "{}/{}_{}".format(
                                self.FUNC_PATH.get(),
                                fno_,
                                self.FILENAME.get("CTOT和COBT历史记录"),
                            )
                            if self.FUNC_PATH_.get()
                            else filedialog.asksaveasfilename(
                                confirmoverwrite=True,
                                parent=history,
                                filetypes=(("Xlsx表格文件", "*.xlsx"),),
                                title="导出调时结果表格",
                                initialdir=self.FUNC_PATH.get(),
                                initialfile=f"{fno_}_{self.FILENAME.get('CTOT和COBT历史记录')}",
                            )
                        ),
                        save,
                        data=result.rename(columns=renamer),
                    )
                    else None
                ),
            )

        ttk.Label(upper, text="航班号").grid(row=0, column=0, padx=3, pady=5)
        fno = ttk.Entry(upper, textvariable=tk.StringVar(history), width=8)
        fno.grid(row=0, column=1, padx=2, pady=5, columnspan=3)
        fno.focus_set()
        fno.bind("<Return>", search)

        ttk.Label(upper, text="日期").grid(row=0, column=5, padx=5, pady=5)
        fdate = tk.StringVar(history, str(datetime.now())[:10])
        DateEntry(
            upper,
            textvariable=fdate,
            font="微软雅黑 10",
            width=10,
            background="lightgrey",
            locale="zh_CN",
            date_pattern="yyyy-MM-dd",
        ).grid(row=0, column=6, padx=2, pady=5)

        ttk.Button(upper, text="查询", command=search, width=5).grid(
            row=0, column=7, padx=2, pady=5
        )
        export = ttk.Button(upper, text="导出", width=5, state=tk.DISABLED)
        export.grid(row=0, column=8, padx=2, pady=5)

        info = []
        for i, j in enumerate(
            ("航班号", "STD", "最晚", "跳变次数", "最长推点", "最短推点")
        ):
            ttk.Label(mid, text=j, width=8, anchor="center").grid(
                row=0, column=i, padx=1, pady=2, sticky="ew"
            )
            info.append(ttk.Label(mid, text="-", width=7, anchor="center"))
            info[-1].grid(row=1, column=i, padx=1, pady=2, sticky="ew")

        columns = ("时间", "COBT", "CTOT", "推点", "状态")
        widths = (70, 70, 70, 70, 70)

        sby = ttk.Scrollbar(lower)
        table = ttk.Treeview(
            lower,
            height=history.winfo_screenheight() // 20,
            show="headings",
            columns=columns,
            yscrollcommand=sby.set,
            selectmode="browse",
        )
        for column, width in zip(columns, widths):
            table.column(
                column, width=width, minwidth=width, anchor="center", stretch=True
            )
            table.heading(column, text=column)
        sby.pack(side=tk.RIGHT, fill=tk.BOTH)
        sby.config(command=table.yview)

        width = sum(widths)
        table.pack(side=tk.LEFT, fill=tk.BOTH)
        upper.pack(side="top")
        mid.pack(side="top")
        lower.pack(side="bottom")

        history.resizable(False, True)
        history.minsize(width=372, height=150)
        history.geometry(
            f"372x300+{self.winfo_rootx() + self.winfo_width() // 3}+{self.winfo_rooty() + self.winfo_height() // 3}"
        )
        history.bind("<Escape>", lambda x: history.destroy())
        history.mainloop()

    def push(self):
        push = tk.Toplevel(self, name="push")
        push.title("企微信息推送")

        img = tk.StringVar(push, "")
        img_ = tk.BooleanVar(push)
        index = tk.StringVar(push)

        def set_img():
            if img_.get():
                if i := filedialog.askopenfilename(
                    filetypes=(
                        ("PNG图片", "*.png"),
                        ("JPG图片", "*.jpg"),
                        ("JPEG图片", "*.jpeg"),
                    ),
                    initialdir=self.INFO_PATH.get(),
                    initialfile=img.get(),
                    parent=push,
                    title="选择图片...",
                ):
                    img.set(i)
                else:
                    img_.set(False)
            img_button.config(state=state(img_.get()))

        def set_info(*args):
            message.delete("0.0", tk.END)
            item = self.PUSH.loc[index.get()]
            message.insert("0.0", item["信息"])
            img_.set(bool(item["图片"]))
            if img_.get():
                img.set(item["图片"])
            img_button.config(state=state(img_.get()))

        def update_push(*args):
            if self.PUSH.index.size and index.get() not in self.PUSH.index:
                index.set(self.PUSH.index[0])
            combobox.config(values=self.PUSH.index.to_list()[::-1])

        def push_info():
            image = img.get()
            text = message.get("0.0", tk.END).strip().strip("\n").strip()
            if "\n" in text:
                return messagebox.showinfo(
                    TITLE, "推送信息不得换行，请修改！", parent=push
                )
            if any([text, image]) and messagebox.askyesno(
                TITLE, "确定立刻推送信息？", parent=push
            ):
                output = {"text": text, "image": image}
                messagebox.showinfo(TITLE, f"测试推送参数：\n{output}", parent=push)

        def resize(*args):
            w = push.winfo_width()
            combobox.config(width=min(30, max(20, 20 + ((w - 400) // 15))))
            message.config(width=w // 9 + w // 100)

        push.bind("<Configure>", resize)
        frames = [ttk.Frame(push) for _ in range(2)]
        combobox = ttk.Combobox(
            frames[0],
            width=20,
            values=self.PUSH.index.to_list()[::-1],
            textvariable=index,
            state="readonly",
        )
        combobox.grid(row=0, column=0, padx=5, pady=5)
        combobox.bind("<Enter>", update_push)
        combobox.bind("<<ComboboxSelected>>", set_info)

        ttk.Checkbutton(
            frames[0], text="附带图片", variable=img_, command=set_img, **ONOFFS
        ).grid(row=0, column=1, padx=5, pady=5)
        img_button = ttk.Button(
            frames[0],
            text="查看图片",
            command=lambda: Image.open(img.get()).show("查看图片"),
            state=state(img_.get()),
            width=9,
        )
        img_button.grid(row=0, column=2, padx=2, pady=5)

        ttk.Button(frames[0], text="推送!", command=push_info, width=6).grid(
            row=0, column=3, padx=5, pady=5
        )

        message = tk.Text(
            frames[1],
            name="message",
            font=("微软雅黑", 10),
            wrap="char",
            height=push.winfo_screenheight() // 20,
            undo=True,
        )
        message.pack(padx=0, pady=0, expand=True, fill=tk.BOTH, side="left")
        msgscroll = ttk.Scrollbar(
            frames[1],
            name="msgscroll",
            orient="vertical",
            command=message.yview,
        )
        message.config(yscrollcommand=msgscroll.set)
        msgscroll.pack(padx=0, pady=0, expand=True, fill=tk.BOTH)

        frames[0].pack(fill="none", anchor="center")
        frames[1].pack(fill="both", expand=True)

        popup = tk.Menu(message, tearoff=False)

        def cut():
            try:
                copy(), message.delete(tk.SEL_FIRST, tk.SEL_LAST)
            except Exception:
                ...

        def copy():
            try:
                (
                    message.clipboard_clear(),
                    message.clipboard_append(message.get(tk.SEL_FIRST, tk.SEL_LAST)),
                )
            except Exception:
                ...

        def paste():
            try:
                message.insert(tk.INSERT, message.selection_get(selection="CLIPBOARD"))
            except Exception:
                ...

        def selall():
            message.tag_add("sel", "0.0", tk.END)
            return "break"

        def delete():
            try:
                message.delete(tk.SEL_FIRST, tk.SEL_LAST)
            except Exception:
                ...

        def clear():
            try:
                message.delete("0.0", tk.END)
            except Exception:
                ...

        popup.add_command(label="删除", command=delete)
        popup.add_command(label="剪切", command=cut)
        popup.add_command(label="复制", command=copy)
        popup.add_command(label="粘贴", command=paste)
        popup.add_separator()
        popup.add_command(label="全选", command=selall)
        popup.add_command(label="清空", command=clear)

        message.bind("<Button-3>", lambda event: popup.post(event.x_root, event.y_root))
        if self.PUSH.index.size:
            index.set(self.PUSH.index[-1])
            set_info()

        push.minsize(width=400, height=150)
        push.geometry(
            f"600x300+{self.winfo_rootx() + self.winfo_width() // 10}+{self.winfo_rooty() + self.winfo_height() // 8}"
        )
        push.mainloop()

    def post_push(self, data: dict):
        pass

    def auto_delay(self):
        autodelay = tk.Toplevel(self, name="autodelay")
        autodelay.attributes("-topmost", self.TOPMOST.get())
        autodelay.title("自动延误判定")
        autodelay.resizable(False, False)
        autodelay.bind("<Escape>", lambda *x: autodelay.destroy())
        autodelay.geometry(
            f"+{self.winfo_rootx() + self.winfo_width() // 5}+{self.winfo_rooty() + self.winfo_height() // 10}"
        )
        if self.winfo_viewable():
            autodelay.transient(self)
        autodelay.grab_set()
        _setup_dialog(autodelay)

        ttk.Label(
            autodelay,
            text="；\n".join(
                (
                    "每次更新数据时自动判定满足条件、已放行延误、未判定延误原因的航班延误原因，判定优先级由前往后递减",
                    "不满足条件则不自动判定延误原因，注意在A-CDM延误判定页核实判定结果",
                    "前两项选择后将无视其余条件，用于本场特殊天气和军事活动导致的延误情景",
                )
            ),
            font=("微软雅黑", 8),
            foreground="dimgrey",
            wraplength=390,
        ).grid(row=0, column=0, columnspan=3, padx=12, pady=8)

        cbs = dict()
        keys = any(list(self.DELAYBY.values())[:2])
        for i, (k, v) in enumerate(self.DELAYBY.items()):
            v = tk.BooleanVar(autodelay, v)
            cbs[k] = (
                v,
                ttk.Checkbutton(
                    autodelay,
                    text=k,
                    variable=v,
                    **ONOFFS,
                    state=state(not (keys and i >= 2)),
                ),
            )
            cbs[k][1].grid(
                row=i + 2, column=0, columnspan=3, padx=10, pady=2, sticky="w"
            )
        keys = list(self.DELAYBY.keys())
        cbs["均判定为本场天气"][1].config(
            command=lambda: (
                cbs["均判定为本场军事活动"][0].set(False),
                [
                    cbs[k][1].config(state=state(not cbs["均判定为本场天气"][0].get()))
                    for k in keys[2:]
                ],
            )
        )
        cbs["均判定为本场军事活动"][1].config(
            command=lambda: (
                cbs["均判定为本场天气"][0].set(False),
                [
                    cbs[k][1].config(
                        state=state(not cbs["均判定为本场军事活动"][0].get())
                    )
                    for k in keys[2:]
                ],
            )
        )
        cbs[keys[-1]][1].config(
            command=lambda: [cbs[k][0].set(False) for k in keys[-4:] if k != keys[-1]]
        )
        cbs[keys[-2]][1].config(
            command=lambda: [cbs[k][0].set(False) for k in keys[-4:] if k != keys[-2]]
        )
        cbs[keys[-3]][1].config(
            command=lambda: [cbs[k][0].set(False) for k in keys[-4:] if k != keys[-3]]
        )
        cbs[keys[-4]][1].config(
            command=lambda: [cbs[k][0].set(False) for k in keys[-4:] if k != keys[-4]]
        )

        def confirm():
            for k, v in cbs.items():
                self.DELAYBY[k] = v[0].get()
            self.AUTODELAY.set(any([i for i in self.DELAYBY.values()]))
            autodelay.destroy()
            self.update_status()

        ttk.Button(autodelay, text="确定", command=confirm, width=15).grid(
            row=i + 3, column=0, columnspan=3, sticky="ws", padx=10, pady=5
        )
        ttk.Button(autodelay, text="取消", command=autodelay.destroy, width=15).grid(
            row=i + 3, column=0, columnspan=3, sticky="es", padx=10, pady=5
        )
        autodelay.focus_set()
        autodelay.wait_window()
        return 0

    def update_delay(self, data: DataFrame):
        try:
            target = (
                data.loc[data["rstDelayReason"].isna()]
                .loc[
                    data["outAtot"].fillna(self.datetime_now()) - data["outLastTot"]
                    > timedelta()
                ]
                .copy()
            )
            target.drop(
                target.loc[target["inAldt"].isna() & target["inSldt"].notna()]
                .loc[(target["outSobt"] - target["inSibt"]) > timedelta()]
                .index,
                inplace=True,
            )
            if not target.__len__():
                return data

            delay_reasons = [f"{i}DelayReason" for i in ("rst", "add", "pri", "sub")]
            delay_type = lambda x: self.delay_type[x] + [
                self.PRIMARY[self.delay_type[x][0][:2]],
                self.delay_type[x][0][5:],
            ]
            if self.DELAYBY["均判定为本场天气"]:
                target[delay_reasons] = delay_type(0)
            elif self.DELAYBY["均判定为本场军事活动"]:
                target[delay_reasons] = delay_type(1)
            else:
                if self.DELAYBY["过站时间严重不足（前序STA晚于后序STD）判为公司计划"]:
                    target.loc[target["outSobt"] <= target["inSibt"], delay_reasons] = (
                        delay_type(4)
                    )
                if self.DELAYBY["根据流控类型判定为外站天气或军事活动"]:
                    for i in (
                        target.loc[target["rstDelayReason"].isna()]
                        .loc[target["outTmi"].notna()]
                        .index
                    ):
                        if "天气" in target.loc[i, "outTmi"]:
                            target.loc[i, delay_reasons] = delay_type(2)
                        elif "其他空域用户" in target.loc[i, "outTmi"]:
                            target.loc[i, delay_reasons] = delay_type(3)
                if self.DELAYBY[
                    "过站时间不足（计划过站时间小于最短过站时间）判为公司计划"
                ]:
                    for i in (
                        target.loc[target["rstDelayReason"].isna()]
                        .loc[target["ttt"] < timedelta()]
                        .index
                    ):
                        target.loc[i, delay_reasons] = delay_type(4)
                for i in enumerate(
                    (
                        "剩余延误原因判定为本场天气",
                        "剩余延误原因判定为本场军事活动",
                        "剩余延误原因判定为外站天气",
                        "剩余延误原因判定为外站军事活动",
                    )
                ):
                    if self.DELAYBY[i[1]]:
                        target.loc[target["rstDelayReason"].isna(), delay_reasons] = (
                            delay_type(i[0])
                        )
                        break

            updates = []
            for i in target.loc[target["rstDelayReason"].notna()].index:
                try:
                    updates.append(
                        {
                            "flightNo": target.loc[i, "outFlightNo"],
                            "guid": int(target.loc[i, "delayGuid"]),
                            "rstDelayReason": target.loc[i, "rstDelayReason"],
                        }
                    )
                    if target.loc[i, "addDelayReason"]:
                        updates[-1]["addDelayReason"] = target.loc[i, "addDelayReason"]
                except (TypeError, ValueError):
                    updates.append(
                        {
                            "flightNo": target.loc[i, "outFlightNo"],
                            "flightDate": str(target.loc[i, "outFlightDate"])[:19],
                            "sobt": str(target.loc[i, "outSobt"])[:19],
                            "rstDelayReason": target.loc[i, "rstDelayReason"],
                        }
                    )
                    if target.loc[i, "addDelayReason"]:
                        updates[-1]["addDelayReason"] = target.loc[i, "addDelayReason"]
                data.loc[i, delay_reasons] = target.loc[i, delay_reasons]

            if updates:
                threading.Thread(target=self.update_delay_data, args=(updates,)).start()
        except Exception as exception:
            tb, exception = format_exc().split("\n", 1)[1], repr(exception)
            self.update_log(
                f"自动延误判定失败 ({exception[: exception.find('(')]})\n{tb}", "warn"
            )
        return data

    def update_delay_data(self, payloads: Iterable[dict], name: str = TITLE):
        now = str(datetime.now())[:19]
        # returns = []
        for payload in payloads:
            if "flightDate" in payload:
                flight = payload.get("flightNo", "")
                payload["insertPer"] = payload["rstModiPer"] = name
                payload["insertTime"] = now
                response = self.get_session().post(
                    self.URLS.get("延误保存"),
                    data=dumps(payload),
                    headers=self.HEADER,
                    timeout=10,
                )
            else:
                flight = payload.pop("flightNo", "")
                payload["updatePer"] = payload["rstModiPer"] = name
                payload["updateTime"] = now
                response = self.get_session().post(
                    self.URLS.get("延误更新"),
                    data=dumps(payload),
                    headers=self.HEADER,
                    timeout=10,
                )
            if response.status_code != 200:
                msg = f"{flight}延误判定上传失败 ({response.json().get('message')})"
                self.update_log(msg, "warn")
                # returns.append(msg)
        # return "；".join(returns)


if __name__ == "__main__":
    hwnd = []
    autogui = win32gui.EnumWindows(
        lambda i, j: (
            j.append(i)
            if win32gui.GetClassName(i) == "TkTopLevel"
            and TITLE in win32gui.GetWindowText(i)
            else None
        ),
        hwnd,
    )
    if hwnd:
        win32gui.ShowWindow(hwnd[0], 5)
        win32gui.BringWindowToTop(hwnd[0])
        messagebox.showinfo("提示", f"{TITLE}已在运行中！")
    else:
        kwargs = {
            "LOGIN": {
                "shi.xx": "4c/3qMk4KGLDHb43aI+BUQ==",
            }
        }
        AutoGui(**kwargs)
