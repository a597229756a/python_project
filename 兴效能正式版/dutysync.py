import os
import base64
import pystray
import win32gui
import threading
import tkinter as tk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from tkinter import ttk, messagebox, filedialog
from tkinter.simpledialog import _setup_dialog
from json import load, dump, JSONDecodeError
from datetime import datetime, timedelta
from traceback import format_exc, format_tb
from dutyicon import icon as dutyicon
from PIL import Image

VERSION = "1.0.2"
TITLE = "值班表同步"
CACHE = os.getenv("LOCALAPPDATA")
CACHE = rf"{CACHE if CACHE else '.'}\pkxds"
LOG = rf"{CACHE}\dutysync.log"
SETTINGS = {
    "path": "",
    "sync": r"\\10.154.61.233\运行管理部\兴效能",
    "sheet": 1,
    "row": 4,
    "column": 2,
    "columns": {
        "dateField_l89vg9u8": "日期",
        "textField_l89uhu9e": "星期",
        "textField_l89uhu9f": "机场总值班",
        "textField_l89uhu9g": "运行总值班 白班",
        "textField_l89uhu9h": "运行总值班 夜班",
        "textField_l89uhu9i": "运行总值班 全天",
        "textField_l89uhu9j": "飞行区管理部",
        "textField_l89uhu9k": "航站楼管理部",
        "textField_l89uhu9l": "公共区管理部",
        "textField_l89uhu9n": "信息管理部",
        "textField_l89uhu9o": "消防管理部",
        "textField_l89uhu9s": "服务品质部",
        "textField_l89uhu9m": "采购工程部",
        "textField_l89uhu9p": "党群工作部",
        "textField_l89uhu9q": "安全质量部",
        "textField_l89uhu9r": "航空安保管理部",
    },
}
SETTINGS_ = SETTINGS.copy()


def save_settings(index: int = 0):
    with open(rf"{CACHE}\dutysync{index}.json", "w") as settings:
        dump(SETTINGS, settings)


def load_settings(index: int = 0):
    json = rf"{CACHE}\dutysync{index}.json"
    if os.path.exists(json):
        try:
            with open(json) as settings:
                settings = load(settings)
            assert isinstance(settings, dict)
            assert isinstance(settings.get("columns", dict()), dict)
            SETTINGS.update(settings)
            return 0
        except (JSONDecodeError, AssertionError):
            os.remove(json)
    save_settings(index)
    return 1


def load_excel(file: str) -> dict[datetime, list[str]]:
    wb = openpyxl.load_workbook(file, True)
    try:
        ws = wb.worksheets[SETTINGS.get("sheet", 1) - 1]
        output = dict()
        for row in ws.iter_rows(
            SETTINGS.get("row", 2),
            ws.max_row,
            SETTINGS.get("column", 2),
            SETTINGS.get("column", 2) + len(SETTINGS.get("columns", [0, 0])) - 1,
        ):
            if row[0].value:
                output[from_excel(row[0].value)] = [
                    str(cell.value).split("：")[-1].replace(" ", "").replace("\u00a0", "") for cell in row[1:]
                ]
        return output
    finally:
        wb.close()


def attribute_editor(
    root: tk.Misc, attr: str, title: str = "参数编辑", tip: str = "", **kwargs
) -> int:
    """
    To edit list or dict `SETTINGS` setted in `root`.

    Parameters
    --
    attr: `str`, attribute name setted in root, <RESET> button is enabled if found in `SETTINGS`.
    title: `str`, editor window name.
    tip: `str`, editor tip showed at lower center.

    Kwargs
    --
    master: `tk.Misc`, default `root`.
    headers: `tuple[str]`, set column text[key, value], example: `('My Keys', 'My Values')`.
    width: `tuple[int]`, set column width[key, value] in pixels, example: `(150, 250)`.
    anchor: `tuple[str]`, set column anchor[key, value], example: `('center', 'w')`.

    key: `Callable`, a function or lambda to check key entry validity. Dict key edit is locked if not set.
    value: `Callable | dict[str, Callable]`, a function or lambda to check value entry validity.
    To check different value of key with a specific method, use dict[str, Callable].

    key_convert: `Callable`, convert key entry `str` to desired `object` upon saving.
    value_convert: `Callable | dict[str, Callable]`, convert value entry `str` to desired `object` upon saving.
    To convert different value of key with a specific method, use dict[str, Callable].
    """

    true = lambda *x: True
    state = lambda x: tk.NORMAL if x else tk.DISABLED
    var = SETTINGS[attr]
    isdict = isinstance(var, dict)
    editable = (isdict and kwargs.get("key")) or not isdict
    anchor = kwargs.pop("anchor", ("center", "w"))
    width = kwargs.pop("width", (100, 450) if isdict else (20, 530))
    headers = {
        "#1": ("键名", {"anchor": anchor[0], "minwidth": 20, "width": width[0]}),
        "#2": ("键值", {"anchor": anchor[1], "minwidth": 20, "width": width[1]}),
    }
    width = sum(width)
    header = kwargs.pop("headers", (headers["#1"][0], headers["#2"][0]))
    col_to_header = {"键名": header[0], "键值": header[1]}
    master = kwargs.pop("master", root)

    editor = tk.Toplevel(master, name="editor")
    editor.title(title)
    editor.resizable(False, False)
    editor.geometry(
        f"+{master.winfo_rootx()+master.winfo_width()//4}+{master.winfo_rooty()+master.winfo_height()//4}"
    )
    if master.winfo_viewable():
        editor.transient(master)
    editor.grab_set()
    _setup_dialog(editor)

    def get_edit(event):
        row = table.identify_row(event.y)
        col = table.identify_column(event.x)
        if row and col:
            if col == "#1" and not editable:
                return 1
            else:
                col = headers[col][0]
            edit(row, col, **kwargs)
        else:
            new_row()

    def edit(row, col, new: bool = False, **kwargs):
        editing = tk.Toplevel(editor, name="editing")
        editing.title(f"{col_to_header[col]}{'新建' if new else '编辑'}")
        editing.resizable(False, False)
        editing.geometry(
            f"+{editor.winfo_rootx()+editor.winfo_width()//6}+{editor.winfo_rooty()+editor.winfo_height()//6}"
        )
        if editor.winfo_viewable():
            editing.transient(editor)

        for i in newb, delb, upb, downb:
            i.config(state=tk.DISABLED)
        table.config(selectmode="none")
        table.unbind("<<TreeviewSelect>>")
        table.unbind("<Escape>")

        item = table.item(row, "values")
        if "值" in col:
            check = kwargs.get("value", true)
            if isinstance(check, dict):
                check = check.get(item[0], true)
            item = item[1]
        else:
            item, check = item[0], kwargs.get("key")
        scrollbar_e = ttk.Scrollbar(editing)
        entry = tk.Text(
            editing,
            width=width // 12,
            height=5,
            undo=True,
            wrap="char",
            font=("微软雅黑", 10),
            yscrollcommand=scrollbar_e.set,
        )
        scrollbar_e.config(command=entry.yview)
        entry.insert("0.0", item)

        popup = tk.Menu(editing, tearoff=False)

        def cut():
            try:
                copy(), entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
            except Exception:
                ...

        def copy():
            try:
                entry.clipboard_clear(), entry.clipboard_append(
                    entry.get(tk.SEL_FIRST, tk.SEL_LAST)
                )
            except Exception:
                ...

        def paste():
            try:
                entry.insert(tk.INSERT, entry.selection_get(selection="CLIPBOARD"))
            except Exception:
                ...

        def selall():
            entry.tag_add("sel", "0.0", tk.END)
            return "break"

        def delete():
            try:
                entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
            except Exception:
                ...

        popup.add_command(label="删除", command=delete)
        popup.add_command(label="剪切", command=cut)
        popup.add_command(label="复制", command=copy)
        popup.add_command(label="粘贴", command=paste)
        popup.add_separator()
        popup.add_command(label="全选", command=selall)

        entry.bind("<Button-3>", lambda event: popup.post(event.x_root, event.y_root))
        entry.grid(row=0, column=0, columnspan=2)
        scrollbar_e.grid(row=0, column=3, sticky="ns")

        def confirm(*args):
            v = entry.get("0.0", tk.END).strip().strip("\n").strip()
            if check(v):
                table.set(row, col, v)
                editing.destroy()
                table.see(row)
            else:
                messagebox.showinfo(
                    TITLE, f"输入{col}无效或无法匹配，请参考提示", parent=editing
                )

        ttk.Button(editing, text="确认", width=width // 24 - 1, command=confirm).grid(
            sticky="ws", row=2, column=0, columnspan=4, padx=5, pady=2
        )
        ttk.Button(
            editing, text="取消", width=width // 24 - 1, command=editing.destroy
        ).grid(sticky="es", row=2, column=0, columnspan=4, padx=5, pady=2)
        editing.grab_set()
        _setup_dialog(editing)
        editing.wait_window()
        if editor.winfo_exists():
            for i in newb, delb, upb, downb:
                i.config(state=state(editable))
            table.config(selectmode="extended")
            table.bind("<<TreeviewSelect>>", button_update)
            table.bind("<Escape>", lambda x: table.selection_set([]))
            table.focus_set()
            table.selection_set([row] if row in table.children else [])
            if table.item(row)["values"][0] == "" and new:
                table.delete(row)
        return 0

    scrollbar = ttk.Scrollbar(editor)
    table = ttk.Treeview(
        editor,
        show="headings",
        columns=list(i[0] for i in headers.values()),
        yscrollcommand=scrollbar.set,
    )
    for k, v in headers.values():
        table.column(k, **v)
        table.heading(k, text=col_to_header[k])
    scrollbar.config(command=table.yview)
    table.grid(row=0, column=0, columnspan=14, sticky="nsew")
    scrollbar.grid(row=0, column=14, sticky="ns")

    for k, v in var.items() if isdict else enumerate(var):
        table.insert("", tk.END, values=(k, str(v)))
    table.bind("<Double-1>", get_edit)

    def new_row():
        if isdict:
            row, col = table.insert("", tk.END), "键名"
            table.set(row, 0, "")
        else:
            row, col = table.insert("", tk.END), "键值"
            table.set(row, 0, int(table.item(table.prev(row), "values")[0]) + 1)
        table.set(row, 1, "")
        edit(row, col, True, **kwargs)

    def del_row():
        if messagebox.askyesno(TITLE, "确定删除？", parent=editor):
            for i in table.selection():
                table.delete(i)

    def move_up():
        sel = [table.selection(), []]
        if table.get_children()[0] in sel[0]:
            return 0
        for i in sel[0]:
            j = table.prev(i)
            k = table.item(j, "values")
            for v in enumerate(table.item(i, "values")):
                if not isdict and v[0] == 0:
                    continue
                table.set(j, *v)
            for v in enumerate(k):
                if not isdict and v[0] == 0:
                    continue
                table.set(i, *v)
            sel[1].append(j)
        table.selection_set(sel[1])

    def move_down():
        sel = [table.selection()[::-1], []]
        if table.get_children()[-1] in sel[0]:
            return 0
        for i in sel[0]:
            j = table.next(i)
            k = table.item(j, "values")
            for v in enumerate(table.item(i, "values")):
                if not isdict and v[0] == 0:
                    continue
                table.set(j, *v)
            for v in enumerate(k):
                if not isdict and v[0] == 0:
                    continue
                table.set(i, *v)
            sel[1].append(j)
        table.selection_set(sel[1])

    def button_update(*args):
        flag = table.selection() and (not isdict or editable)
        for i in delb, upb, downb:
            i.config(state=state(flag))
        newb.config(state=state(editable))

    button_style = (
        ("EditorIcon.TButton", "↑", "↓", width // 2 - 10, "－", "＋")
        if width <= 400
        else (
            "Editor.TButton",
            "↑ 上移",
            "下移 ↓",
            (width - 80) // 2,
            "－删除",
            "新建＋",
        )
    )
    upb = ttk.Button(
        editor,
        text=button_style[1],
        style=button_style[0],
        command=move_up,
        state=tk.DISABLED,
    )
    upb.grid(sticky="ws", row=1, column=0, columnspan=2, padx=2, pady=5)
    downb = ttk.Button(
        editor,
        text=button_style[2],
        style=button_style[0],
        command=move_down,
        state=tk.DISABLED,
    )
    downb.grid(sticky="ws", row=1, column=2, columnspan=2, padx=2, pady=5)
    ttk.Label(
        editor,
        text=tip,
        font=("微软雅黑", 8),
        foreground="dimgrey",
        wraplength=button_style[3],
    ).grid(row=1, column=3, columnspan=8, padx=2, pady=2)
    delb = ttk.Button(
        editor,
        text=button_style[4],
        style=button_style[0],
        command=del_row,
        state=tk.DISABLED,
    )
    delb.grid(sticky="es", row=1, column=0, columnspan=12, padx=2, pady=5)
    newb = ttk.Button(
        editor,
        text=button_style[5],
        style=button_style[0],
        command=new_row,
        state=state(editable),
    )
    newb.grid(sticky="es", row=1, column=0, columnspan=14, padx=2, pady=5)
    table.bind("<<TreeviewSelect>>", button_update)
    table.bind("<Escape>", lambda x: table.selection_set([]))

    def confirm():
        if isdict:
            k, v = kwargs.get("key_convert", str), kwargs.get("value_convert", str)
            update = dict()
            for i in table.get_children():
                value = table.item(i, "values")
                if any(value):
                    if k(value[0]) in update:
                        messagebox.showinfo(
                            TITLE, "存在相同的键名，请确保键名唯一", parent=editor
                        )
                        try:
                            table.focus_set()
                            table.selection_set(i)
                        except Exception:
                            ...
                        return 1
                    update[k(value[0])] = (
                        v.get(k(value[0]), str)(value[1])
                        if isinstance(v, dict)
                        else v(value[1])
                    )
        else:
            update = list()
            for i in table.get_children():
                value = table.item(i, "values")[1]
                update.append(kwargs.get("value_convert", str)(value))
        SETTINGS[attr] = update
        editor.destroy()

    def reset():
        if messagebox.askyesno(TITLE, "确认恢复默认？", parent=editor):
            for i in table.get_children():
                table.delete(i)
            for k, v in (
                SETTINGS_.get(attr).items()
                if isdict
                else enumerate(SETTINGS_.get(attr))
            ):
                table.insert("", tk.END, values=(k, str(v)))
            table.focus_set()

    ttk.Button(editor, text="保存", width=width // 25, command=confirm).grid(
        sticky="w", row=2, column=0, padx=5, pady=5, columnspan=15
    )
    ttk.Button(
        editor,
        text="恢复默认",
        width=width // 25,
        command=reset,
        state=state(attr in SETTINGS),
    ).grid(row=2, column=0, padx=5, pady=5, columnspan=15)
    ttk.Button(editor, text="取消", width=width // 25, command=editor.destroy).grid(
        sticky="e", row=2, column=0, padx=5, pady=5, columnspan=15
    )
    editor.wait_window()


def main(index: int = 0):
    global TITLE
    TITLE += f" - 进程{index}"
    if not os.path.exists(CACHE):
        os.mkdir(CACHE)

    root = tk.Tk()
    root.title(TITLE)
    if load_settings(index):
        root.after_idle(messagebox.showinfo, TITLE, "应用初次运行或设置加载失败，使用默认设置！")

    path = tk.StringVar(root, SETTINGS.get("path", ""))
    sheet = tk.IntVar(root, SETTINGS.get("sheet", 1))
    row = tk.IntVar(root, SETTINGS.get("row", 4))
    column = tk.IntVar(root, SETTINGS.get("column", 2))
    def check(file: str):
        try:
            return load_excel(file)
        except Exception as exception:
            tb, exception = format_exc().split("\n", 1)[1], repr(exception)
            messagebox.showerror(
                TITLE,
                f"表格无法正确读取，请检查参数设置和表格！\n错误类型：{exception[: exception.find('(')]}\n{tb}",
            )
            return {}

    def ask_file():
        file = SETTINGS.get("path", "")
        if update_spinbox():
            return file
        if s := filedialog.askopenfilename(
            initialfile=file,
            filetypes=(("Xlsx表格", "*.xlsx"),),
            parent=root,
            title="选择表格文件",
        ):
            result = check(s)
            if result and messagebox.askyesno(
                TITLE,
                "读取结果如下，是否确认？\n{}".format(
                    "\n".join(
                        [
                            "，".join([str(k)[:10]] + [v for v in v][:4])
                            + ("..." if len(v) > 3 else "")
                            for k, v in list(result.items())[:3]
                        ]
                    )
                ),
            ):
                file = s
                SETTINGS["last"] = 0
        path.set(file)
        SETTINGS["path"] = file
        save_settings(index)
        return file

    ttk.Button(root, text="选择值班表格...", command=ask_file).grid(
        row=0, column=0, columnspan=3, padx=5, pady=5, sticky="we"
    )
    ttk.Label(
        root,
        textvariable=path,
        font=("微软雅黑", 8),
        foreground="dimgrey",
        wraplength=290,
    ).grid(row=1, column=0, columnspan=6, padx=8, pady=2, sticky="w")

    ttk.Button(
        root,
        text="修改列名...",
        command=lambda: (
            attribute_editor(
                root,
                "columns",
                f"编辑{get_column_letter(column.get())}列起数据",
                "双击编辑",
                key=str,
                headers=("宜搭表单空格唯一标识", "列名"),
                width=(150, 150),
                anchor=("center", "center"),
            )
        ),
    ).grid(row=0, column=3, columnspan=3, padx=5, pady=5, sticky="we")

    def update_spinbox():
        for name, key, var in zip(
            ("表单页", "首日行", "日期列"),
            ("sheet", "row", "column"),
            (sheet, row, column),
        ):
            try:
                value = int(var.get())
            except Exception:
                messagebox.showerror(TITLE, f"请检查{name}参数！")
                return 1
            var.set(min(max(value, 1), 999))
            SETTINGS[key] = value

    ttk.Label(root, text="表单页").grid(row=2, column=0, padx=5, pady=2)
    sheetbox = ttk.Spinbox(
        root, textvariable=sheet, increment=1, width=3, from_=1, to=999
    )
    sheetbox.grid(row=2, column=1, padx=5, pady=2)
    ttk.Label(root, text="首日行").grid(row=2, column=2, padx=5, pady=2)
    rowbox = ttk.Spinbox(root, textvariable=row, increment=1, width=3, from_=1, to=999)
    rowbox.grid(row=2, column=3, padx=5, pady=2)
    ttk.Label(root, text="日期列").grid(row=2, column=4, padx=5, pady=2)
    columnbox = ttk.Spinbox(
        root, textvariable=column, increment=1, width=3, from_=1, to=999
    )
    columnbox.grid(row=2, column=5, padx=5, pady=2)

    schedule = tk.StringVar(root)
    running = tk.BooleanVar(root)

    def update():
        try:
            now = datetime.now()
            file = SETTINGS.get("path")
            last = os.path.getmtime(file)

            updated = []
            for k, v in load_excel(file).items():
                updated.append(
                    dict(
                        [
                            (i, j)
                            for i, j in zip(
                                SETTINGS.get("columns", {}).keys(),
                                [int(k.timestamp() * 1000)] + v,
                            )
                        ]
                    )
                )

            with open(rf"{SETTINGS.get('sync', '.')}\duty.json", "w") as json:
                dump(updated, json)
            SETTINGS["last"] = last
            with open(LOG, "a", encoding="UTF-8") as log:
                log.write("{} 同步成功\n".format(str(now)[:19]))
            status.set("同步成功：{}".format(now.strftime(r"%H:%M:%S")))
            schedule.set(root.after(1000, run))
        except Exception as exception:
            tb, exception = format_exc().split("\n", 1)[1], repr(exception)
            status.set(
                "同步失败：{} ({})".format(
                    now.strftime(r"%H:%M:%S"), exception[: exception.find("(")]
                )
            )
            with open(LOG, "a", encoding="UTF-8") as log:
                log.write("{} {}\n".format(str(now)[:19], tb))
            schedule.set(root.after(10000, run))

    def run():
        if i := schedule.get():
            root.after_cancel(i)
        mtime, now = os.path.getmtime(SETTINGS.get("path")), datetime.now().timestamp()
        if mtime == SETTINGS.get("last", 0) or now - mtime < 4:
            schedule.set(root.after(5000, run))
        else:
            update()
            save_settings(index)
            if os.path.exists(LOG):
                days = datetime.now() - timedelta(1)
                with open(LOG, "r", encoding="UTF-8") as log:
                    lines = log.readlines()
                for i, line in enumerate(lines):
                    try:
                        if datetime.fromisoformat(line[:19]) >= days:
                            break
                    except Exception:
                        continue
                if i > 0:
                    try:
                        with open(LOG, "w", encoding="UTF-8") as log:
                            log.writelines(lines[i:])
                    except PermissionError:
                        ...

    def launch():
        if i := SETTINGS.get("path"):
            update_spinbox()
            if i := check(i):
                sync_button.config(text="停 止 同 步", command=terminate)
                for i in sheetbox, rowbox, columnbox:
                    i.config(state=tk.DISABLED)
                save_settings(index)
                threading.Thread(target=run).start()
                status.set("同步开始：{} ".format(datetime.now().strftime(r"%H:%M:%S")))
                running.set(True)
                icon.update_menu()
        else:
            messagebox.showinfo(TITLE, "请选择表格文件并配置参数以开始同步！")

    def terminate():
        sync_button.config(text="开 始 同 步", command=launch)
        for i in sheetbox, rowbox, columnbox:
            i.config(state=tk.NORMAL)
        if i := schedule.get():
            root.after_cancel(i)
        status.set("同步停止：{} ".format(datetime.now().strftime(r"%H:%M:%S")))
        running.set(False)
        icon.update_menu()
        save_settings(index)

    style = ttk.Style()
    style.configure("Sync.TButton", font=("微软雅黑", 12, "bold"))
    style.configure("Editor.TButton", font=("微软雅黑", 8), width=8)
    style.configure("EditorIcon.TButton", font=("微软雅黑", 9), width=3)

    def ask_sync():
        if s := filedialog.askdirectory(
            initialdir=SETTINGS.get("sync", "."),
            mustexist=True,
            parent=root,
            title="选择同步文件夹",
        ):
            SETTINGS["sync"] = s
        save_settings(index)
        return SETTINGS.get("sync", ".")

    sync_button = ttk.Button(
        root, style="Sync.TButton", text="开 始 同 步", command=launch
    )
    sync_button.grid(row=3, column=0, columnspan=4, padx=5, pady=5, sticky="we")
    ttk.Button(root, text="同步至...", command=ask_sync).grid(
        row=3, column=4, columnspan=2, padx=5, pady=5, sticky="nswe"
    )

    status = tk.StringVar(root, "初始化成功")
    status_bar = ttk.Label(
        root, textvariable=status, background="white", wraplength=290
    )
    status_bar.bind("<Double-1>", lambda *args: status.set("状态栏已重置"))
    status_bar.grid(row=4, column=0, columnspan=6, sticky="nswe")

    def handle_exception(*args):
        if isinstance(args[0], threading.ExceptHookArgs):
            args = args[0][:3]
        status.set("出现异常：{} ({})".format(args[1], args[0].__name__))
        with open(LOG, "a", encoding="UTF-8") as log:
            log.write(
                "{} {}\n".format(
                    str(datetime.now())[:19], "".join(format_tb(args[2]))[:-1]
                )
            )

    def quit_window(icon, item):
        icon.stop()
        if i := schedule.get():
            root.after_cancel(i)
        save_settings(index)
        root.destroy()

    icon = open("dutyicon.ico", "wb+")
    icon.write(base64.b64decode(dutyicon))
    icon.close()
    image = Image.open("dutyicon.ico").copy()
    root.iconbitmap("dutyicon.ico", "dutyicon.ico")
    os.remove("dutyicon.ico")

    launch_icon = pystray.MenuItem("开始同步", lambda *x: launch(), enabled=lambda x: not running.get())
    terminate_icon = pystray.MenuItem("停止同步", lambda *x: terminate(), enabled=lambda x: running.get())
    menu = (
        launch_icon,
        terminate_icon,
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("显示面板", lambda *x: root.after(100, root.deiconify), default=True, visible=False),
        pystray.MenuItem("退出", quit_window),
    )
    icon = pystray.Icon(TITLE, image, TITLE, menu)
    threading.Thread(target=icon.run, daemon=True).start()

    root.report_callback_exception = handle_exception
    root.resizable(0, 0)
    w, h = 300, 130
    root.geometry(
        f"+{(root.winfo_screenwidth()-w)//2}+{(root.winfo_screenheight()-h)//2}"
    )
    root.wm_protocol("WM_DELETE_WINDOW", root.withdraw)

    root.mainloop()


if __name__ == "__main__":
    def detect_windows(window, outputs: list):
        title = win32gui.GetWindowText(window)
        if win32gui.GetClassName(window) == "TkTopLevel" and TITLE in title:
            try:
                i = int(title[title.find("进程") + 2 :])
            except:
                i = 1
            outputs.append(i)
    hwnd = []
    win32gui.EnumWindows(detect_windows, hwnd)
    i = 1
    for _ in hwnd:
        if i in hwnd:
            i += 1
    main(i)
